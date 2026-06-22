from __future__ import annotations

import asyncio
import contextvars
import json
import math
import os
from io import BytesIO
from pathlib import Path
import random
import re
import socket
import string
import time
import uuid
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from typing import Any

from fastapi import FastAPI, Request, WebSocket, WebSocketDisconnect
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fastapi.responses import HTMLResponse, JSONResponse, Response, StreamingResponse

app = FastAPI(title='ReMap Core')

STATIC_DIR = Path(__file__).resolve().parent / 'static'
STATIC_DIR.mkdir(exist_ok=True)
app.mount('/static', StaticFiles(directory=str(STATIC_DIR)), name='static')

def get_lan_ip() -> str:
    """Return a likely LAN IPv4 address for phones/tablets on the same network."""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            # No packet is actually sent; connect() lets the OS choose the outbound interface.
            sock.connect(('8.8.8.8', 80))
            ip = sock.getsockname()[0]
            if ip and not ip.startswith('127.'):
                return ip
    except OSError:
        pass
    try:
        ip = socket.gethostbyname(socket.gethostname())
        if ip and not ip.startswith('127.'):
            return ip
    except OSError:
        pass
    return '127.0.0.1'


def get_network_base_url(request: Request) -> str:
    scheme = request.headers.get('x-forwarded-proto') or request.url.scheme or 'http'
    host = request.url.hostname or get_lan_ip()
    port = request.url.port
    if host in {'localhost', '127.0.0.1', '0.0.0.0', '::1'}:
        host = get_lan_ip()
    default_port = 443 if scheme == 'https' else 80
    netloc = host if not port or port == default_port else f'{host}:{port}'
    return f'{scheme}://{netloc}'


def normalize_room_code(value: str | None) -> str:
    allowed = string.ascii_uppercase + string.digits
    return ''.join(ch for ch in str(value or '').upper() if ch in allowed)[:4]

DEFAULTS = {
    'room_title': 'REMAP 방제목',
    'teacher_owner': '',
    'room_code': '',
    'game_mode': 'solo',
    'team_count': 4,
    'map_type': 'open',
    'map_width': 1060,
    'map_height': 612,
    'question_count': 1,
    'question_time_limit': 20,
    'total_game_time': 180,
    'score_win': 3,
    'score_draw': 1,
    'score_lose': 0,
    'player_speed': 7.5,
    'background_data_url': None,
}

MAX_NICKNAME_LEN = 12
MAX_ROOM_TITLE_LEN = 50
MAX_QUESTION_TEXT_LEN = 240
MAX_CHOICE_TEXT_LEN = 100
MAX_BACKGROUND_DATA_URL_LEN = 2_000_000
MOVE_BROADCAST_INTERVAL = 1 / 30
DISCONNECT_GRACE_SECONDS = 35


def clean_text(value: Any, limit: int) -> str:
    text = str(value or '')
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    # Public-service hardening: keep user text printable, but do not preserve
    # raw HTML tag brackets in room titles, nicknames, questions, or choices.
    text = text.replace('<', '').replace('>', '')
    return text.strip()[:limit]


def clamp_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        number = int(float(value))
    except (TypeError, ValueError, OverflowError):
        return default
    return max(minimum, min(maximum, number))


def clamp_float(value: Any, default: float, minimum: float, maximum: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(number):
        return default
    return max(minimum, min(maximum, number))


BLOCKED_PROFANITY_TERMS = [
    '시발', '씨발', 'ㅅㅂ', 'ㅆㅂ', '병신', '븅신', '개새끼', '새끼', '지랄',
    '꺼져', '좆', '존나', '졸라', '개같', '미친놈', '미친년', 'fuck', 'shit'
]

def _normalize_for_profanity(value: Any) -> str:
    text = str(value or '').lower()
    return re.sub(r'[^0-9a-zㄱ-ㅎ가-힣]+', '', text)

def find_blocked_profanity(value: Any) -> str | None:
    normalized = _normalize_for_profanity(value)
    for term in BLOCKED_PROFANITY_TERMS:
        if term in normalized:
            return term
    return None

TEAM_ORDER = ['A', 'B', 'C', 'D', 'E']
TEAM_COLORS = {
    'A': '#3b82f6',  # 파랑
    'B': '#ef4444',  # 빨강
    'C': '#facc15',  # 노랑
    'D': '#22c55e',  # 초록
    'E': '#94a3b8',  # 회색
}
CHAR_COLORS = [
    '#fca5a5', '#ef4444', '#b91c1c', '#ec4899', '#fdba74', '#f97316',
    '#f59e0b', '#fde047', '#bef264', '#22c55e', '#6ee7b7', '#14b8a6',
    '#67e8f9', '#3b82f6', '#1e40af', '#1e3a8a', '#c4b5fd', '#8b5cf6',
    '#7c3aed', '#d946ef', '#8b5a2b', '#ffffff', '#111827'
]


def normalize_hex_color(value: str | None, default: str = '#60a5fa') -> str:
    raw = str(value or '').strip().lstrip('#')
    if len(raw) == 3 and all(ch in '0123456789abcdefABCDEF' for ch in raw):
        raw = ''.join(ch * 2 for ch in raw)
    if len(raw) != 6 or any(ch not in '0123456789abcdefABCDEF' for ch in raw):
        return default
    return '#' + raw.lower()


def hex_to_rgb_tuple(value: str) -> tuple[int, int, int]:
    normalized = normalize_hex_color(value)
    return tuple(int(normalized[i:i+2], 16) for i in (1, 3, 5))


def blend_colors(value: str, target: str, amount: float) -> str:
    amount = max(0.0, min(1.0, amount))
    sr, sg, sb = hex_to_rgb_tuple(value)
    tr, tg, tb = hex_to_rgb_tuple(target)
    rr = round(sr + (tr - sr) * amount)
    rg = round(sg + (tg - sg) * amount)
    rb = round(sb + (tb - sb) * amount)
    return f'#{rr:02x}{rg:02x}{rb:02x}'


def lighten_hex(value: str, amount: float) -> str:
    return blend_colors(value, '#ffffff', amount)


def darken_hex(value: str, amount: float) -> str:
    return blend_colors(value, '#000000', amount)


def build_character_svg(color_value: str | None) -> str:
    """Return the uploaded mascot SVG recolored to the requested player color."""
    body = normalize_hex_color(color_value)
    dark = darken_hex(body, 0.34)
    darker = darken_hex(body, 0.58)
    light = lighten_hex(body, 0.34)
    asset = STATIC_DIR / 'remap_character.svg'
    try:
        svg = asset.read_text(encoding='utf-8')
    except Exception:
        return f'<svg xmlns="http://www.w3.org/2000/svg" width="128" height="128" viewBox="0 0 128 128"><rect width="128" height="128" rx="24" fill="{body}"/><circle cx="44" cy="50" r="9" fill="#fff"/><circle cx="84" cy="50" r="9" fill="#fff"/><rect x="24" y="80" width="80" height="12" rx="6" fill="{darker}"/></svg>'

    # Canvas drawImage can fail on SVG files that keep an external DOCTYPE.
    # Strip XML/DOCTYPE headers and keep the pure inline SVG content.
    svg = re.sub(r'<\?xml[^>]*>\s*', '', svg, flags=re.IGNORECASE)
    svg = re.sub(r'<!DOCTYPE[^>]*(?:\[[\s\S]*?\]\s*)?>\s*', '', svg, flags=re.IGNORECASE)

    replacements = {
        '#4285D3': body,
        '#3868AA': dark,
        '#192653': darker,
        '#7BA1D2': light,
    }
    for src, dst in replacements.items():
        svg = re.sub(src, dst, svg, flags=re.IGNORECASE)
    return svg


@dataclass
class Question:
    text: str
    choices: list[str]
    answer: int


@dataclass
class Player:
    id: str
    nickname: str
    team: str
    color: str
    x: float
    y: float
    direction: str = 'right'
    score: int = 0
    state: str = 'moving'
    battles_played: int = 0
    correct_count: int = 0
    answer_count: int = 0
    ws: WebSocket | None = None
    questions: list[Question] = field(default_factory=list)
    last_move_at: float = 0.0
    reconnect_token: str = ''
    disconnected_at: float | None = None


@dataclass
class PendingParticipant:
    id: str
    nickname: str
    created_at: float = field(default_factory=time.time)


@dataclass
class BattleSide:
    player_id: str
    questions: list[Question]
    current_index: int = 0
    answers: list[dict[str, Any]] = field(default_factory=list)
    finished: bool = False


@dataclass
class Battle:
    id: str
    player_a: str
    player_b: str
    started_at: float
    side_a: BattleSide
    side_b: BattleSide
    status: str = 'active'


def make_room_state() -> dict[str, Any]:
    return {
        'settings': dict(DEFAULTS),
        'players': {},
        'pending_participants': {},
        'encounters': set(),
        'battles': {},
        'game_status': 'idle',
        'game_started_at': None,
        'game_end_at': None,
        'countdown_end_at': None,
        'teacher_clients': set(),
        'logs': [],
        'team_scores': {t: 0 for t in TEAM_ORDER},
        'last_move_broadcast_at': 0.0,
        'ai_reviews': {},
        'ai_review_meta': {},
        # 게임 종료 시점의 결과 스냅샷입니다.
        # 학생이 시상식 화면에서 나가도 최종 순위가 변하지 않도록 보관합니다.
        'final_results': None,
    }


rooms: dict[str, dict[str, Any]] = {}
default_state: dict[str, Any] = make_room_state()
current_room_code: contextvars.ContextVar[str] = contextvars.ContextVar('current_room_code', default='')


def get_room_state(code: str | None, create: bool = False) -> dict[str, Any] | None:
    room_code = normalize_room_code(code)
    if not room_code:
        return default_state if create else None
    if create and room_code not in rooms:
        room = make_room_state()
        room['settings']['room_code'] = room_code
        rooms[room_code] = room
    return rooms.get(room_code)


def bind_room(code: str | None, create: bool = False):
    room_code = normalize_room_code(code)
    if not room_code:
        return None
    if get_room_state(room_code, create=create) is None:
        return None
    return current_room_code.set(room_code)


def unbind_room(token) -> None:
    if token is not None:
        try:
            current_room_code.reset(token)
        except Exception:
            pass


def current_room_state() -> dict[str, Any]:
    code = current_room_code.get('')
    if code and code in rooms:
        return rooms[code]
    return default_state


class StateProxy:
    def _target(self) -> dict[str, Any]:
        return current_room_state()

    def __getitem__(self, key: str) -> Any:
        return self._target()[key]

    def __setitem__(self, key: str, value: Any) -> None:
        self._target()[key] = value

    def __contains__(self, key: str) -> bool:
        return key in self._target()

    def get(self, key: str, default: Any = None) -> Any:
        return self._target().get(key, default)

    def keys(self):
        return self._target().keys()

    def values(self):
        return self._target().values()

    def items(self):
        return self._target().items()


state = StateProxy()


def log(message: str, kind: str = 'system') -> None:
    now = time.strftime('%H:%M:%S')
    state['logs'].append({'time': now, 'message': message, 'kind': kind})
    state['logs'] = state['logs'][-200:]


def student_logs(limit: int = 30) -> list[dict[str, Any]]:
    return [row for row in state['logs'] if row.get('kind') == 'battle'][-limit:]


def allowed_teams() -> list[str]:
    if state['settings'].get('game_mode') != 'team':
        return []
    return TEAM_ORDER[: max(2, min(5, int(state['settings']['team_count'])))]


def new_room_code() -> str:
    alphabet = string.ascii_uppercase + string.digits
    for _ in range(2000):
        code = ''.join(random.choice(alphabet) for _ in range(4))
        if code not in rooms:
            return code
    return ''.join(random.choice(alphabet) for _ in range(6))[:4]


def reset_runtime(keep_room: bool = True) -> None:
    preserved = {}
    if keep_room:
        for k in ['room_title', 'teacher_owner', 'room_code', 'game_mode', 'team_count', 'map_type', 'map_width', 'map_height', 'question_count', 'question_time_limit', 'total_game_time', 'score_win', 'score_draw', 'score_lose', 'player_speed', 'background_data_url']:
            preserved[k] = state['settings'].get(k, DEFAULTS[k])
    state['settings'] = dict(DEFAULTS)
    state['settings'].update(preserved)
    state['players'].clear()
    state['pending_participants'].clear()
    state['encounters'].clear()
    state['battles'].clear()
    state['game_status'] = 'lobby' if state['settings']['room_code'] else 'idle'
    state['game_started_at'] = None
    state['game_end_at'] = None
    state['countdown_end_at'] = None
    state['team_scores'] = {t: 0 for t in TEAM_ORDER}
    state['last_move_broadcast_at'] = 0.0
    state['ai_reviews'] = {}
    state['ai_review_meta'] = {}
    state['final_results'] = None
    state['logs'] = []


def remaining_time() -> int:
    if state['game_status'] != 'running' or not state['game_end_at']:
        return max(0, int(state['settings']['total_game_time']))
    return max(0, int(state['game_end_at'] - time.time()))


def countdown_remaining() -> int:
    if state['game_status'] != 'countdown' or not state['countdown_end_at']:
        return 0
    return max(0, int(math.ceil(state['countdown_end_at'] - time.time())))


def team_rankings() -> list[dict[str, Any]]:
    # 게임 종료 후에는 학생이 나가도 시상식 결과가 흔들리지 않도록
    # 종료 당시 저장해 둔 스냅샷을 우선 사용합니다.
    final_results = state.get('final_results')
    if state.get('game_status') == 'finished' and isinstance(final_results, dict):
        frozen = final_results.get('team_rankings')
        if isinstance(frozen, list):
            return [dict(row) for row in frozen]
    if state['settings'].get('game_mode') != 'team':
        return []
    rows = [{'team': t, 'score': state['team_scores'].get(t, 0)} for t in allowed_teams()]
    rows.sort(key=lambda x: (-x['score'], x['team']))
    for i, row in enumerate(rows, 1):
        row['rank'] = i
    return rows


def rankings() -> list[dict[str, Any]]:
    # 게임 종료 후에는 현재 접속 중인 플레이어 목록이 아니라
    # 종료 순간의 순위 스냅샷을 사용합니다.
    # 그래야 시상식에서 학생이 나가도 1,2,3위가 변하지 않습니다.
    final_results = state.get('final_results')
    if state.get('game_status') == 'finished' and isinstance(final_results, dict):
        frozen = final_results.get('rankings')
        if isinstance(frozen, list):
            return [dict(row) for row in frozen]
    rows = sorted(state['players'].values(), key=lambda p: (-p.score, -p.correct_count, p.nickname.lower()))
    out = []
    for idx, p in enumerate(rows, 1):
        out.append({
            'rank': idx,
            'player_id': p.id,
            'nickname': p.nickname,
            'score': p.score,
            'team': p.team,
            'state': p.state,
            'color': p.color,
            'correct_count': p.correct_count,
            'answer_count': p.answer_count,
            'battles_played': p.battles_played,
        })
    return out


def get_public_player(player: Player) -> dict[str, Any]:
    battled_ids: list[str] = []
    for a_id, b_id in state['encounters']:
        if player.id == a_id:
            battled_ids.append(b_id)
        elif player.id == b_id:
            battled_ids.append(a_id)
    return {
        'id': player.id,
        'nickname': player.nickname,
        'team': player.team,
        'team_color': TEAM_COLORS.get(player.team, '#64748b'),
        'color': player.color,
        'x': player.x,
        'y': player.y,
        'direction': player.direction,
        'score': player.score,
        'state': player.state,
        'battles_played': player.battles_played,
        'correct_count': player.correct_count,
        'answer_count': player.answer_count,
        'battled_ids': battled_ids,
        'connected': bool(player.ws),
    }


def battle_progress_payload() -> list[dict[str, Any]]:
    rows = []
    for b in state['battles'].values():
        if b.status != 'active':
            continue
        pa = state['players'].get(b.player_a)
        pb = state['players'].get(b.player_b)
        if not pa or not pb:
            continue
        rows.append({
            'battle_id': b.id,
            'players': [pa.nickname, pb.nickname],
            'progress': max(b.side_a.current_index, b.side_b.current_index),
            'total': max(len(b.side_a.questions), len(b.side_b.questions)),
        })
    return rows


async def safe_send(ws: WebSocket, payload: dict[str, Any]) -> None:
    try:
        await ws.send_json(payload)
    except Exception:
        pass


def full_payload() -> dict[str, Any]:
    return {
        'type': 'state',
        'room': {
            'title': state['settings']['room_title'],
            'owner': state['settings'].get('teacher_owner', ''),
            'code': state['settings']['room_code'],
            'game_mode': state['settings'].get('game_mode', 'solo'),
            'team_count': state['settings']['team_count'],
            'map_type': state['settings'].get('map_type', 'open'),
            'map_label': map_label(),
        },
        'players': [get_public_player(p) for p in state['players'].values()],
        'participants': participants_payload(),
        'unsubmitted_count': unsubmitted_count(),
        'team_counts': team_counts_payload(),
        'rankings': rankings(),
        'team_rankings': team_rankings(),
        'battles': battle_progress_payload(),
        'settings': state['settings'],
        'game_status': state['game_status'],
        'remaining_time': remaining_time(),
        'countdown_remaining': countdown_remaining(),
        'map_walls': get_map_walls(),
        'logs': state['logs'][-30:],
        'student_logs': student_logs(30),
        'ai_reviews': list(state.get('ai_reviews', {}).values()),
        'ai_review_meta': state.get('ai_review_meta', {}),
    }


def cleanup_stale_disconnected_players() -> None:
    if state.get('game_status') == 'finished':
        return
    now = time.time()
    stale_ids = []
    for pid, p in list(state['players'].items()):
        if p.ws is None and p.disconnected_at and now - p.disconnected_at > DISCONNECT_GRACE_SECONDS:
            stale_ids.append(pid)
    for pid in stale_ids:
        p = state['players'].pop(pid, None)
        if p:
            log(f'{p.nickname} 재접속 시간 초과 - 퇴장 처리')


async def broadcast_state() -> None:
    cleanup_stale_disconnected_players()
    payload = full_payload()
    for p in list(state['players'].values()):
        if p.ws:
            await safe_send(p.ws, payload)
    for ws in list(state['teacher_clients']):
        await safe_send(ws, payload)


async def send_to_player(player_id: str, payload: dict[str, Any]) -> None:
    p = state['players'].get(player_id)
    if p and p.ws:
        await safe_send(p.ws, payload)


def map_label() -> str:
    return '미로형 맵' if state['settings'].get('map_type') == 'maze' else '오픈 스퀘어'


def get_map_walls() -> list[dict[str, float]]:
    if state['settings'].get('map_type') != 'maze':
        return []
    w = int(state['settings']['map_width'])
    h = int(state['settings']['map_height'])

    # PPT로 전달된 참고안처럼, 맵 전체에 정사각형 장애물을 규칙적으로 배치한다.
    # 길이 완전히 막히지 않도록 각 블록 사이 간격을 충분히 두고,
    # 현재 오픈맵과 잘 어울리도록 부드러운 밝은 블루 계열 오브젝트로 렌더링한다.
    cols = 7
    rows = 4
    block = max(29, round(min(w * 0.055, h * 0.11) * 0.85))
    margin_x = max(32, round(w * 0.11))
    margin_y = max(24, round(h * 0.12))
    gap_x = max(26, (w - (margin_x * 2) - (block * cols)) / max(1, cols - 1))
    gap_y = max(24, (h - (margin_y * 2) - (block * rows)) / max(1, rows - 1))

    walls: list[dict[str, float]] = []
    for row in range(rows):
        for col in range(cols):
            x = margin_x + col * (block + gap_x)
            y = margin_y + row * (block + gap_y)
            walls.append({'x': round(x, 2), 'y': round(y, 2), 'w': float(block), 'h': float(block)})
    return walls


def collides_with_walls(x: float, y: float, size: float = 14) -> bool:
    if state['settings'].get('map_type') != 'maze':
        return False
    left = x - size
    top = y - size
    right = x + size
    bottom = y + size
    for wall in get_map_walls():
        if right > wall['x'] and left < wall['x'] + wall['w'] and bottom > wall['y'] and top < wall['y'] + wall['h']:
            return True
    return False


def participants_payload() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pending in state['pending_participants'].values():
        rows.append({'id': pending.id, 'nickname': pending.nickname, 'team': '', 'submitted': False})
    for p in state['players'].values():
        rows.append({'id': p.id, 'nickname': p.nickname, 'team': p.team, 'submitted': True})
    rows.sort(key=lambda x: (not x['submitted'], x['nickname'].lower()))
    return rows


def team_counts_payload() -> list[dict[str, Any]]:
    if state['settings'].get('game_mode') != 'team':
        return []
    rows = []
    for team in allowed_teams():
        count = sum(1 for p in state['players'].values() if p.team == team)
        rows.append({'team': team, 'count': count})
    return rows


def unsubmitted_count() -> int:
    return len(state['pending_participants'])


def random_spawn() -> tuple[float, float]:
    w = int(state['settings']['map_width'])
    h = int(state['settings']['map_height'])
    for _ in range(50):
        x = random.randint(50, max(60, w - 50))
        y = random.randint(50, max(60, h - 50))
        if collides_with_walls(x, y):
            continue
        if all(math.dist((x, y), (p.x, p.y)) > 70 for p in state['players'].values()):
            return x, y
    return 60, 60


def can_battle(a: Player, b: Player) -> bool:
    if not a.ws or not b.ws:
        return False
    if a.id == b.id:
        return False
    if state['settings'].get('game_mode') == 'team' and a.team == b.team:
        return False
    if a.state == 'battling' or b.state == 'battling':
        return False
    if state['game_status'] != 'running':
        return False
    if tuple(sorted([a.id, b.id])) in state['encounters']:
        return False
    if not a.questions or not b.questions:
        return False
    return True


def possible_battle_pairs() -> set[tuple[str, str]]:
    players = [p for p in state['players'].values() if p.questions and p.ws]
    pairs: set[tuple[str, str]] = set()
    team_mode = state['settings'].get('game_mode') == 'team'
    for i, a in enumerate(players):
        for b in players[i + 1:]:
            if team_mode and a.team == b.team:
                continue
            pairs.add(tuple(sorted([a.id, b.id])))
    return pairs


def has_active_battles() -> bool:
    return any(b.status == 'active' for b in state['battles'].values())


async def maybe_auto_end_all_battles_complete() -> None:
    if state['game_status'] != 'running':
        return
    if has_active_battles():
        return
    possible = possible_battle_pairs()
    if not possible:
        return
    completed = {pair for pair in state['encounters'] if pair in possible}
    if len(completed) >= len(possible):
        log('모든 가능한 배틀 완료 - 자동 종료')
        await end_game('all_battles_complete')


async def auto_end_after_results_delay() -> None:
    # Let the last battle result briefly show before moving everyone to the ceremony.
    await asyncio.sleep(2.8)
    await maybe_auto_end_all_battles_complete()

def pick_questions_no_repeat(source: list[Question], count: int) -> list[Question]:
    if not source:
        return []
    return random.sample(source, min(count, len(source)))


async def try_collisions_for(player: Player) -> None:
    if player.state == 'battling':
        return
    for other in state['players'].values():
        if not can_battle(player, other):
            continue
        if math.dist((player.x, player.y), (other.x, other.y)) <= 28:
            await start_battle(player, other)
            return


async def start_battle(a: Player, b: Player) -> None:
    count = int(state['settings']['question_count'])
    if len(a.questions) < count or len(b.questions) < count:
        return
    state['encounters'].add(tuple(sorted([a.id, b.id])))
    a.state = 'battling'
    b.state = 'battling'
    battle_id = str(uuid.uuid4())
    battle = Battle(
        id=battle_id,
        player_a=a.id,
        player_b=b.id,
        started_at=time.time(),
        side_a=BattleSide(player_id=a.id, questions=pick_questions_no_repeat(b.questions, count)),
        side_b=BattleSide(player_id=b.id, questions=pick_questions_no_repeat(a.questions, count)),
    )
    state['battles'][battle_id] = battle
    log(f'{a.nickname} vs {b.nickname} 배틀', 'battle')
    await send_to_player(a.id, {'type': 'battle_intro', 'me': a.nickname, 'opponent': b.nickname})
    await send_to_player(b.id, {'type': 'battle_intro', 'me': b.nickname, 'opponent': a.nickname})
    await broadcast_state()
    await asyncio.sleep(0.9)
    await send_battle_question(battle, a.id)
    await send_battle_question(battle, b.id)
    await broadcast_state()


async def send_battle_question(battle: Battle, player_id: str) -> None:
    side = battle.side_a if battle.side_a.player_id == player_id else battle.side_b
    if side.current_index >= len(side.questions):
        side.finished = True
        await maybe_finish_battle(battle)
        return
    q = side.questions[side.current_index]
    opponent_id = battle.player_b if player_id == battle.player_a else battle.player_a
    await send_to_player(player_id, {
        'type': 'battle_question',
        'battle_id': battle.id,
        'index': side.current_index + 1,
        'total': len(side.questions),
        'time_limit': int(state['settings']['question_time_limit']),
        'question': {'text': q.text, 'choices': q.choices},
        'opponent': state['players'][opponent_id].nickname,
        'room_title': state['settings']['room_title'],
    })
    asyncio.create_task(question_timeout(battle.id, player_id, side.current_index))


async def question_timeout(battle_id: str, player_id: str, expected_index: int) -> None:
    await asyncio.sleep(int(state['settings']['question_time_limit']))
    battle = state['battles'].get(battle_id)
    if not battle or battle.status != 'active':
        return
    side = battle.side_a if battle.side_a.player_id == player_id else battle.side_b
    if side.current_index != expected_index:
        return
    side.answers.append({'selected': None, 'correct': False, 'time_used': int(state['settings']['question_time_limit'])})
    player = state['players'].get(player_id)
    if player:
        player.answer_count += 1
    side.current_index += 1
    await send_to_player(player_id, {'type': 'battle_feedback', 'correct': False, 'selected': None, 'timed_out': True})
    await send_battle_question(battle, player_id)
    await broadcast_state()


async def maybe_finish_battle(battle: Battle) -> None:
    if not (battle.side_a.finished and battle.side_b.finished) or battle.status != 'active':
        return
    battle.status = 'finished'
    a = state['players'][battle.player_a]
    b = state['players'][battle.player_b]
    a_correct = sum(1 for ans in battle.side_a.answers if ans['correct'])
    b_correct = sum(1 for ans in battle.side_b.answers if ans['correct'])
    a.battles_played += 1
    b.battles_played += 1
    team_mode = state['settings'].get('game_mode') == 'team'
    if a_correct > b_correct:
        a.score += int(state['settings']['score_win'])
        b.score += int(state['settings']['score_lose'])
        if team_mode:
            state['team_scores'][a.team] += int(state['settings']['score_win'])
            state['team_scores'][b.team] += int(state['settings']['score_lose'])
        a_result, b_result = 'win', 'lose'
        result_text = f'{a.nickname} 승리 / {b.nickname} 패배'
    elif b_correct > a_correct:
        a.score += int(state['settings']['score_lose'])
        b.score += int(state['settings']['score_win'])
        if team_mode:
            state['team_scores'][a.team] += int(state['settings']['score_lose'])
            state['team_scores'][b.team] += int(state['settings']['score_win'])
        a_result, b_result = 'lose', 'win'
        result_text = f'{b.nickname} 승리 / {a.nickname} 패배'
    else:
        a.score += int(state['settings']['score_draw'])
        b.score += int(state['settings']['score_draw'])
        if team_mode:
            state['team_scores'][a.team] += int(state['settings']['score_draw'])
            state['team_scores'][b.team] += int(state['settings']['score_draw'])
        a_result = b_result = 'draw'
        result_text = f'{a.nickname} / {b.nickname} 무승부'
    a.state = 'moving'
    b.state = 'moving'
    log(result_text, 'battle')
    await send_to_player(a.id, {'type': 'battle_result', 'result': a_result, 'opponent': b.nickname, 'my_correct': a_correct, 'opponent_correct': b_correct, 'my_score': a.score})
    await send_to_player(b.id, {'type': 'battle_result', 'result': b_result, 'opponent': a.nickname, 'my_correct': b_correct, 'opponent_correct': a_correct, 'my_score': b.score})
    await broadcast_state()
    asyncio.create_task(auto_end_after_results_delay())


def build_game_end_payload() -> dict[str, Any]:
    ranked = rankings()
    best_correct_player = max(state['players'].values(), key=lambda p: (p.correct_count, -p.answer_count, p.nickname), default=None)
    most_battles_player = max(state['players'].values(), key=lambda p: (p.battles_played, p.correct_count, p.nickname), default=None)
    return {
        'type': 'game_end',
        'room_title': state['settings'].get('room_title', DEFAULTS['room_title']),
        'game_mode': state['settings'].get('game_mode', 'solo'),
        'rankings': ranked,
        'team_rankings': team_rankings(),
        'logs': student_logs(100),
        'player_stats': [
            {
                'nickname': p.nickname,
                'team': p.team,
                'score': p.score,
                'correct_count': p.correct_count,
                'answer_count': p.answer_count,
                'battles_played': p.battles_played,
                'color': p.color,
            } for p in sorted(state['players'].values(), key=lambda p: (-p.score, -p.correct_count, p.nickname.lower()))
        ],
        'best_correct': ({'nickname': best_correct_player.nickname, 'correct_count': best_correct_player.correct_count} if best_correct_player else None),
        'most_battles': ({'nickname': most_battles_player.nickname, 'battles_played': most_battles_player.battles_played} if most_battles_player else None),
        'winner_team': (team_rankings()[0] if team_rankings() else None),
    }


async def end_game(triggered_by: str = 'system') -> None:
    if state['game_status'] == 'finished':
        return
    state['game_status'] = 'finished'
    log(f'게임 종료 ({triggered_by})')
    payload = build_game_end_payload()
    # 종료 시점 결과를 고정합니다. 이후 학생이 퇴장/접속 종료해도
    # 교사/학생 시상식과 순위표는 이 스냅샷을 기준으로 유지됩니다.
    state['final_results'] = json.loads(json.dumps(payload, ensure_ascii=False))
    for p in list(state['players'].values()):
        if p.ws:
            await safe_send(p.ws, payload)
    for ws in list(state['teacher_clients']):
        await safe_send(ws, payload)
    await broadcast_state()


async def game_timer_loop() -> None:
    while True:
        await asyncio.sleep(1)
        for room_code in list(rooms.keys()):
            token = bind_room(room_code)
            if token is None:
                continue
            try:
                if state['game_status'] == 'countdown':
                    if countdown_remaining() <= 0:
                        state['game_status'] = 'running'
                        state['game_started_at'] = time.time()
                        state['game_end_at'] = time.time() + int(state['settings']['total_game_time'])
                        state['countdown_end_at'] = None
                        log('게임 시작')
                        await broadcast_state()
                    else:
                        await broadcast_state()
                elif state['game_status'] == 'running':
                    if remaining_time() <= 0:
                        await end_game('timer')
                    else:
                        await maybe_auto_end_all_battles_complete()
                        if state['game_status'] == 'running':
                            await broadcast_state()
            finally:
                unbind_room(token)


@app.on_event('startup')
async def startup_event() -> None:
    asyncio.create_task(game_timer_loop())


@app.get('/api/access_info')
async def access_info(request: Request, code: str = '') -> JSONResponse:
    room_code = normalize_room_code(code)
    base_url = get_network_base_url(request)
    student_url = f'{base_url}/?code={room_code}' if room_code else f'{base_url}/'
    return JSONResponse({
        'base_url': base_url,
        'lan_ip': get_lan_ip(),
        'student_url': student_url,
        'teacher_url': f'{base_url}/teacher',
        'room_code': room_code,
    })


@app.get('/api/room/qr.svg')
async def room_qr_svg(request: Request, code: str = '') -> Response:
    room_code = normalize_room_code(code)
    base_url = get_network_base_url(request)
    target_url = f'{base_url}/?code={room_code}' if room_code else f'{base_url}/'
    try:
        import qrcode
        import qrcode.image.svg
    except Exception:
        fallback = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="240" height="240" viewBox="0 0 240 240">'
            '<rect width="240" height="240" rx="18" fill="#ffffff"/>'
            '<text x="120" y="108" text-anchor="middle" font-family="Arial" font-size="16" fill="#111827">QR 라이브러리 필요</text>'
            '<text x="120" y="134" text-anchor="middle" font-family="Arial" font-size="13" fill="#475569">pip install -r requirements.txt</text>'
            '</svg>'
        )
        return Response(fallback, media_type='image/svg+xml')
    img = qrcode.make(
        target_url,
        image_factory=qrcode.image.svg.SvgPathImage,
        box_size=8,
        border=3,
    )
    buffer = BytesIO()
    img.save(buffer)
    return Response(buffer.getvalue(), media_type='image/svg+xml', headers={'Cache-Control': 'no-store'})


@app.get('/character/{color_value}.svg')
async def character_svg(color_value: str) -> Response:
    return Response(content=build_character_svg(color_value), media_type='image/svg+xml')


@app.get('/')
async def index() -> HTMLResponse:
    return HTMLResponse(STUDENT_HTML)


@app.get('/teacher')
async def teacher_page() -> HTMLResponse:
    return HTMLResponse(TEACHER_HTML)


@app.get('/teacher/ceremony')
async def teacher_ceremony_page(room: str = '') -> HTMLResponse:
    token = bind_room(room) if room else None
    try:
        stored = state.get('final_results')
        payload = stored if isinstance(stored, dict) else build_game_end_payload()
        payload_json = json.dumps(payload, ensure_ascii=False)
        html = TEACHER_CEREMONY_HTML.replace('__PAYLOAD__', payload_json)
        return HTMLResponse(html)
    finally:
        unbind_room(token)


@app.get('/api/room/info')
async def room_info(code: str) -> JSONResponse:
    room_code = normalize_room_code(code)
    token = bind_room(room_code)
    if token is None:
        return JSONResponse({'ok': False, 'title': '', 'game_mode': 'solo', 'team_count': 0, 'question_count': 0, 'question_time_limit': 0, 'map_width': 0, 'map_height': 0, 'map_type': 'open', 'map_label': '', 'status': 'invalid'})
    try:
        valid = bool(state['settings']['room_code']) and room_code == state['settings']['room_code']
        return JSONResponse({'ok': valid, 'title': state['settings']['room_title'] if valid else '', 'teacher_owner': state['settings'].get('teacher_owner', '') if valid else '', 'game_mode': state['settings'].get('game_mode', 'solo') if valid else 'solo', 'team_count': state['settings']['team_count'] if valid and state['settings'].get('game_mode') == 'team' else 0, 'question_count': state['settings']['question_count'] if valid else 0, 'question_time_limit': state['settings']['question_time_limit'] if valid else 0, 'map_width': state['settings']['map_width'] if valid else 0, 'map_height': state['settings']['map_height'] if valid else 0, 'map_type': state['settings'].get('map_type', 'open') if valid else 'open', 'map_label': map_label() if valid else '', 'status': state['game_status'] if valid else 'invalid'})
    finally:
        unbind_room(token)


@app.post('/api/room/prepare')
async def room_prepare(payload: dict[str, Any]) -> JSONResponse:
    code = normalize_room_code(payload.get('code'))
    token = bind_room(code)
    if token is None:
        return JSONResponse({'ok': False, 'message': '존재하지 않는 방 코드입니다.'}, status_code=400)
    nickname = clean_text(payload.get('nickname'), MAX_NICKNAME_LEN)
    if not state['settings']['room_code'] or code != state['settings']['room_code']:
        unbind_room(token)
        return JSONResponse({'ok': False, 'message': '존재하지 않는 방 코드입니다.'}, status_code=400)
    if state['game_status'] in ('running', 'countdown'):
        return JSONResponse({'ok': False, 'message': '이미 시작된 방입니다.'}, status_code=400)
    if not nickname:
        return JSONResponse({'ok': False, 'message': '닉네임을 입력하세요.'}, status_code=400)
    all_names = {p.nickname.lower() for p in state['players'].values()} | {pp.nickname.lower() for pp in state['pending_participants'].values()}
    if nickname.lower() in all_names:
        return JSONResponse({'ok': False, 'message': '이미 사용 중인 닉네임입니다.'}, status_code=400)
    pending_id = str(uuid.uuid4())
    state['pending_participants'][pending_id] = PendingParticipant(id=pending_id, nickname=nickname)
    await broadcast_state()
    return JSONResponse({'ok': True, 'pending_id': pending_id, 'title': state['settings']['room_title'], 'game_mode': state['settings'].get('game_mode', 'solo'), 'team_count': state['settings']['team_count'] if state['settings'].get('game_mode') == 'team' else 0, 'question_count': state['settings']['question_count'], 'question_time_limit': state['settings']['question_time_limit'], 'map_width': state['settings']['map_width'], 'map_height': state['settings']['map_height'], 'map_type': state['settings'].get('map_type', 'open'), 'map_label': map_label(), 'status': state['game_status']})


@app.post('/api/room/cancel_prepare')
async def room_cancel_prepare(payload: dict[str, Any]) -> JSONResponse:
    pending_id = str(payload.get('pending_id') or '').strip()
    room_code = normalize_room_code(payload.get('code'))
    candidate_codes = [room_code] if room_code else list(rooms.keys())
    for code in candidate_codes:
        token = bind_room(code)
        if token is None:
            continue
        try:
            if pending_id and pending_id in state['pending_participants']:
                state['pending_participants'].pop(pending_id, None)
                await broadcast_state()
                break
        finally:
            unbind_room(token)
    return JSONResponse({'ok': True})


@app.get('/api/state')
async def get_state(room: str = '') -> JSONResponse:
    room_code = normalize_room_code(room)
    token = bind_room(room_code) if room_code else None
    try:
        return JSONResponse(full_payload())
    finally:
        unbind_room(token)


@app.post('/api/teacher/create_room')
async def create_room(payload: dict[str, Any]) -> JSONResponse:
    # Multiple teachers can use the same server. Every teacher-created room gets
    # its own state bucket keyed by the room code, so rooms do not interfere.
    room_code = new_room_code()
    token = bind_room(room_code, create=True)
    if token is None:
        return JSONResponse({'ok': False, 'message': '방 코드를 생성하지 못했습니다.'}, status_code=500)
    reset_runtime(keep_room=False)
    room_title = clean_text(payload.get('room_title'), MAX_ROOM_TITLE_LEN) or DEFAULTS['room_title']
    teacher_owner = clean_text(payload.get('teacher_owner'), MAX_NICKNAME_LEN)
    map_type = str(payload.get('map_type') or DEFAULTS['map_type']).strip()
    if map_type not in {'open', 'maze'}:
        map_type = DEFAULTS['map_type']
    background_data_url = payload.get('background_data_url')
    if isinstance(background_data_url, str) and background_data_url.startswith('data:image/'):
        background_data_url = background_data_url[:MAX_BACKGROUND_DATA_URL_LEN]
    else:
        background_data_url = None
    state['settings']['room_title'] = room_title
    state['settings']['teacher_owner'] = teacher_owner
    state['settings']['room_code'] = room_code
    state['settings']['game_mode'] = 'team' if str(payload.get('game_mode') or 'solo') == 'team' else 'solo'
    state['settings']['team_count'] = clamp_int(payload.get('team_count'), DEFAULTS['team_count'], 2, 5) if state['settings']['game_mode'] == 'team' else 0
    state['settings']['map_type'] = map_type
    state['settings']['map_width'] = clamp_int(payload.get('map_width'), DEFAULTS['map_width'], 480, 2400)
    state['settings']['map_height'] = clamp_int(payload.get('map_height'), DEFAULTS['map_height'], 360, 1600)
    state['settings']['question_count'] = clamp_int(payload.get('question_count'), DEFAULTS['question_count'], 1, 20)
    state['settings']['question_time_limit'] = clamp_int(payload.get('question_time_limit'), DEFAULTS['question_time_limit'], 5, 180)
    state['settings']['total_game_time'] = clamp_int(payload.get('total_game_time'), DEFAULTS['total_game_time'], 30, 3600)
    state['settings']['score_win'] = clamp_int(payload.get('score_win'), DEFAULTS['score_win'], -20, 100)
    state['settings']['score_draw'] = clamp_int(payload.get('score_draw'), DEFAULTS['score_draw'], -20, 100)
    state['settings']['score_lose'] = clamp_int(payload.get('score_lose'), DEFAULTS['score_lose'], -20, 100)
    state['settings']['player_speed'] = clamp_float(payload.get('player_speed'), DEFAULTS['player_speed'], 1.0, 20.0)
    state['settings']['background_data_url'] = background_data_url
    state['game_status'] = 'lobby'
    log(f"방 생성: {state['settings']['room_title']} ({state['settings']['room_code']})")
    await broadcast_state()
    response = JSONResponse({'ok': True, 'room_code': state['settings']['room_code'], 'settings': state['settings']})
    unbind_room(token)
    return response


@app.post('/api/teacher/start')
async def start_game(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': False, 'message': '관리할 방을 찾을 수 없습니다.'}, status_code=404)
    if not state['settings']['room_code']:
        unbind_room(token)
        return JSONResponse({'ok': False, 'message': '먼저 방을 생성하세요.'}, status_code=400)
    state['game_status'] = 'countdown'
    state['game_started_at'] = None
    state['game_end_at'] = None
    state['countdown_end_at'] = time.time() + 4
    log('게임 시작 카운트다운')
    await broadcast_state()
    unbind_room(token)
    return JSONResponse({'ok': True})


@app.post('/api/teacher/end')
async def teacher_end_game(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': False, 'message': '관리할 방을 찾을 수 없습니다.'}, status_code=404)
    try:
        await end_game('teacher')
        return JSONResponse({'ok': True})
    finally:
        unbind_room(token)


@app.post('/api/teacher/new_room_setup')
async def teacher_new_room_setup(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': True})
    # [새 방 설정]은 기존 방을 닫고 학생 화면을 코드 입력 첫 화면으로 돌려보낸다.
    # 게임이 아직 시작되지 않은 lobby 상태에서는 /api/teacher/end가 동작하지 않으므로
    # 별도의 방 닫기 흐름으로 참가자/대기자를 모두 정리한다.
    for p in list(state['players'].values()):
        if p.ws:
            await safe_send(p.ws, {'type': 'reset', 'target': 'home', 'message': '선생님이 새 방 설정을 시작했습니다. 방 코드를 다시 입력하세요.'})
            try:
                await p.ws.close()
            except Exception:
                pass
    reset_runtime(keep_room=False)
    log('새 방 설정을 위해 기존 방을 닫았습니다.')
    await broadcast_state()
    unbind_room(token)
    return JSONResponse({'ok': True})


@app.post('/api/teacher/clear_background')
async def clear_background(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': False, 'message': '관리할 방을 찾을 수 없습니다.'}, status_code=404)
    state['settings']['background_data_url'] = None
    log('기본 배경 복원')
    await broadcast_state()
    unbind_room(token)
    return JSONResponse({'ok': True})

@app.post('/api/teacher/reset')
async def teacher_reset_game(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': False, 'message': '관리할 방을 찾을 수 없습니다.'}, status_code=404)
    for p in list(state['players'].values()):
        if p.ws:
            await safe_send(p.ws, {'type': 'reset', 'message': '선생님이 다음 게임 준비를 시작했습니다.'})
            try:
                await p.ws.close()
            except Exception:
                pass
    reset_runtime(keep_room=True)
    log('다음 게임 준비 완료')
    await broadcast_state()
    unbind_room(token)
    return JSONResponse({'ok': True})



AI_REVIEW_STATUSES = {'통과', '확인 필요', '오류 가능성', '표현 모호', '검토 불가'}
MAX_AI_REVIEW_QUESTIONS = 80


def collect_question_review_items() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    players = sorted(state['players'].values(), key=lambda p: (p.team if p.team else 'Z', p.nickname.lower()))
    for player in players:
        for idx, q in enumerate(player.questions, start=1):
            answer_text = q.choices[q.answer] if 0 <= q.answer < len(q.choices) else ''
            items.append({
                'key': f'{player.id}:{idx}',
                'player_id': player.id,
                'nickname': player.nickname,
                'team': player.team,
                'question_no': idx,
                'question': q.text,
                'choices': list(q.choices),
                'selected_answer_number': int(q.answer) + 1,
                'selected_answer': answer_text,
            })
    return items


def answer_candidate_numbers(text: Any) -> set[int]:
    raw = str(text or '')
    found: set[int] = set()
    circled = {'①': 1, '②': 2, '③': 3, '④': 4, '➀': 1, '➁': 2, '➂': 3, '➃': 4}
    for ch, num in circled.items():
        if ch in raw:
            found.add(num)
    for m in re.finditer(r'(?<!\d)([1-4])\s*(?:번|[.)]|$)', raw):
        found.add(int(m.group(1)))
    return found


def needs_external_fact_check(question: Any) -> bool:
    q = str(question or '')
    external_terms = [
        '위치', '지역구', '주소', '소재', '소재지', '어디', '현재', '최근', '올해', '작년', '내년',
        '학교', '기관', '회사', '업체', '교장', '교감', '대통령', '총리', '시장', '교육감',
        '날짜', '시간', '가격', '요금', '전화', '홈페이지', '인구', '순위', '랭킹'
    ]
    return any(term in q for term in external_terms)


def sanitize_ai_review_item(raw: dict[str, Any], item_by_key: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    key = str(raw.get('key') or '').strip()
    base = item_by_key.get(key)
    if not base:
        return None
    status = str(raw.get('status') or '검토 불가').strip()
    if status not in AI_REVIEW_STATUSES:
        status = '검토 불가'
    summary = clean_text(raw.get('summary'), 220) or 'AI 검토 의견 없음'
    suggested_answer = clean_text(raw.get('suggested_answer'), 120)
    confidence = str(raw.get('confidence') or '').strip()
    if confidence not in {'높음', '보통', '낮음'}:
        confidence = '보통'

    # 보수적 후처리: AI가 여러 선택지를 동시에 정답처럼 제시하면 최종 판정하지 않는다.
    candidates = answer_candidate_numbers(suggested_answer)
    if len(candidates) >= 2:
        status = '표현 모호'
        confidence = '낮음'
        summary = 'AI가 여러 선택지를 정답 후보로 판단했습니다. 문제 표현 또는 선택지 구성을 교사가 확인해야 합니다.'
        suggested_answer = '교사 확인 필요'

    # 외부 사실 확인형 문항은 AI가 맞다/틀리다를 단정하지 않도록 낮춘다.
    # 예: 학교 위치, 지역구, 최신 정보, 기관/인물/날짜/가격 등
    if needs_external_fact_check(base.get('question')) and status in {'통과', '오류 가능성'}:
        status = '확인 필요'
        confidence = '낮음'
        summary = '학교 위치, 지역구, 주소, 현재 정보처럼 외부 사실 확인이 필요한 문항입니다. AI 단독 판단보다 교사 확인이 필요합니다.'
        suggested_answer = '교사 확인 필요'

    return {
        **base,
        'status': status,
        'summary': summary,
        'suggested_answer': suggested_answer,
        'confidence': confidence,
    }


def parse_json_from_text(text: str) -> dict[str, Any]:
    text = str(text or '').strip()
    if not text:
        raise ValueError('empty response')
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    match = re.search(r'```(?:json)?\s*(\{[\s\S]*?\})\s*```', text, re.IGNORECASE)
    if match:
        return json.loads(match.group(1))
    first = text.find('{')
    last = text.rfind('}')
    if first >= 0 and last > first:
        return json.loads(text[first:last+1])
    raise ValueError('json not found')




def make_external_fact_review(base: dict[str, Any]) -> dict[str, Any]:
    return {
        **base,
        'status': '확인 필요',
        'summary': '위치, 주소, 지역구, 학교/기관 정보, 최신 정보처럼 외부 사실 확인이 필요한 문항입니다. AI가 검색 없이 단정하지 않고 교사 확인이 필요합니다.',
        'suggested_answer': '교사 확인 필요',
        'confidence': '낮음',
    }


def make_ai_unavailable_review(base: dict[str, Any], message: str) -> dict[str, Any]:
    return {
        **base,
        'status': '검토 불가',
        'summary': clean_text(message, 220) or 'AI 검토를 일시적으로 사용할 수 없습니다. 게임 진행에는 영향이 없습니다.',
        'suggested_answer': '교사 확인 필요',
        'confidence': '낮음',
    }


def friendly_gemini_error(code: int | None = None, detail: str = '') -> str:
    detail_text = str(detail or '')
    if code == 503:
        return 'Gemini API 서버가 일시적으로 혼잡합니다. 잠시 후 다시 AI 문제 검토를 눌러주세요. 학생 입장, 게임 진행, 배틀 기능에는 영향이 없습니다.'
    if code == 429:
        return 'Gemini 무료 API 요청 한도 또는 사용량 제한에 도달했을 수 있습니다. 잠시 후 다시 시도하거나 나중에 검토해주세요.'
    if code in {401, 403}:
        return 'Gemini API 키 또는 권한을 확인해야 합니다. Render 환경변수 GEMINI_API_KEY가 올바른지 확인해주세요.'
    if code == 404:
        return 'Gemini 모델명을 찾을 수 없습니다. Render 환경변수 GEMINI_MODEL 값을 gemini-2.5-flash처럼 사용 가능한 모델명으로 확인해주세요.'
    if code in {500, 502, 504}:
        return 'Gemini API가 일시적으로 응답하지 않습니다. 잠시 후 다시 시도해주세요.'
    if detail_text:
        compact = re.sub(r'\s+', ' ', detail_text).strip()[:180]
        return f'AI 검토 API 오류가 발생했습니다. 잠시 후 다시 시도해주세요. ({compact})'
    return 'AI 검토 API 오류가 발생했습니다. 잠시 후 다시 시도해주세요.'


def gemini_model_candidates() -> list[str]:
    primary = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash').strip() or 'gemini-2.5-flash'
    fallback_raw = os.environ.get('GEMINI_FALLBACK_MODELS') or os.environ.get('GEMINI_FALLBACK_MODEL') or 'gemini-2.5-flash-lite'
    values = [primary] + [v.strip() for v in str(fallback_raw).split(',') if v.strip()]
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            out.append(value)
    return out

def call_gemini_review(items: list[dict[str, Any]]) -> dict[str, Any]:
    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        return {
            'ok': False,
            'error_kind': 'missing_key',
            'message': 'GEMINI_API_KEY가 설정되어 있지 않습니다. Render Environment 또는 실행 환경 변수에 Gemini API 키를 넣으면 AI 검토를 사용할 수 있습니다.',
            'items': [],
        }

    payload_items = []
    for item in items[:MAX_AI_REVIEW_QUESTIONS]:
        payload_items.append({
            'key': item['key'],
            'question': item['question'],
            'choices': item['choices'],
            'selected_answer_number': item['selected_answer_number'],
            'selected_answer': item['selected_answer'],
        })

    prompt_instructions = """너는 교사용 참고 의견을 제공하는 보조 검토자다. 최종 정답 판정자가 아니다.
학생 개인정보를 추론하지 말고, 제공된 문제와 선택지만 보고 매우 보수적으로 판단한다.
학생이 고른 정답이 명확히 맞아 보일 때만 "통과"로 분류한다.
학생이 고른 정답이 명확히 틀렸다고 확신할 때만 "오류 가능성"으로 분류한다.
조금이라도 확실하지 않으면 "확인 필요" 또는 "검토 불가"로 분류한다.
정답 후보는 최대 1개만 제시한다. 여러 선택지가 모두 맞을 수 있거나 둘 이상을 정답 후보로 보게 되면 "표현 모호"로 분류한다.
학교 위치, 지역구, 주소, 현재 정보, 기관 정보, 인물, 날짜, 가격처럼 외부 사실 확인이나 최신 정보가 필요한 문항은 추측하지 말고 "검토 불가" 또는 "확인 필요"로 분류한다.
AI의 내부 지식만으로 맞다/틀리다를 단정하지 않는다.
출력은 반드시 JSON 객체 하나만 사용한다. 설명 문장은 JSON 밖에 쓰지 않는다.

상태값은 다음 중 하나만 사용한다: 통과, 확인 필요, 오류 가능성, 표현 모호, 검토 불가.
confidence는 다음 중 하나만 사용한다: 높음, 보통, 낮음. 확신이 없으면 반드시 낮음으로 둔다.
summary는 교사가 참고할 수 있게 한 문장으로 짧게 쓴다. 최종 판정이 아니라 참고 의견임을 전제로 쓴다.
JSON 형식:
{"items":[{"key":"문제키","status":"통과|확인 필요|오류 가능성|표현 모호|검토 불가","summary":"짧은 검토 의견","suggested_answer":"추천 정답 1개 또는 교사 확인 필요 또는 빈 문자열","confidence":"높음|보통|낮음"}]}
"""
    prompt = prompt_instructions + "\n검토할 문제 목록:\n" + json.dumps(payload_items, ensure_ascii=False)

    body = {
        'contents': [{'parts': [{'text': prompt}]}],
        'generationConfig': {'temperature': 0.05, 'response_mime_type': 'application/json'},
    }
    data = json.dumps(body, ensure_ascii=False).encode('utf-8')

    retry_codes = {429, 500, 502, 503, 504}
    last_message = 'AI 검토 API 오류가 발생했습니다. 잠시 후 다시 시도해주세요.'
    last_code: int | None = None
    used_models: list[str] = []

    for model in gemini_model_candidates():
        used_models.append(model)
        url = f'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}'
        for attempt in range(3):
            req = urllib.request.Request(url, data=data, headers={'Content-Type': 'application/json'}, method='POST')
            try:
                with urllib.request.urlopen(req, timeout=70) as resp:
                    raw = resp.read().decode('utf-8', errors='replace')
                parsed = json.loads(raw)
                text = parsed.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                result = parse_json_from_text(text)
                return {
                    'ok': True,
                    'message': 'AI 문제 검토 완료',
                    'items': result.get('items', []) if isinstance(result, dict) else [],
                    'model': model,
                    'used_models': used_models,
                }
            except urllib.error.HTTPError as e:
                detail = e.read().decode('utf-8', errors='replace')[:700]
                last_code = int(e.code)
                last_message = friendly_gemini_error(last_code, detail)
                if last_code in retry_codes and attempt < 2:
                    time.sleep(1.2 * (attempt + 1))
                    continue
                break
            except Exception as e:
                last_message = f'AI 검토 연결 오류: {e}'
                if attempt < 1:
                    time.sleep(1.0)
                    continue
                break

    return {
        'ok': False,
        'error_kind': 'api_unavailable',
        'http_status': last_code,
        'message': last_message,
        'items': [],
        'used_models': used_models,
    }


@app.post('/api/teacher/ai_review')
async def teacher_ai_review(room: str = '') -> JSONResponse:
    token = bind_room(room)
    if token is None:
        return JSONResponse({'ok': False, 'message': '관리할 방을 찾을 수 없습니다.', 'items': []}, status_code=404)
    try:
        items = collect_question_review_items()
        if not items:
            return JSONResponse({'ok': False, 'message': '아직 제출된 학생 문제가 없습니다.', 'items': []}, status_code=400)
        if len(items) > MAX_AI_REVIEW_QUESTIONS:
            items = items[:MAX_AI_REVIEW_QUESTIONS]

        reviews: dict[str, dict[str, Any]] = {}
        ai_items: list[dict[str, Any]] = []
        warnings: list[str] = []
        used_model = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash').strip() or 'gemini-2.5-flash'

        # 검색 없이 판단하기 위험한 최신/외부 사실형 문항은 API에 보내기 전에 교사 확인으로 분류한다.
        # 이렇게 해야 AI가 위치, 주소, 지역구 같은 사실을 그럴듯하게 추측해 오답을 정답처럼 제시하는 일을 줄일 수 있다.
        for item in items:
            if needs_external_fact_check(item.get('question')):
                reviews[item['key']] = make_external_fact_review(item)
            else:
                ai_items.append(item)

        if ai_items:
            item_by_key = {item['key']: item for item in ai_items}
            api_result = await asyncio.to_thread(call_gemini_review, ai_items)
            if api_result.get('ok'):
                used_model = str(api_result.get('model') or used_model)
                for raw in api_result.get('items') or []:
                    if isinstance(raw, dict):
                        cleaned = sanitize_ai_review_item(raw, item_by_key)
                        if cleaned:
                            reviews[cleaned['key']] = cleaned
                # AI가 누락한 문제는 검토 불가로 채워 엑셀/화면에서 빠지지 않도록 한다.
                for key, base in item_by_key.items():
                    if key not in reviews:
                        reviews[key] = make_ai_unavailable_review(base, 'AI 응답에서 이 문제가 누락되었습니다. 교사가 확인해주세요.')
            else:
                message = str(api_result.get('message') or 'AI 검토를 일시적으로 사용할 수 없습니다.')
                warnings.append(message)
                for item in ai_items:
                    reviews[item['key']] = make_ai_unavailable_review(item, message)

        if any(needs_external_fact_check(item.get('question')) for item in items):
            warnings.append('위치·주소·지역구·최신 정보처럼 외부 사실 확인이 필요한 문항은 AI가 단정하지 않고 교사 확인 필요로 표시했습니다.')

        state['ai_reviews'] = reviews
        state['ai_review_meta'] = {
            'reviewed_at': time.strftime('%H:%M:%S'),
            'count': len(reviews),
            'model': used_model,
            'warning': ' / '.join(dict.fromkeys(warnings)) if warnings else '',
        }
        if warnings:
            log(f"AI 문제 검토 완료(일부 확인 필요): {len(reviews)}문항", 'system')
        else:
            log(f"AI 문제 검토 완료: {len(reviews)}문항", 'system')
        await broadcast_state()
        return JSONResponse({'ok': True, 'message': 'AI 문제 검토 완료', 'items': list(reviews.values()), 'meta': state['ai_review_meta']})
    finally:
        unbind_room(token)


def build_questions_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '문제정답정리'

    title = state['settings'].get('room_title') or 'REMAP 방제목'
    mode = '팀전' if state['settings'].get('game_mode') == 'team' else '개인전'

    ws.merge_cells('A1:L1')
    ws['A1'] = f"{title} - 학생 문제/정답 정리"
    ws['A2'] = '게임 모드'
    ws['B2'] = mode
    ws['D2'] = '방 코드'
    ws['E2'] = state['settings'].get('room_code') or '-'
    ws['G2'] = '배틀 문제 수'
    ws['H2'] = state['settings'].get('question_count') or 0

    headers = ['닉네임', '팀', '문제 번호', '문제', '선지 1', '선지 2', '선지 3', '선지 4', '정답', 'AI 검토', 'AI 의견', 'AI 추천답']
    header_row = 4
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Side(style='thin', color='D9E5FF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    players = sorted(
        state['players'].values(),
        key=lambda p: (p.team if p.team else 'Z', p.nickname.lower())
    )
    row = header_row + 1
    for player in players:
        team_label = f"{player.team}팀" if player.team else '-'
        for idx, q in enumerate(player.questions, start=1):
            review = state.get('ai_reviews', {}).get(f'{player.id}:{idx}', {})
            values = [
                player.nickname,
                team_label,
                idx,
                q.text,
                q.choices[0] if len(q.choices) > 0 else '',
                q.choices[1] if len(q.choices) > 1 else '',
                q.choices[2] if len(q.choices) > 2 else '',
                q.choices[3] if len(q.choices) > 3 else '',
                q.choices[q.answer] if 0 <= q.answer < len(q.choices) else '',
                review.get('status', ''),
                review.get('summary', ''),
                review.get('suggested_answer', ''),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='center' if col in (2,3,10) else 'left')
            row += 1

    if row == header_row + 1:
        ws.cell(row=row, column=1, value='아직 제출된 문제가 없습니다.')
        ws.cell(row=row, column=1).font = Font(italic=True, color='5D7AA6')

    ws['A1'].font = Font(size=15, bold=True, color='173B7A')
    ws['A1'].alignment = Alignment(horizontal='center')

    widths = {'A': 14, 'B': 10, 'C': 10, 'D': 38, 'E': 22, 'F': 22, 'G': 22, 'H': 22, 'I': 22, 'J': 14, 'K': 44, 'L': 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@app.get('/api/teacher/export_questions.xlsx')
async def export_questions_xlsx(room: str = '') -> StreamingResponse:
    token = bind_room(room)
    if token is None:
        token = None
    filename = f"remap_questions_{state['settings'].get('room_code') or 'room'}.xlsx"
    payload = build_questions_workbook_bytes()
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    response = StreamingResponse(
        BytesIO(payload),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )
    unbind_room(token)
    return response

@app.websocket('/ws/player')
@app.websocket('//ws/player')
async def ws_player(ws: WebSocket) -> None:
    await ws.accept()
    player_id: str | None = None
    room_token = None
    try:
        while True:
            msg = await ws.receive_json()
            msg_type = msg.get('type')
            if msg_type == 'join':
                room_code = normalize_room_code(msg.get('room_code'))
                if room_token is None:
                    room_token = bind_room(room_code)
                if room_token is None or not state['settings']['room_code'] or room_code != state['settings']['room_code']:
                    await safe_send(ws, {'type': 'error', 'message': '방이 종료되었거나 새 방으로 변경되었습니다. 나가기를 누른 뒤 다시 입장해주세요.'})
                    continue
                if state['game_status'] in ('running', 'countdown'):
                    await safe_send(ws, {'type': 'error', 'message': '이미 시작된 방입니다.'})
                    continue
                pending_id = str(msg.get('pending_id') or '').strip()
                pending = state['pending_participants'].pop(pending_id, None) if pending_id else None
                nickname = clean_text(pending.nickname if pending else msg.get('nickname'), MAX_NICKNAME_LEN)
                raw_team = str(msg.get('team') or '').strip().upper()
                team = raw_team if state['settings'].get('game_mode') == 'team' else ''
                if state['settings'].get('game_mode') == 'team':
                    color = TEAM_COLORS.get(team, TEAM_COLORS['A'])
                else:
                    raw_color = str(msg.get('color') or CHAR_COLORS[0])
                    color = raw_color if raw_color in CHAR_COLORS else CHAR_COLORS[0]
                if not nickname:
                    await safe_send(ws, {'type': 'error', 'message': '닉네임을 입력하세요.'})
                    continue
                if state['settings'].get('game_mode') == 'team' and team not in allowed_teams():
                    await safe_send(ws, {'type': 'error', 'message': '선택할 수 없는 팀입니다.'})
                    continue
                questions_raw = msg.get('questions') or []
                if not isinstance(questions_raw, list):
                    questions_raw = []
                questions: list[Question] = []
                invalid_question_message: str | None = None
                for q_index, item in enumerate(questions_raw, start=1):
                    if not isinstance(item, dict):
                        continue
                    text = clean_text(item.get('text'), MAX_QUESTION_TEXT_LEN)
                    raw_choices = item.get('choices') or []
                    if not isinstance(raw_choices, list):
                        raw_choices = []
                    choices = [clean_text(x, MAX_CHOICE_TEXT_LEN) for x in raw_choices[:4]]
                    blocked = find_blocked_profanity(text) or next((term for choice in choices for term in [find_blocked_profanity(choice)] if term), None)
                    if blocked:
                        invalid_question_message = f'{q_index}번 문제 또는 선택지에 사용할 수 없는 표현이 포함되어 있습니다. 다시 작성해주세요.'
                        break
                    answer = clamp_int(item.get('answer'), 0, 0, 3)
                    if text and len(choices) == 4 and all(choices) and 0 <= answer < 4:
                        questions.append(Question(text=text, choices=choices, answer=answer))
                if invalid_question_message:
                    await safe_send(ws, {'type': 'error', 'message': invalid_question_message})
                    continue
                required_count = int(state['settings']['question_count'])
                if len(questions) != required_count:
                    await safe_send(ws, {'type': 'error', 'message': f'이 방은 문제 {required_count}개를 정확히 입력해야 합니다.'})
                    continue
                x, y = random_spawn()
                player_id = str(uuid.uuid4())
                reconnect_token = str(uuid.uuid4())
                state['players'][player_id] = Player(id=player_id, nickname=nickname, team=team, color=color, x=x, y=y, ws=ws, questions=questions, last_move_at=time.time(), reconnect_token=reconnect_token)
                team_log = f'({team})' if team else ''
                log(f'{nickname}{team_log} 입장')
                await safe_send(ws, {'type': 'joined', 'player_id': player_id, 'reconnect_token': reconnect_token, 'settings': state['settings'], 'room_title': state['settings']['room_title'], 'room_code': state['settings']['room_code']})
                await broadcast_state()
            elif msg_type == 'resume':
                room_code = normalize_room_code(msg.get('room_code'))
                if room_token is None:
                    room_token = bind_room(room_code)
                resume_id = str(msg.get('player_id') or '')
                resume_token = str(msg.get('reconnect_token') or '')
                p = state['players'].get(resume_id)
                if not p or not p.reconnect_token or p.reconnect_token != resume_token:
                    await safe_send(ws, {'type': 'error', 'message': '이전 접속을 복구하지 못했습니다. 새로 입장해주세요.'})
                    continue
                if room_code != state['settings'].get('room_code'):
                    await safe_send(ws, {'type': 'error', 'message': '방 코드가 달라 이전 접속을 복구하지 못했습니다.'})
                    continue
                player_id = p.id
                p.ws = ws
                p.disconnected_at = None
                p.last_move_at = time.time()
                await safe_send(ws, {'type': 'resumed', 'player_id': p.id, 'reconnect_token': p.reconnect_token, 'settings': state['settings'], 'room_title': state['settings']['room_title'], 'room_code': state['settings']['room_code']})
                log(f'{p.nickname} 재접속')
                await broadcast_state()
            elif msg_type == 'move' and player_id:
                p = state['players'].get(player_id)
                if not p or p.state == 'battling' or state['game_status'] == 'finished':
                    continue
                dx = clamp_float(msg.get('dx'), 0.0, -1.0, 1.0)
                dy = clamp_float(msg.get('dy'), 0.0, -1.0, 1.0)
                mag = math.hypot(dx, dy)
                if mag > 1:
                    dx /= mag
                    dy /= mag
                if abs(dx) >= abs(dy) and dx > 0:
                    p.direction = 'right'
                elif abs(dx) >= abs(dy) and dx < 0:
                    p.direction = 'left'
                elif dy > 0:
                    p.direction = 'down'
                elif dy < 0:
                    p.direction = 'up'
                now = time.time()
                last_move_at = float(getattr(p, 'last_move_at', 0.0) or 0.0)
                elapsed = (now - last_move_at) if last_move_at else (1 / 60)
                if elapsed < 0:
                    elapsed = 0
                # Server-authoritative movement for public hosting.
                # player_speed keeps the old 60fps-per-frame feel, but movement is
                # now based on elapsed time so clients cannot move faster by sending
                # more WebSocket packets. Large gaps are capped to prevent jumps.
                elapsed = min(elapsed, 1 / 15)
                p.last_move_at = now
                speed = clamp_float(state['settings'].get('player_speed'), DEFAULTS['player_speed'], 1.0, 20.0)
                distance = speed * 60 * elapsed
                map_width = clamp_int(state['settings'].get('map_width'), DEFAULTS['map_width'], 480, 2400)
                map_height = clamp_int(state['settings'].get('map_height'), DEFAULTS['map_height'], 360, 1600)
                next_x = max(16, min(map_width - 16, p.x + dx * distance))
                next_y = max(16, min(map_height - 16, p.y + dy * distance))
                if not collides_with_walls(next_x, p.y):
                    p.x = next_x
                if not collides_with_walls(p.x, next_y):
                    p.y = next_y
                await try_collisions_for(p)
                if now - float(state.get('last_move_broadcast_at') or 0.0) >= MOVE_BROADCAST_INTERVAL:
                    state['last_move_broadcast_at'] = now
                    await broadcast_state()
            elif msg_type == 'answer' and player_id:
                battle_id = msg.get('battle_id')
                selected = clamp_int(msg.get('selected'), -1, -1, 3)
                time_used = clamp_float(msg.get('time_used'), 0.0, 0.0, float(state['settings'].get('question_time_limit') or DEFAULTS['question_time_limit']))
                battle = state['battles'].get(battle_id)
                if not battle or battle.status != 'active':
                    continue
                side = battle.side_a if battle.side_a.player_id == player_id else battle.side_b
                if side.finished or side.current_index >= len(side.questions):
                    continue
                q = side.questions[side.current_index]
                if selected < 0:
                    continue
                correct = selected == q.answer
                side.answers.append({'selected': selected, 'correct': correct, 'time_used': time_used})
                player = state['players'].get(player_id)
                if player:
                    player.answer_count += 1
                    if correct:
                        player.correct_count += 1
                side.current_index += 1
                await send_to_player(player_id, {'type': 'battle_feedback', 'correct': correct, 'selected': selected, 'timed_out': False})
                await send_battle_question(battle, player_id)
                await broadcast_state()
            elif msg_type == 'leave' and player_id:
                p = state['players'].pop(player_id, None)
                if p:
                    log(f'{p.nickname} 퇴장')
                    await broadcast_state()
                try:
                    await ws.close()
                except Exception:
                    pass
                break
    except (WebSocketDisconnect, OSError, RuntimeError, ConnectionError):
        # Mobile browsers can close WebSocket connections abruptly when rotating,
        # locking the screen, or switching networks. On Windows this can surface
        # as WinError 121 (semaphore timeout) from the websocket stack. Treat it
        # as a normal disconnect instead of printing a scary traceback.
        pass
    finally:
        if player_id and player_id in state['players']:
            p = state['players'][player_id]
            if p.ws is ws:
                p.ws = None
                p.disconnected_at = time.time()
                log(f"{p.nickname} 연결 일시 끊김 - 재접속 대기")
                await broadcast_state()
        unbind_room(room_token)


@app.websocket('/ws/teacher')
@app.websocket('//ws/teacher')
async def ws_teacher(ws: WebSocket) -> None:
    room_code = normalize_room_code(ws.query_params.get('room'))
    room_token = bind_room(room_code) if room_code else None
    await ws.accept()
    state['teacher_clients'].add(ws)
    try:
        await safe_send(ws, full_payload())
        while True:
            await ws.receive_text()
    except (WebSocketDisconnect, OSError, RuntimeError, ConnectionError):
        pass
    finally:
        state['teacher_clients'].discard(ws)
        unbind_room(room_token)




TEACHER_CEREMONY_HTML = """
<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>ReMap 시상식</title>
<style>
:root{--bg0:#061326;--bg1:#0b2442;--card:#eef7ff;--text:#0f2545}
*{box-sizing:border-box}html,body{margin:0;min-height:100%;font-family:Arial,'Noto Sans KR',sans-serif;background:radial-gradient(circle at 30% 0,rgba(56,189,248,.18),transparent 34%),linear-gradient(135deg,#061326,#0b2442 55%,#141b4c);color:#eff8ff}
.topbar{height:64px;display:flex;align-items:center;justify-content:center;position:sticky;top:0;z-index:5;background:rgba(4,14,28,.88);border-bottom:1px solid rgba(125,211,252,.22);backdrop-filter:blur(10px)}.brand{position:absolute;left:18px;font-weight:1000;font-size:24px;background:linear-gradient(135deg,#38bdf8,#8b5cf6);-webkit-background-clip:text;color:transparent}.title{font-size:38px;font-weight:1000;letter-spacing:-1px;text-shadow:0 0 20px rgba(255,255,255,.42)}.close{position:absolute;right:18px;border:0;border-radius:999px;padding:10px 18px;font-weight:1000;color:#0f2545;background:linear-gradient(180deg,#fff,#dbeafe);cursor:pointer}
.wrap{width:min(1180px,96vw);margin:18px auto 34px;background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(232,242,255,.96));border:2px solid rgba(15,31,58,.65);border-radius:28px;color:#0f2545;box-shadow:0 28px 90px rgba(0,0,0,.32);padding:22px;overflow:hidden}.kicker{display:inline-flex;gap:8px;align-items:center;padding:8px 18px;border-radius:999px;background:#071b33;color:#fde68a;border:2px solid #fbbf24;font-size:13px;font-weight:1000}.head{text-align:center}.head h1{margin:12px 0 5px;font-size:clamp(34px,5vw,58px);letter-spacing:-1px}.sub{color:#52708f;font-weight:800}.banner{display:flex;gap:14px;align-items:center;margin:18px 0 12px;padding:16px;border-radius:22px;background:linear-gradient(135deg,#dff6ff,#dfe7ff);border:1px solid rgba(91,141,204,.28)}.bannerIcon{font-size:34px}.banner strong{display:block;font-size:24px;color:#0f2545}.grid{display:grid;grid-template-columns:minmax(0,1.25fr) minmax(320px,.75fr);gap:14px}.panel{background:linear-gradient(180deg,#10223c,#07182d);color:#eef7ff;border-radius:24px;padding:16px;border:1px solid rgba(191,219,254,.22);box-shadow:0 18px 48px rgba(15,31,58,.18)}.panel h2{margin:0 0 12px;font-size:20px}.podium{display:grid;gap:10px}.podiumItem{position:relative;display:grid;grid-template-columns:68px minmax(0,1fr) 88px;align-items:center;gap:12px;padding:14px;border-radius:18px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12)}.podiumItem.rank1{background:linear-gradient(135deg,rgba(251,191,36,.28),rgba(255,255,255,.08));border-color:rgba(251,191,36,.45)}.avatar{width:58px;height:58px;object-fit:contain;filter:drop-shadow(0 10px 16px rgba(0,0,0,.22))}.rankBadge{display:inline-flex;align-items:center;gap:6px;padding:4px 10px;border-radius:999px;background:#e2e8f0;color:#0f2545;font-weight:1000;font-size:12px}.rank1 .rankBadge{background:linear-gradient(135deg,#fde68a,#f59e0b)}.name{display:block;margin-top:6px;color:#fff;font-size:21px;font-weight:1000;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.meta{color:#bad0ec;font-size:12px;margin-top:3px;font-weight:800}.score{text-align:right;color:#fbbf24;font-size:24px;font-weight:1000}.list{display:grid;gap:8px}.item{display:flex;justify-content:space-between;gap:10px;align-items:center;padding:10px 12px;border-radius:14px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.10);color:#eef7ff}.item strong{color:#fbbf24}.mini{font-size:12px;color:#aecaec}.mvp{display:grid;gap:9px}.mvpCard{display:flex;align-items:center;gap:12px;padding:12px;border-radius:16px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.10)}.mvpIcon{width:42px;height:42px;border-radius:14px;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,rgba(56,189,248,.25),rgba(139,92,246,.20));font-size:22px}.log{max-height:190px;overflow:auto;display:grid;gap:6px}.empty{padding:28px;text-align:center;color:#aecaec;border:1px dashed rgba(255,255,255,.15);border-radius:18px}.foot{display:flex;justify-content:center;margin-top:16px}.home{border:0;border-radius:16px;padding:13px 34px;font-weight:1000;font-size:16px;background:linear-gradient(135deg,#38bdf8,#7c3aed);color:#fff;cursor:pointer}
@media(max-width:760px){.topbar{height:56px}.title{font-size:26px}.wrap{margin:8px auto 18px;padding:14px;border-radius:20px}.grid{grid-template-columns:1fr}.head h1{font-size:30px}.podiumItem{grid-template-columns:58px minmax(0,1fr) 64px;padding:10px}.avatar{width:50px;height:50px}.name{font-size:16px}.score{font-size:18px}.close{right:8px;padding:8px 12px}.brand{left:10px;font-size:18px}.banner strong{font-size:18px}}

/* ===== v3.26 teacher credit footer patch ===== */
@media(max-width:1180px){
  #createScreen .teacherCredit{
    text-align:center!important;
    margin-top:12px!important;
    padding:10px 12px!important;
    border-radius:14px!important;
    background:rgba(255,255,255,.50)!important;
    border:1px solid rgba(91,141,204,.18)!important;
    color:#35516f!important;
    font-size:12px!important;
  }
  #createScreen .teacherCredit div{white-space:normal!important;}
}
@media(min-width:1181px){
  #createScreen .teacherCredit{
    max-width:520px;
    margin-left:auto;
    padding-right:4px;
  }
}

.aiReviewPanel{border-color:rgba(34,211,238,.30)!important;background:linear-gradient(180deg,rgba(14,165,233,.11),rgba(255,255,255,.052))!important}
.aiReviewActions{display:flex;gap:8px;align-items:center;margin:8px 0 10px}.aiReviewActions button{width:100%;min-height:40px;padding:8px 10px;font-size:14px}.aiReviewBox{display:grid;gap:8px;max-height:220px;overflow:auto;padding-right:3px}.aiReviewSummary{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:4px}.aiPill{display:inline-flex;align-items:center;gap:4px;border-radius:999px;padding:5px 8px;font-size:11px;font-weight:1000;border:1px solid rgba(255,255,255,.14);background:rgba(255,255,255,.08)}.aiPill.ok{color:#dcfce7;background:rgba(22,163,74,.20)}.aiPill.warn{color:#fef3c7;background:rgba(245,158,11,.22)}.aiPill.bad{color:#fee2e2;background:rgba(239,68,68,.20)}.aiReviewItem{border:1px solid rgba(255,255,255,.10);background:rgba(255,255,255,.065);border-radius:13px;padding:8px}.aiReviewItem b{display:block;color:#0f2544;font-size:13px;margin-bottom:4px}.aiReviewItem .mini{font-size:11px}.aiStatus{display:inline-block;border-radius:999px;padding:3px 7px;margin-right:5px;font-size:11px;font-weight:1000}.aiStatus.ok{background:rgba(34,197,94,.20);color:#dcfce7}.aiStatus.warn{background:rgba(245,158,11,.20);color:#fef3c7}.aiStatus.bad{background:rgba(239,68,68,.22);color:#fee2e2}.aiStatus.muted{background:rgba(148,163,184,.18);color:#e2e8f0}.aiReviewWarning{border:1px solid rgba(245,158,11,.28);border-radius:12px;padding:7px 8px;background:rgba(245,158,11,.12);color:#78350f;font-weight:900}
#aiReviewBtn{background:linear-gradient(135deg,#22c55e,#0891b2 60%,#2563eb)!important;color:#fff!important;box-shadow:0 12px 24px rgba(14,165,233,.20)!important}
@media(max-width:1180px){.aiReviewBox{max-height:180px}.aiReviewActions{margin-top:6px}}


.winnerCrown{position:absolute;left:50%;top:-20px;transform:translateX(-50%) rotate(-7deg);z-index:6;font-size:34px;filter:drop-shadow(0 7px 9px rgba(0,0,0,.28)) drop-shadow(0 0 10px rgba(251,191,36,.52));animation:crownFloat 1.8s ease-in-out infinite alternate;pointer-events:none}
@keyframes crownFloat{from{transform:translateX(-50%) translateY(0) rotate(-7deg)}to{transform:translateX(-50%) translateY(-3px) rotate(6deg)}}
@media(max-width:760px){.winnerCrown{font-size:26px;top:-15px}}


/* v3.32 teacher mobile AI review visibility */
@media(max-width:1180px){
  #operateScreen .rightCol{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;overflow:visible!important;padding-right:0!important;}
  #operateScreen .rightCol .aiReviewPanel{order:99!important;width:100%!important;display:block!important;margin-bottom:10px!important;}
  #operateScreen .aiReviewBox{max-height:none!important;overflow:visible!important;}
}
@media(max-width:820px) and (orientation:landscape){
  #operateScreen .rightCol{grid-template-columns:1fr 1fr!important;align-items:start!important;}
  #operateScreen .rightCol .aiReviewPanel{grid-column:1 / -1!important;}
  #operateScreen .rightCol .aiReviewBox{max-height:160px!important;overflow:auto!important;}
}


/* ===== v3.34 student info compact grid + teacher setup 3-row patch ===== */
/* Student prep: title stays wide, the six small room/nickname fields form 3 columns x 2 rows on phones. */
#prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:1 / -1;}
#prepScreen .studentRoomInfoGrid .compactInfoCard .value{word-break:keep-all;}
@media (orientation:portrait) and (pointer:coarse), (max-width:760px){
  #prepScreen .studentRoomInfoGrid{
    grid-template-columns:repeat(3,minmax(0,1fr))!important;
    gap:8px!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard{
    grid-column:1 / -1!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard{
    min-height:76px!important;
    padding:10px 8px!important;
    display:flex!important;
    flex-direction:column!important;
    justify-content:center!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .label{
    font-size:clamp(10px,2.9vw,12px)!important;
    line-height:1.15!important;
    margin-bottom:6px!important;
    white-space:normal!important;
    letter-spacing:-.45px!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{
    font-size:clamp(18px,5vw,28px)!important;
    line-height:1.05!important;
    letter-spacing:-1.2px!important;
    white-space:nowrap!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard .value{
    font-size:clamp(21px,6.2vw,34px)!important;
  }
}
@media (orientation:portrait) and (pointer:coarse) and (max-width:380px){
  #prepScreen .studentRoomInfoGrid{gap:6px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard{padding:8px 6px!important;min-height:68px!important;border-radius:14px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{font-size:clamp(16px,4.7vw,22px)!important;}
}
/* Student landscape keeps the previous compact side panel to avoid squeezing the question editor. */
@media (orientation:landscape) and (pointer:coarse){
  #prepScreen .studentRoomInfoGrid{grid-template-columns:1fr!important;}
  #prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:auto!important;}
}
/* Teacher PC setup: compact numeric settings become 4 columns, so 10 fields render as 3 rows instead of 5. */
@media (min-width:821px){
  #createScreen .teacherSettingsGrid{
    grid-template-columns:repeat(4,minmax(0,1fr))!important;
    gap:10px!important;
  }
  #createScreen .teacherSettingsGrid .field input{
    min-height:40px!important;
    padding:8px 10px!important;
  }
}

</style>
</head>
<body>
<div class="topbar"><div class="brand">ReMap</div><div class="title">REMAP</div><button class="close" onclick="window.close()">닫기</button></div>
<main class="wrap">
  <section class="head"><div class="kicker">🏆 FINAL CEREMONY</div><h1>최종 결과 시상식</h1><div class="sub">교사용 시상식 화면입니다. 학생들과 함께 결과를 확인하세요.</div></section>
  <div id="banner" class="banner"></div>
  <section class="grid">
    <div class="panel"><h2>개인 시상대</h2><div id="podium" class="podium"></div><h2 style="margin-top:18px">전체 개인 순위</h2><div id="fullRank" class="list"></div></div>
    <div class="panel"><h2>MVP</h2><div id="mvp" class="mvp"></div><h2 style="margin-top:18px">팀 순위</h2><div id="teamRank" class="list"></div><h2 style="margin-top:18px">배틀 로그</h2><div id="logs" class="log"></div></div>
  </section>
  <div class="foot"><button class="home" onclick="window.close()">닫기</button></div>
</main>
<script>
const payload=__PAYLOAD__;
function escapeHtml(v){return String(v??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));}
function charUrl(c){return '/character/'+encodeURIComponent(String(c||'#60a5fa').replace('#',''))+'.svg';}
function medal(r){return r===1?'🥇':r===2?'🥈':r===3?'🥉':'🏅';}
const ranks=payload.rankings||[], stats=payload.player_stats||[], teams=payload.team_rankings||[], logs=payload.logs||[];
const byRank={};ranks.slice(0,3).forEach(r=>byRank[Number(r.rank)||0]=r);
document.getElementById('banner').innerHTML=payload.winner_team?`<div class="bannerIcon">🏆</div><div><span class="mini">팀전 우승</span><strong>${escapeHtml(payload.winner_team.team)}팀</strong><span>${Number(payload.winner_team.score||0)}점으로 우승!</span></div>`:`<div class="bannerIcon">🎮</div><div><span class="mini">게임 모드</span><strong>개인전 결과</strong><span>개인 순위와 MVP를 확인합니다.</span></div>`;
function podiumSlot(p,rank){return p?`<div class="podiumItem rank${rank}">${rank===1?'<div class="winnerCrown" aria-hidden="true">👑</div>':''}<img class="avatar" src="${charUrl(p.color)}" alt=""><div><span class="rankBadge">${medal(rank)} ${rank}위</span><span class="name">${escapeHtml(p.nickname)}</span><div class="meta">${p.team?escapeHtml(p.team)+'팀 · ':''}정답 ${Number(p.correct_count||0)}개 · 배틀 ${Number(p.battles_played||0)}회</div></div><div class="score">${Number(p.score||0)}점</div></div>`:`<div class="podiumItem rank${rank}"><div></div><div><span class="rankBadge">${rank}위</span><span class="name">참가자 없음</span><div class="meta">기록 없음</div></div><div class="score">-</div></div>`;}
document.getElementById('podium').innerHTML=ranks.length?[1,2,3].map(r=>podiumSlot(byRank[r],r)).join(''):'<div class="empty">아직 결과가 없습니다.</div>';
document.getElementById('fullRank').innerHTML=stats.length?stats.map((p,i)=>`<div class="item"><span>${i+1}. ${escapeHtml(p.nickname)}${p.team?` <span class="mini">(${escapeHtml(p.team)})</span>`:''}<br><span class="mini">정답 ${Number(p.correct_count||0)}/${Number(p.answer_count||0)} · 배틀 ${Number(p.battles_played||0)}회</span></span><strong>${Number(p.score||0)}점</strong></div>`).join(''):'<div class="empty">참가자 기록이 없습니다.</div>';
const mvp=ranks[0]||null,best=payload.best_correct||null,most=payload.most_battles||null;
document.getElementById('mvp').innerHTML=`<div class="mvpCard"><div class="mvpIcon">👑</div><div><b>${mvp?escapeHtml(mvp.nickname):'-'}</b><div class="mini">최고 점수 MVP · ${mvp?Number(mvp.score||0):0}점</div></div></div><div class="mvpCard"><div class="mvpIcon">✅</div><div><b>${best?escapeHtml(best.nickname):'-'}</b><div class="mini">가장 많이 맞힌 학생 · ${best?Number(best.correct_count||0):0}개</div></div></div><div class="mvpCard"><div class="mvpIcon">⚔️</div><div><b>${most?escapeHtml(most.nickname):'-'}</b><div class="mini">가장 많은 배틀 · ${most?Number(most.battles_played||0):0}회</div></div></div>`;
document.getElementById('teamRank').innerHTML=teams.length?teams.map(t=>`<div class="item"><span>${Number(t.rank||0)}. ${escapeHtml(t.team)}팀</span><strong>${Number(t.score||0)}점</strong></div>`).join(''):'<div class="empty">개인전 모드</div>';
document.getElementById('logs').innerHTML=logs.length?logs.slice().reverse().map(l=>`<div class="item"><span>${escapeHtml(l.time||'')}</span><span>${escapeHtml(l.message||'')}</span></div>`).join(''):'<div class="empty">배틀 기록 없음</div>';
</script>
</body>
</html>
"""

STUDENT_HTML = """
<!doctype html>
<html lang='ko'>
<head>
<meta charset='utf-8' />
<meta name='viewport' content='width=device-width, initial-scale=1, viewport-fit=cover' />
<title>ReMap 코어</title>
<style>
:root{
  --bg0:#06111f;
  --bg1:#071b35;
  --bg2:#0b2852;
  --card:rgba(15,31,58,.74);
  --card2:rgba(12,26,48,.86);
  --line:rgba(148,197,255,.22);
  --line2:rgba(255,255,255,.12);
  --text:#eef7ff;
  --muted:#9fb7d5;
  --blue:#38bdf8;
  --blue2:#2563eb;
  --cyan:#22d3ee;
  --violet:#8b5cf6;
  --green:#22c55e;
  --red:#ef4444;
  --amber:#f59e0b;
}
html,body{margin:0;height:100%;font-family:Arial,"Malgun Gothic",sans-serif;background:var(--bg0);color:var(--text);overflow:hidden}*{box-sizing:border-box}
body{display:flex;flex-direction:column;background:
  radial-gradient(circle at 18% 8%,rgba(56,189,248,.22),transparent 32%),
  radial-gradient(circle at 82% 16%,rgba(139,92,246,.20),transparent 30%),
  radial-gradient(circle at 55% 92%,rgba(34,211,238,.13),transparent 36%),
  linear-gradient(135deg,var(--bg0),var(--bg1) 48%,var(--bg2));
}
button,input,textarea,select{font:inherit;border-radius:14px;border:1px solid var(--line);padding:10px 12px;background:rgba(255,255,255,.08);color:var(--text);outline:none}
textarea{min-height:72px;resize:vertical}select option{color:#111827;background:#fff}input::placeholder,textarea::placeholder{color:rgba(223,244,255,.52)}
input:focus,textarea:focus,select:focus{border-color:rgba(56,189,248,.72);box-shadow:0 0 0 3px rgba(56,189,248,.14)}
button{cursor:pointer;border:none;color:#fff;font-weight:900;background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed);box-shadow:0 14px 30px rgba(37,99,235,.22);transition:.16s ease transform,.16s ease filter,.16s ease box-shadow}
button:hover{transform:translateY(-1px);filter:brightness(1.07);box-shadow:0 18px 36px rgba(56,189,248,.25)}button:active{transform:translateY(0)}button:disabled{opacity:.58;cursor:not-allowed;transform:none;filter:none}
button.secondary{background:linear-gradient(135deg,#22d3ee,#3b82f6)}button.ghost{background:rgba(255,255,255,.08);color:#dff4ff;border:1px solid var(--line);box-shadow:none}
.topbar{height:64px;background:rgba(6,17,31,.78);border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:0 18px;gap:14px;backdrop-filter:blur(16px);box-shadow:0 12px 34px rgba(0,0,0,.22);z-index:5;position:relative}
.brand{font-weight:1000;font-size:20px;letter-spacing:.2px;background:linear-gradient(135deg,#eff6ff,#38bdf8 45%,#a78bfa);-webkit-background-clip:text;color:transparent;text-shadow:0 0 24px rgba(56,189,248,.14)}
.roomTitle{position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);width:min(720px,58vw);font-weight:1000;text-align:center;font-size:32px;line-height:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:#ffffff!important;text-shadow:0 3px 12px rgba(255,255,255,.38),0 0 22px rgba(255,255,255,.22);letter-spacing:.8px;pointer-events:none;}
.statusBadge{display:inline-flex;align-items:center;gap:6px;padding:8px 13px;border-radius:999px;background:linear-gradient(135deg,rgba(34,211,238,.20),rgba(59,130,246,.22));border:1px solid rgba(125,211,252,.38);color:#dff9ff;font-size:12px;font-weight:1000;box-shadow:0 0 24px rgba(34,211,238,.12)}
.statusBadge::before{content:"";width:8px;height:8px;border-radius:999px;background:#22c55e;box-shadow:0 0 14px rgba(34,197,94,.85)}
.statusBadge.status-running{background:linear-gradient(135deg,#16a34a,#22c55e)!important;border-color:rgba(187,247,208,.88)!important;color:#ffffff!important;box-shadow:0 0 28px rgba(34,197,94,.36),0 10px 24px rgba(22,163,74,.22)!important}
.statusBadge.status-running::before{background:#dcfce7!important;box-shadow:0 0 16px rgba(220,252,231,.95)!important}
.statusBadge.status-finished{background:linear-gradient(135deg,#ef4444,#b91c1c)!important;border-color:rgba(254,202,202,.88)!important;color:#ffffff!important;box-shadow:0 0 28px rgba(239,68,68,.36),0 10px 24px rgba(185,28,28,.22)!important}
.statusBadge.status-finished::before{background:#fee2e2!important;box-shadow:0 0 16px rgba(254,226,226,.95)!important}
.statusBadge.status-waiting{background:linear-gradient(135deg,#facc15,#f59e0b)!important;border-color:rgba(254,240,138,.92)!important;color:#172033!important;box-shadow:0 0 28px rgba(245,158,11,.34),0 10px 24px rgba(180,83,9,.18)!important}
.statusBadge.status-waiting::before{background:#14532d!important;box-shadow:0 0 12px rgba(20,83,45,.55)!important}
.topLeaveBtn{display:none;align-items:center;justify-content:center;height:32px;padding:0 12px;border-radius:999px;background:linear-gradient(135deg,#eef6ff,#dbeafe);color:#12315a;font-weight:1000;font-size:12px;border:1px solid rgba(125,211,252,.45);box-shadow:0 8px 20px rgba(15,31,58,.16);white-space:nowrap}
#joinScreen,#prepScreen,#gameScreen,#endScreen{width:100%;height:calc(100vh - 64px)}#prepScreen,#gameScreen,#endScreen{display:none}
#joinScreen,#prepScreen{justify-content:center;padding:20px;overflow:auto;background:
  radial-gradient(circle at 18% 18%,rgba(56,189,248,.16),transparent 30%),
  radial-gradient(circle at 88% 10%,rgba(139,92,246,.12),transparent 28%);
}
#joinScreen{display:flex;align-items:center}#prepScreen{display:none;align-items:flex-start;padding-top:28px;padding-bottom:28px}.card{width:min(920px,95vw);background:linear-gradient(180deg,rgba(15,31,58,.86),rgba(10,24,45,.88));border:1px solid var(--line);border-radius:28px;padding:24px;box-shadow:0 24px 70px rgba(0,0,0,.30),inset 0 1px 0 rgba(255,255,255,.08);position:relative;overflow:hidden}.card::before{content:"";position:absolute;inset:-1px;background:radial-gradient(circle at 12% 0,rgba(56,189,248,.18),transparent 28%),radial-gradient(circle at 88% 18%,rgba(236,72,153,.10),transparent 25%);pointer-events:none}.card>*{position:relative}.card h1{color:#fff!important;line-height:1.25;padding-top:2px}.row{display:flex;gap:10px;flex-wrap:wrap}.stack{display:flex;flex-direction:column;gap:8px}.mini{font-size:12px;color:var(--muted);line-height:1.45}.badge{padding:5px 9px;border-radius:999px;background:rgba(34,211,238,.15);border:1px solid rgba(125,211,252,.28);font-size:12px;color:#bdefff;font-weight:900}
.panel{background:linear-gradient(180deg,rgba(255,255,255,.095),rgba(255,255,255,.052));border:1px solid var(--line);border-radius:22px;padding:14px;box-shadow:0 18px 44px rgba(0,0,0,.20);backdrop-filter:blur(14px);margin-bottom:10px}.panel h3{margin:0 0 10px 0;font-size:16px;color:#f4fbff}.panelTitle{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-bottom:8px}
#questionsList{max-height:320px;overflow:auto;padding-right:6px}#questionsList .qitem{padding:10px;margin:8px 0;background:rgba(255,255,255,.06);border:1px solid var(--line);border-radius:16px}.teamWrap,.colorWrap{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}.teamBtn,.colorBtn{padding:9px 12px;border-radius:14px;border:1px solid var(--line);background:rgba(255,255,255,.08);cursor:pointer;font-weight:900;color:#dff4ff;box-shadow:none}.teamBtn.active{background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed);color:#fff;border-color:rgba(125,211,252,.65);box-shadow:0 16px 34px rgba(37,99,235,.24)}.colorBtn{width:32px;height:32px;padding:0;border-radius:999px}.colorBtn.active{outline:3px solid rgba(125,211,252,.72);outline-offset:2px}
.infoGrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}.infoCard{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.11);border-radius:18px;padding:10px}.label{font-size:12px;color:#b8cce8;margin-bottom:5px;font-weight:800}.value{font-weight:900;color:#fff}
#gameScreen{grid-template-columns:280px minmax(0,1fr) 300px;gap:12px;padding:12px;height:calc(100vh - 64px);overflow:hidden}.studentCol{min-height:0;overflow:auto;padding-right:2px}.studentMain{min-width:0;min-height:0;display:flex}.studentMapPanel{width:100%;min-width:0;min-height:0;display:flex;flex-direction:column;padding:15px}.mapHeader{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;gap:10px}.mapTitleWrap h3{margin:0;color:#fff}.mapTitleWrap .mini{margin-top:3px}#mapWrap{flex:1;min-height:0;display:flex;justify-content:center;align-items:center;background:linear-gradient(180deg,rgba(10,28,54,.74),rgba(8,20,38,.80));border:1px solid rgba(255,255,255,.09);border-radius:20px;padding:10px;position:relative;overflow:hidden}#mapWrap::before{content:"";position:absolute;inset:0;background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.10),transparent 45%);pointer-events:none}
canvas{position:relative;max-width:100%;max-height:100%;border:2px solid rgba(125,211,252,.38);border-radius:20px;background:#d5e4f7;box-shadow:0 26px 60px rgba(0,0,0,.32),0 0 0 8px rgba(255,255,255,.035)}
#gameCanvas{transform:scale(var(--student-map-scale,1));transform-origin:center center;transition:transform .12s ease}
.mapZoomControls{position:absolute;right:12px;bottom:12px;z-index:18;display:none;align-items:center;gap:6px;padding:6px;border-radius:999px;background:rgba(238,247,255,.88);border:1px solid rgba(125,211,252,.45);box-shadow:0 10px 24px rgba(15,31,58,.20);backdrop-filter:blur(10px)}
.mapZoomControls button{width:34px;height:30px;min-height:30px;padding:0;border-radius:999px;background:linear-gradient(135deg,#eef6ff,#dbeafe);color:#12315a;font-weight:1000;box-shadow:none}
.mapZoomControls span{min-width:44px;text-align:center;color:#12315a;font-weight:1000;font-size:12px}
#topInfo{position:static;width:100%;z-index:auto;background:transparent;backdrop-filter:none;border:none;border-radius:0;padding:0;box-shadow:none}.statHero{padding:12px;border-radius:18px;background:linear-gradient(135deg,rgba(34,211,238,.14),rgba(99,102,241,.16));border:1px solid rgba(125,211,252,.22);margin-bottom:8px}.statHero strong{display:block;margin-top:4px;font-size:18px;color:#fff}.rankItem,.battleItem,.logItem{display:flex;justify-content:space-between;gap:10px;align-items:center;padding:8px 0;border-bottom:1px solid rgba(255,255,255,.08);font-size:14px;color:#dceeff}.rankItem:last-child,.battleItem:last-child,.logItem:last-child{border-bottom:none}.rankItem span:first-child,.battleItem span:first-child,.logItem span:first-child{color:#b7c9e2}.rankItem strong,.battleItem strong{color:#fff}.teamPill{display:inline-flex;align-items:center;padding:5px 9px;border-radius:999px;background:rgba(34,197,94,.14);border:1px solid rgba(34,197,94,.28);color:#bbf7d0;font-size:12px;font-weight:900}.leaveWrap{padding-top:4px}#leaveBtn{width:100%}
#battleOverlay{position:fixed;inset:0;background:radial-gradient(circle at 50% 42%,rgba(56,189,248,.24),transparent 34%),radial-gradient(circle at 20% 20%,rgba(139,92,246,.22),transparent 28%),rgba(3,7,18,.76);display:none;align-items:center;justify-content:center;z-index:20;padding:16px;backdrop-filter:blur(10px)}#battleCard{width:min(760px,94vw);background:linear-gradient(180deg,rgba(15,31,58,.96),rgba(8,18,35,.98));border:1px solid rgba(125,211,252,.36);border-radius:28px;padding:22px;box-shadow:0 30px 90px rgba(0,0,0,.46),0 0 0 1px rgba(255,255,255,.06),inset 0 1px 0 rgba(255,255,255,.10);position:relative;overflow:hidden;animation:battleCardIn .26s ease-out}#battleCard::before{content:"";position:absolute;inset:-2px;background:radial-gradient(circle at 18% 0,rgba(56,189,248,.20),transparent 28%),radial-gradient(circle at 90% 12%,rgba(236,72,153,.14),transparent 28%);pointer-events:none}#battleCard>*{position:relative}#battleCard.intro{width:min(680px,92vw);text-align:center}.battleVsStage{padding:36px 0 22px;text-align:center}.battleVsText{font-size:clamp(34px,7vw,72px);font-weight:1000;line-height:1.08;letter-spacing:-1px;background:linear-gradient(135deg,#fff,#38bdf8 45%,#fbbf24);-webkit-background-clip:text;color:transparent;text-shadow:0 18px 45px rgba(56,189,248,.22);animation:vsPop .62s cubic-bezier(.17,.89,.29,1.28)}.battleVsSub{margin-top:12px;color:#dff4ff;font-weight:900}.choice{display:block;width:100%;text-align:left;margin:10px 0;padding:14px 15px;background:linear-gradient(135deg,rgba(255,255,255,.09),rgba(255,255,255,.045));border-radius:16px;border:1px solid rgba(125,211,252,.20);color:#eef7ff;box-shadow:0 10px 24px rgba(0,0,0,.14);transition:.16s ease transform,.16s ease background,.16s ease border-color,.16s ease box-shadow;position:relative;overflow:hidden}.choice:hover{transform:translateY(-2px);background:linear-gradient(135deg,rgba(56,189,248,.26),rgba(99,102,241,.18));border-color:rgba(125,211,252,.58);box-shadow:0 16px 34px rgba(56,189,248,.18)}.choice:active{transform:scale(.985)}.choice.selected{border-color:rgba(250,204,21,.78);background:linear-gradient(135deg,rgba(250,204,21,.26),rgba(245,158,11,.16))}.choice.correct{background:linear-gradient(135deg,rgba(34,197,94,.94),rgba(22,163,74,.86));border-color:rgba(187,247,208,.82);color:#fff;animation:choiceGood .45s ease}.choice.wrong{background:linear-gradient(135deg,rgba(239,68,68,.96),rgba(185,28,28,.88));border-color:rgba(254,202,202,.74);color:#fff;animation:choiceBad .42s ease}.choice.locked{pointer-events:none;opacity:.82}.battleHeader{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:10px}.battleTitleText{margin:0;font-size:24px;color:#f8fbff}.battleTimerPill{display:inline-flex;align-items:center;justify-content:center;min-width:112px;padding:9px 12px;border-radius:999px;background:linear-gradient(135deg,rgba(14,165,233,.26),rgba(99,102,241,.24));border:1px solid rgba(125,211,252,.32);font-weight:1000;color:#e0f7ff}.battleQuestionBox{font-size:22px;line-height:1.45;margin:12px 0 16px;color:#f4fbff;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.10);border-radius:18px;padding:16px}.battleWaiting{font-size:18px;text-align:center;color:#dbeafe;padding:18px 0}.battleProgressBar{height:8px;background:rgba(255,255,255,.09);border-radius:999px;overflow:hidden;margin:8px 0 12px}.battleProgressBar span{display:block;height:100%;background:linear-gradient(90deg,#38bdf8,#8b5cf6);border-radius:999px}#battleEffect{position:fixed;top:50%;left:50%;transform:translate(-50%,-50%) scale(.86);display:none;z-index:35;padding:22px 38px;border-radius:26px;color:#fff;font-size:clamp(36px,8vw,78px);font-weight:1000;text-align:center;letter-spacing:-1px;box-shadow:0 28px 80px rgba(0,0,0,.38);pointer-events:none;text-shadow:0 6px 22px rgba(0,0,0,.28)}#battleEffect.success{display:block;background:linear-gradient(135deg,#22c55e,#16a34a);animation:effectPop .68s ease-out}#battleEffect.error{display:block;background:linear-gradient(135deg,#ef4444,#b91c1c);animation:effectShake .68s ease-out}#toast{position:fixed;left:50%;padding:10px 16px;border-radius:14px;background:#0f172a;color:#fff;display:none;z-index:40;box-shadow:0 10px 30px rgba(15,23,42,.25);pointer-events:none;text-align:center}#toast.bottom{bottom:16px;transform:translateX(-50%)}#toast.center{top:50%;transform:translate(-50%,-50%);font-size:30px;font-weight:800;padding:18px 30px;background:rgba(15,23,42,.92)}#toast.feedback{top:17%;transform:translateX(-50%);font-size:34px;font-weight:900;padding:18px 34px;min-width:220px;background:rgba(15,23,42,.94);border:2px solid rgba(255,255,255,.18);text-shadow:0 2px 8px rgba(15,23,42,.22)}#toast.success{background:rgba(22,163,74,.96)}#toast.error{background:rgba(220,38,38,.96)}#countdownOverlay{position:fixed;inset:0;background:radial-gradient(circle at 50% 45%,rgba(56,189,248,.25),transparent 35%),rgba(3,7,18,.62);display:none;align-items:center;justify-content:center;z-index:28;pointer-events:none;backdrop-filter:blur(6px)}#countdownText{font-size:min(20vw,142px);font-weight:1000;color:#fff;text-shadow:0 16px 50px rgba(56,189,248,.35),0 2px 0 rgba(15,23,42,.26);letter-spacing:2px;animation:countPulse .82s ease-out}#countdownText.start{font-size:min(15vw,112px);background:linear-gradient(135deg,#fff,#38bdf8,#fbbf24);-webkit-background-clip:text;color:transparent}#resultOverlay{position:fixed;inset:0;background:radial-gradient(circle at 50% 42%,rgba(56,189,248,.20),transparent 34%),rgba(3,7,18,.66);display:none;align-items:center;justify-content:center;z-index:25;backdrop-filter:blur(9px)}.resultCard{width:min(520px,92vw);background:linear-gradient(180deg,rgba(15,31,58,.96),rgba(8,18,35,.98));border:1px solid rgba(125,211,252,.36);border-radius:28px;padding:28px 24px;text-align:center;box-shadow:0 30px 90px rgba(0,0,0,.46),inset 0 1px 0 rgba(255,255,255,.10);animation:resultRise .32s ease-out}.resultCard.win{border-color:rgba(187,247,208,.62);box-shadow:0 30px 90px rgba(0,0,0,.46),0 0 42px rgba(34,197,94,.18)}.resultCard.lose{border-color:rgba(254,202,202,.55);box-shadow:0 30px 90px rgba(0,0,0,.46),0 0 42px rgba(239,68,68,.15)}.resultCard.draw{border-color:rgba(191,219,254,.58)}#resultTitle{font-size:clamp(44px,9vw,78px)!important;margin:0;background:linear-gradient(135deg,#fff,#38bdf8);-webkit-background-clip:text;color:transparent;animation:resultTitlePop .6s cubic-bezier(.17,.89,.29,1.28)}.resultCard.win #resultTitle{background:linear-gradient(135deg,#fff,#22c55e,#fbbf24);-webkit-background-clip:text;color:transparent}.resultCard.lose #resultTitle{background:linear-gradient(135deg,#fff,#ef4444);-webkit-background-clip:text;color:transparent}.resultCard.draw #resultTitle{background:linear-gradient(135deg,#fff,#60a5fa);-webkit-background-clip:text;color:transparent}#resultText{font-size:18px!important;line-height:1.65;color:#dbeafe;margin:16px 0 18px!important}.podium{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:10px}.podiumCard{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.11);border-radius:18px;padding:12px;text-align:center}.winnerBanner{background:linear-gradient(135deg,rgba(34,211,238,.18),rgba(99,102,241,.20));border:1px solid rgba(125,211,252,.28);border-radius:18px;padding:14px;margin-bottom:12px;color:#fff;font-weight:900;font-size:22px}.statBox{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.11);border-radius:16px;padding:10px;margin:10px 0}@keyframes battleCardIn{from{opacity:0;transform:translateY(18px) scale(.96)}to{opacity:1;transform:translateY(0) scale(1)}}@keyframes vsPop{0%{opacity:0;transform:scale(.52) rotate(-2deg)}70%{opacity:1;transform:scale(1.08)}100%{transform:scale(1)}}@keyframes choiceGood{0%,100%{transform:translateX(0)}35%{transform:translateX(5px)}}@keyframes choiceBad{0%,100%{transform:translateX(0)}20%{transform:translateX(-6px)}40%{transform:translateX(6px)}60%{transform:translateX(-3px)}}@keyframes effectPop{0%{opacity:0;transform:translate(-50%,-50%) scale(.55)}35%{opacity:1;transform:translate(-50%,-50%) scale(1.08)}100%{opacity:0;transform:translate(-50%,-50%) scale(1)}}@keyframes effectShake{0%{opacity:0;transform:translate(-50%,-50%) scale(.75)}20%{opacity:1;transform:translate(calc(-50% - 8px),-50%) scale(1.02)}40%{transform:translate(calc(-50% + 8px),-50%) scale(1.02)}65%{transform:translate(-50%,-50%) scale(1)}100%{opacity:0;transform:translate(-50%,-50%) scale(.96)}}@keyframes countPulse{0%{opacity:0;transform:scale(.45)}45%{opacity:1;transform:scale(1.12)}100%{transform:scale(1)}}@keyframes resultRise{from{opacity:0;transform:translateY(26px) scale(.94)}to{opacity:1;transform:translateY(0) scale(1)}}@keyframes resultTitlePop{from{transform:scale(.62);opacity:0}to{transform:scale(1);opacity:1}}
#endScreen{padding:24px;overflow:auto;background:transparent}
.awardShell{width:min(1180px,96vw);margin:0 auto;position:relative;overflow:hidden;padding:22px;background:linear-gradient(180deg,rgba(15,31,58,.88),rgba(8,18,35,.92));border:1px solid rgba(125,211,252,.32);box-shadow:0 30px 90px rgba(0,0,0,.42),inset 0 1px 0 rgba(255,255,255,.10);animation:awardFadeIn .45s ease-out}.awardShell::before{content:"";position:absolute;inset:-1px;background:radial-gradient(circle at 20% 0,rgba(56,189,248,.20),transparent 28%),radial-gradient(circle at 88% 8%,rgba(250,204,21,.16),transparent 25%),radial-gradient(circle at 50% 100%,rgba(139,92,246,.18),transparent 40%);pointer-events:none}.awardShell>*{position:relative}.awardHeader{display:grid;grid-template-columns:150px 1fr 150px;gap:14px;align-items:center;margin-bottom:12px}.brandMascotWrap{display:flex;align-items:center;justify-content:center}.brandMascot{width:120px;height:120px;object-fit:contain;border-radius:0;border:0;background:transparent;box-shadow:none;filter:drop-shadow(0 14px 22px rgba(15,31,58,.30)) drop-shadow(0 0 8px rgba(56,189,248,.18));animation:mascotFloat 2.4s ease-in-out infinite}.awardTitleBox{text-align:center}.awardKicker{display:inline-flex;align-items:center;gap:7px;padding:7px 12px;border-radius:999px;background:rgba(250,204,21,.12);border:1px solid rgba(250,204,21,.28);color:#fde68a;font-weight:1000;font-size:12px}.awardMainTitle{margin:8px 0 2px;font-size:clamp(38px,6vw,70px);font-weight:1000;letter-spacing:-1px;background:linear-gradient(135deg,#fff,#38bdf8 45%,#fbbf24);-webkit-background-clip:text;color:transparent;text-shadow:0 20px 60px rgba(56,189,248,.18);animation:awardTitlePop .7s cubic-bezier(.17,.89,.29,1.22)}.awardSubtitle{color:#b8cce8;font-weight:800}.awardGrid{display:grid;grid-template-columns:minmax(0,1.45fr) minmax(300px,.8fr);gap:14px;margin-top:12px}.awardCard{background:linear-gradient(180deg,rgba(255,255,255,.09),rgba(255,255,255,.045));border:1px solid rgba(255,255,255,.12);border-radius:24px;padding:16px;box-shadow:0 18px 50px rgba(0,0,0,.20);backdrop-filter:blur(14px)}.awardSectionTitle{margin:0 0 12px;color:#f8fbff;font-size:18px}.podiumStage{height:338px;display:grid;grid-template-columns:1fr 1.08fr 1fr;gap:12px;align-items:end;padding:8px 6px 0;position:relative}.podiumStage::before{content:"";position:absolute;left:0;right:0;bottom:0;height:42%;border-radius:28px;background:linear-gradient(180deg,rgba(56,189,248,.07),rgba(15,23,42,.14));border:1px solid rgba(255,255,255,.08)}.podiumSpot{position:relative;position:relative;display:flex;flex-direction:column;align-items:center;justify-content:flex-end;min-width:0;animation:podiumRise .65s ease-out both}.podiumSpot.rank2{animation-delay:.18s}.podiumSpot.rank1{animation-delay:.02s}.podiumSpot.rank3{animation-delay:.32s}.winnerGlow{position:absolute;top:-16px;width:116px;height:116px;border-radius:999px;background:radial-gradient(circle,rgba(255,255,255,.92),rgba(255,255,255,.38) 42%,transparent 72%);filter:blur(2px);animation:glowPulse 1.6s ease-in-out infinite}.winnerConfetti{position:absolute;top:-24px;left:50%;width:250px;height:190px;transform:translateX(-50%);pointer-events:none;overflow:visible;z-index:3}.confettiPiece{position:absolute;width:10px;height:16px;border-radius:2px;opacity:0;transform-origin:center;animation:confettiScatter 2.8s ease-in-out infinite;box-shadow:0 2px 8px rgba(0,0,0,.12);left:var(--sx);top:var(--sy)}.confettiPiece:nth-child(1){background:#fbbf24;animation-delay:0s;--sx:10%;--sy:40%;--dx:-8px;--dy:44px;--rot:-200deg}.confettiPiece:nth-child(2){background:#38bdf8;animation-delay:.16s;--sx:22%;--sy:18%;--dx:-12px;--dy:52px;--rot:170deg}.confettiPiece:nth-child(3){background:#f43f5e;animation-delay:.42s;--sx:34%;--sy:34%;--dx:10px;--dy:42px;--rot:-240deg}.confettiPiece:nth-child(4){background:#22c55e;animation-delay:.08s;--sx:18%;--sy:62%;--dx:-14px;--dy:36px;--rot:210deg}.confettiPiece:nth-child(5){background:#a855f7;animation-delay:.34s;--sx:42%;--sy:8%;--dx:8px;--dy:56px;--rot:-180deg}.confettiPiece:nth-child(6){background:#fb923c;animation-delay:.12s;--sx:58%;--sy:14%;--dx:14px;--dy:54px;--rot:250deg}.confettiPiece:nth-child(7){background:#eab308;animation-delay:.28s;--sx:70%;--sy:40%;--dx:10px;--dy:42px;--rot:-150deg}.confettiPiece:nth-child(8){background:#ec4899;animation-delay:.52s;--sx:82%;--sy:22%;--dx:16px;--dy:58px;--rot:165deg}.confettiPiece:nth-child(9){background:#06b6d4;animation-delay:.68s;--sx:66%;--sy:62%;--dx:-8px;--dy:32px;--rot:190deg}.confettiPiece:nth-child(10){background:#84cc16;animation-delay:.84s;--sx:88%;--sy:54%;--dx:6px;--dy:48px;--rot:-200deg}.confettiPiece:nth-child(11){background:#60a5fa;animation-delay:.46s;--sx:50%;--sy:0%;--dx:-6px;--dy:60px;--rot:220deg}.confettiPiece:nth-child(12){background:#f97316;animation-delay:.72s;--sx:4%;--sy:8%;--dx:10px;--dy:52px;--rot:-170deg}.podiumAvatar{width:74px;height:74px;object-fit:cover;border-radius:22px;border:1px solid rgba(255,255,255,.18);background:rgba(255,255,255,.08);box-shadow:0 16px 34px rgba(0,0,0,.24);z-index:1}.rank1 .podiumAvatar{width:98px;height:98px;border-radius:28px;box-shadow:0 20px 48px rgba(250,204,21,.22),0 0 0 7px rgba(250,204,21,.10)}.medalBadge{margin-top:8px;padding:6px 11px;border-radius:999px;font-weight:1000;color:#0f172a;background:#e5e7eb}.rank1 .medalBadge{background:linear-gradient(135deg,#fde68a,#f59e0b)}.rank2 .medalBadge{background:linear-gradient(135deg,#f8fafc,#94a3b8)}.rank3 .medalBadge{background:linear-gradient(135deg,#fdba74,#c2410c);color:#fff}.podiumName{margin-top:7px;font-size:clamp(16px,2vw,21px);font-weight:1000;color:#fff;text-align:center;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.podiumMeta{margin-top:3px;color:#aecaec;font-size:12px;font-weight:800;text-align:center}.podiumScore{margin-top:5px;font-size:22px;font-weight:1000;color:#38bdf8}.rank1 .podiumScore{font-size:28px;color:#fbbf24}.podiumBase{width:100%;border-radius:18px 18px 10px 10px;margin-top:10px;display:flex;align-items:center;justify-content:center;font-size:46px;font-weight:1000;color:rgba(255,255,255,.92);text-shadow:0 8px 20px rgba(0,0,0,.25);box-shadow:inset 0 1px 0 rgba(255,255,255,.20),0 14px 34px rgba(0,0,0,.20)}.rank1 .podiumBase{height:112px;background:linear-gradient(180deg,#fbbf24,#d97706)}.rank2 .podiumBase{height:86px;background:linear-gradient(180deg,#cbd5e1,#64748b)}.rank3 .podiumBase{height:70px;background:linear-gradient(180deg,#fb923c,#9a3412)}.teamAwardBanner{display:flex;gap:14px;align-items:center;padding:16px;border-radius:22px;background:linear-gradient(135deg,rgba(34,211,238,.18),rgba(99,102,241,.22));border:1px solid rgba(125,211,252,.30);margin-bottom:12px;animation:slideInRight .55s ease-out}.teamAwardIcon{font-size:38px}.teamAwardText strong{display:block;font-size:24px;color:#fff}.mvpGrid{display:grid;grid-template-columns:1fr;gap:10px}.mvpCard{display:flex;align-items:center;gap:12px;padding:12px;border-radius:18px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.10);animation:slideInRight .55s ease-out both}.mvpCard:nth-child(2){animation-delay:.08s}.mvpCard:nth-child(3){animation-delay:.16s}.mvpIcon{width:42px;height:42px;border-radius:14px;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,rgba(56,189,248,.25),rgba(139,92,246,.20));font-size:24px}.mvpText{min-width:0}.mvpText b{display:block;color:#fff;font-size:16px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.mvpText span{display:block;color:#aecaec;font-size:12px;margin-top:3px}.fullRankList{max-height:250px;overflow:auto;padding-right:4px}.awardLogList{max-height:170px;overflow:auto;padding-right:4px}.awardFooter{display:flex;justify-content:center;margin-top:14px}.emptyPodium{height:220px;display:flex;align-items:center;justify-content:center;color:#aecaec;text-align:center;border:1px dashed rgba(255,255,255,.15);border-radius:20px}
/* Award podium character: background-free, uses each student's selected color */
.podiumCharacter{width:86px;height:86px;border-radius:24px;background:linear-gradient(145deg,var(--pc-light),var(--pc));border:7px solid var(--pc-dark);box-shadow:0 18px 38px rgba(0,0,0,.24),inset 7px 7px 0 rgba(255,255,255,.18),inset -9px -9px 0 rgba(0,0,0,.10);position:relative;z-index:1;overflow:hidden;animation:mascotFloat 2.4s ease-in-out infinite}
.rank1 .podiumCharacter{width:114px;height:114px;border-radius:30px;border-width:8px;box-shadow:0 22px 54px rgba(250,204,21,.24),0 0 0 8px rgba(250,204,21,.10),inset 8px 8px 0 rgba(255,255,255,.20),inset -12px -12px 0 rgba(0,0,0,.12)}
.podiumCharacterImg{width:100%;height:100%;object-fit:contain;display:block}.pcGloss,.pcEye{display:none!important}
.pcGloss{position:absolute;left:10%;top:9%;right:11%;height:28%;border-radius:20px 20px 12px 12px;background:linear-gradient(135deg,rgba(255,255,255,.32),rgba(255,255,255,.04));pointer-events:none}
.pcGloss::after{content:"";position:absolute;right:0;top:0;width:18px;height:18px;background:rgba(255,255,255,.22);clip-path:polygon(100% 0,0 0,100% 100%)}
.pcEye{position:absolute;top:42%;width:28%;height:18%;background:#050505;border-radius:8px;box-shadow:inset 0 -2px 0 rgba(255,255,255,.04)}
.pcEyeLeft{left:23%}.pcEyeRight{right:23%}
.pcEye span{position:absolute;left:13%;top:13%;width:20%;height:34%;background:rgba(255,255,255,.92);border-radius:999px}
.rank1 .pcEye{border-radius:10px}

.fullRankList .rankItem{
  padding-left:14px!important;
  padding-right:14px!important;
  gap:18px!important;
}
.fullRankList .rankItem strong{
  margin-left:16px;
  padding-right:4px;
}


.topRoomTitle,.teacherRoomTitle{
  color:#fde047!important;
  font-size:34px!important;
  font-weight:1000!important;
  text-shadow:0 3px 14px rgba(250,204,21,.55),0 0 24px rgba(250,204,21,.38)!important;
  letter-spacing:.9px!important;
}

@media (max-width:880px){.awardHeader{grid-template-columns:1fr}.brandMascotWrap:last-child{display:none}.awardGrid{grid-template-columns:1fr}.podiumStage{height:auto;grid-template-columns:1fr}.podiumStage::before{display:none}.podiumBase{height:54px!important}.podiumSpot{margin-bottom:10px}.winnerConfetti{width:190px;height:150px;top:-12px}}@keyframes awardFadeIn{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:translateY(0)}}@keyframes awardTitlePop{from{opacity:0;transform:scale(.72)}to{opacity:1;transform:scale(1)}}@keyframes mascotFloat{0%,100%{transform:translateY(0)}50%{transform:translateY(-8px)}}@keyframes podiumRise{from{opacity:0;transform:translateY(60px) scale(.94)}to{opacity:1;transform:translateY(0) scale(1)}}@keyframes glowPulse{0%,100%{opacity:.72;transform:scale(.94)}50%{opacity:1;transform:scale(1.08)}}@keyframes confettiScatter{0%{opacity:0;transform:translate(0,0) rotate(0deg) scale(.5)}10%{opacity:1}78%{opacity:1}100%{opacity:0;transform:translate(var(--dx),var(--dy)) rotate(var(--rot)) scale(1)}}@keyframes slideInRight{from{opacity:0;transform:translateX(24px)}to{opacity:1;transform:translateX(0)}}

@media(max-width:1180px){html,body{overflow:auto}#gameScreen{grid-template-columns:1fr;height:auto;overflow:auto}.studentMain{min-height:560px}.infoGrid{grid-template-columns:1fr}.row{flex-direction:column}.card{padding:18px}}


/* ===== ReMap contrast patch: bright content cards on dark gradient background ===== */
.card,
.createCard{
  background:linear-gradient(180deg,rgba(255,255,255,.97),rgba(235,244,255,.95))!important;
  border:1px solid rgba(148,197,255,.48)!important;
  color:#10233f!important;
  box-shadow:0 26px 70px rgba(0,0,0,.26), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
.card::before,
.createCard::before{
  background:radial-gradient(circle at 10% 0,rgba(56,189,248,.16),transparent 30%),radial-gradient(circle at 90% 12%,rgba(139,92,246,.12),transparent 30%)!important;
}
.card h1,
.createCard h1{
  color:#0f2a4d!important;
}
.card .mini,
.createCard .mini,
.panel .mini,
.miniCard .mini,
.helpBox .mini{
  color:#5b718f!important;
}
.panel,
.infoCard,
.miniCard,
.helpBox,
#questionsList .qitem{
  background:linear-gradient(180deg,rgba(255,255,255,.88),rgba(225,237,252,.82))!important;
  border:1px solid rgba(91,141,204,.28)!important;
  color:#10233f!important;
  box-shadow:0 14px 34px rgba(15,31,58,.10)!important;
}
.panel h3,
.sectionTitle,
.mapTitleWrap h3{
  color:#12315a!important;
}
.label,
.field label,
.item span:first-child,
.rankItem span:first-child,
.battleItem span:first-child,
.logItem span:first-child{
  color:#526a8a!important;
}
.value,
.item strong,
.rankItem strong,
.battleItem strong,
.statHero strong,
.miniCard strong,
.codeValue{
  color:#0f2544!important;
  text-shadow:none!important;
}
.item,
.rankItem,
.battleItem,
.logItem{
  color:#173455!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
}
input,
textarea,
select{
  background:rgba(255,255,255,.82)!important;
  border:1px solid rgba(91,141,204,.35)!important;
  color:#10233f!important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.92)!important;
}
input::placeholder,
textarea::placeholder{
  color:rgba(82,106,138,.68)!important;
}
input:focus,
textarea:focus,
select:focus{
  border-color:rgba(37,99,235,.62)!important;
  box-shadow:0 0 0 3px rgba(37,99,235,.14), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
button.ghost,
button.soft,
.modeBtn,
.mapBtn,
.teamBtn{
  background:linear-gradient(180deg,rgba(255,255,255,.92),rgba(226,238,255,.86))!important;
  color:#12315a!important;
  border:1px solid rgba(91,141,204,.34)!important;
}
button.ghost:hover,
button.soft:hover,
.modeBtn:hover,
.mapBtn:hover,
.teamBtn:hover{
  filter:brightness(1.02)!important;
}
.modeBtn.active,
.mapBtn.active,
.teamBtn.active{
  background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed)!important;
  color:#fff!important;
  border-color:rgba(37,99,235,.50)!important;
}
.badge,
.codeTag,
.teamChip{
  background:rgba(14,116,144,.10)!important;
  border:1px solid rgba(14,116,144,.22)!important;
  color:#0f5f78!important;
}
.statHero{
  background:linear-gradient(135deg,rgba(224,242,254,.92),rgba(226,232,255,.90))!important;
  border:1px solid rgba(91,141,204,.25)!important;
}
#mapWrap,
.mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.88),rgba(213,228,247,.82))!important;
  border:1px solid rgba(91,141,204,.25)!important;
}
#mapWrap::before,
.mapWrap::before{
  background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.12),transparent 45%)!important;
}
.pillState,
.teamPill{
  background:rgba(34,197,94,.12)!important;
  border:1px solid rgba(34,197,94,.26)!important;
  color:#15803d!important;
}
.submitDone{color:#15803d!important}.submitWait{color:#b91c1c!important}
.colorBtn{border-color:rgba(15,31,58,.18)!important}


/* ===== ReMap hierarchy contrast refinement =====
   Dark background -> bright main card -> darker section panels -> bright inner cells.
   This keeps the big container easy to see while restoring clear boundaries inside it. */
.card,
.createCard{
  background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(231,241,255,.96))!important;
  border:1px solid rgba(191,219,254,.72)!important;
  color:#0f2544!important;
}
.panel,
.helpBox,
.miniCard,
.studentMapPanel,
.mapPanel{
  background:linear-gradient(145deg,rgba(34,58,92,.94),rgba(19,42,72,.92))!important;
  border:1px solid rgba(96,165,250,.38)!important;
  color:#eef7ff!important;
  box-shadow:0 18px 44px rgba(15,31,58,.20)!important;
}
.panel h3,
.sectionTitle,
.mapTitleWrap h3{
  color:#f8fbff!important;
}
.panel .mini,
.helpBox .mini,
.miniCard .mini,
.mapTitleWrap .mini{
  color:#c7d8ee!important;
}
.infoCard,
#questionsList .qitem,
.item,
.rankItem,
.battleItem,
.logItem,
.statHero{
  background:linear-gradient(180deg,rgba(255,255,255,.95),rgba(230,241,255,.90))!important;
  border:1px solid rgba(147,197,253,.46)!important;
  color:#10233f!important;
  box-shadow:0 10px 26px rgba(15,31,58,.10)!important;
}
.infoGrid .infoCard{
  background:linear-gradient(180deg,rgba(255,255,255,.97),rgba(221,236,255,.92))!important;
}
.label,
.field label,
.item span:first-child,
.rankItem span:first-child,
.battleItem span:first-child,
.logItem span:first-child{
  color:#4d6688!important;
}
.value,
.item strong,
.rankItem strong,
.battleItem strong,
.statHero strong,
.miniCard strong,
.codeValue{
  color:#0f2544!important;
}
input,
textarea,
select{
  background:rgba(255,255,255,.96)!important;
  border:1px solid rgba(91,141,204,.42)!important;
  color:#10233f!important;
}
#questionsList .qitem textarea,
#questionsList .qitem input{
  background:rgba(255,255,255,.98)!important;
}
.teamBtn{
  color:#fff!important;
  border:1px solid rgba(255,255,255,.45)!important;
  text-shadow:0 1px 2px rgba(0,0,0,.28)!important;
  box-shadow:0 8px 18px rgba(15,31,58,.14)!important;
}
.teamBtn.active{
  color:#fff!important;
  border-color:rgba(255,255,255,.88)!important;
  box-shadow:0 14px 28px rgba(15,31,58,.22),0 0 0 3px rgba(255,255,255,.78)!important;
}
.team-A{background:#3b82f6!important}.team-B{background:#ef4444!important}.team-C{background:#facc15!important;color:#fff!important;text-shadow:0 1px 2px rgba(0,0,0,.38)!important}.team-D{background:#22c55e!important}.team-E{background:#64748b!important}
.modeBtn,
.mapBtn,
button.ghost,
button.soft{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(226,238,255,.88))!important;
  color:#12315a!important;
}
.modeBtn.active,
.mapBtn.active{
  background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed)!important;
  color:#fff!important;
}
.badge,
.codeTag,
.teamChip{
  background:rgba(219,246,255,.92)!important;
  border:1px solid rgba(14,116,144,.22)!important;
  color:#0f5f78!important;
}



/* ===== ReMap contrast hierarchy fix v2 =====
   1) Keep operation dashboard clean/light after room creation.
   2) Fixed team color chips use real team colors.
   3) Student team selection keeps its original team color; selection uses border/glow only. */
#operateScreen .panel,
#operateScreen .miniCard,
#operateScreen .mapPanel{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(222,235,252,.90))!important;
  border:1px solid rgba(147,197,253,.42)!important;
  color:#10233f!important;
  box-shadow:0 16px 38px rgba(15,31,58,.16)!important;
}
#operateScreen .sectionTitle,
#operateScreen .mapTitleWrap h3{
  color:#12315a!important;
}
#operateScreen .mini,
#operateScreen .miniCard .mini,
#operateScreen .mapTitleWrap .mini{
  color:#607692!important;
}
#operateScreen .miniCard strong,
#operateScreen .item strong,
#operateScreen .codeValue{
  color:#0f2544!important;
  text-shadow:none!important;
}
#operateScreen .item{
  background:transparent!important;
  color:#173455!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
  box-shadow:none!important;
}
#operateScreen .item span:first-child{
  color:#526a8a!important;
}
#operateScreen .mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.96),rgba(213,228,247,.90))!important;
  border:1px solid rgba(91,141,204,.28)!important;
}
.teamFixedBox{
  background:linear-gradient(145deg,rgba(34,58,92,.96),rgba(19,42,72,.94))!important;
  color:#eef7ff!important;
}
.teamFixedBox b{color:#fff!important;}
.teamLegend .teamChip,
.teamWrap .teamBtn{
  color:#fff!important;
  border:1px solid rgba(255,255,255,.54)!important;
  text-shadow:0 1px 2px rgba(0,0,0,.34)!important;
  box-shadow:0 8px 18px rgba(15,31,58,.14)!important;
}
.teamLegend .teamChip.team-A,
.teamWrap .teamBtn.team-A{background:#3b82f6!important;}
.teamLegend .teamChip.team-B,
.teamWrap .teamBtn.team-B{background:#ef4444!important;}
.teamLegend .teamChip.team-C,
.teamWrap .teamBtn.team-C{background:#facc15!important;color:#fff!important;text-shadow:0 1px 2px rgba(0,0,0,.45)!important;}
.teamLegend .teamChip.team-D,
.teamWrap .teamBtn.team-D{background:#22c55e!important;}
.teamLegend .teamChip.team-E,
.teamWrap .teamBtn.team-E{background:#64748b!important;}
.teamWrap .teamBtn.active{
  color:#fff!important;
  border-color:rgba(255,255,255,.96)!important;
  outline:3px solid rgba(255,255,255,.82)!important;
  outline-offset:2px!important;
  box-shadow:0 12px 24px rgba(15,31,58,.22)!important;
  transform:none!important;
}
.teamWrap .teamBtn:hover{
  filter:brightness(1.05)!important;
}


/* ===== ReMap student clean light dashboard patch =====
   Match the student running screen to the teacher operation screen:
   dark navy background + clean bright cards + centered bright map. */
#gameScreen .panel,
#gameScreen .studentMapPanel{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(222,235,252,.90))!important;
  border:1px solid rgba(147,197,253,.42)!important;
  color:#10233f!important;
  box-shadow:0 16px 38px rgba(15,31,58,.16)!important;
  backdrop-filter:none!important;
}
#gameScreen .panel h3,
#gameScreen .mapTitleWrap h3{
  color:#12315a!important;
}
#gameScreen .mini,
#gameScreen .mapTitleWrap .mini{
  color:#607692!important;
}
#gameScreen #mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.96),rgba(213,228,247,.90))!important;
  border:1px solid rgba(91,141,204,.28)!important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.80)!important;
}
#gameScreen #mapWrap::before{
  background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.10),transparent 45%)!important;
}
#gameScreen canvas{
  border:2px solid rgba(125,211,252,.48)!important;
  background:#d5e4f7!important;
  box-shadow:0 26px 60px rgba(15,31,58,.20),0 0 0 8px rgba(255,255,255,.30)!important;
}
#gameScreen .rankItem,
#gameScreen .battleItem,
#gameScreen .logItem,
#gameScreen .statHero{
  background:transparent!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
  color:#173455!important;
  box-shadow:none!important;
}
#gameScreen .rankItem span:first-child,
#gameScreen .battleItem span:first-child,
#gameScreen .logItem span:first-child,
#gameScreen .label{
  color:#526a8a!important;
}
#gameScreen .rankItem strong,
#gameScreen .battleItem strong,
#gameScreen .statHero strong,
#gameScreen .value{
  color:#0f2544!important;
  text-shadow:none!important;
}
#gameScreen .teamPill{
  background:rgba(34,197,94,.12)!important;
  border:1px solid rgba(34,197,94,.26)!important;
  color:#15803d!important;
}
#gameScreen button.ghost{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(226,238,255,.88))!important;
  color:#12315a!important;
  border:1px solid rgba(91,141,204,.34)!important;
  box-shadow:none!important;
}

.teacherAwardModal{position:fixed;inset:0;z-index:80;display:none;align-items:center;justify-content:center;padding:22px;background:rgba(2,8,23,.72);backdrop-filter:blur(10px)}
.teacherAwardModal.show{display:flex}
.teacherAwardCard{position:relative;width:min(980px,96vw);max-height:92vh;overflow:auto;padding:24px;border-radius:28px;background:linear-gradient(145deg,#eaf7ff,#eef2ff 56%,#dcecff);border:1px solid rgba(125,211,252,.42);box-shadow:0 30px 90px rgba(0,0,0,.38);color:#0f2545}
.teacherAwardClose{position:absolute;right:16px;top:14px;width:38px;height:38px;border-radius:999px;background:#0f2545!important;color:#fff!important;padding:0;font-size:24px;line-height:1;box-shadow:none!important}
.teacherAwardCard h1{margin:8px 0 4px;text-align:center;color:#0f2545;font-size:38px}
.teacherAwardBody{display:grid;grid-template-columns:minmax(0,1.25fr) minmax(260px,.75fr);gap:14px;margin-top:16px}
.teacherAwardPanel{background:linear-gradient(180deg,#10223c,#07182d);color:#eef7ff;border-radius:20px;padding:16px;border:1px solid rgba(191,219,254,.25);box-shadow:0 16px 38px rgba(15,31,58,.22)}
.teacherAwardPanel h3{margin:0 0 12px;color:#fff}
.teacherPodiumList{display:grid;grid-template-columns:1fr;gap:10px}
.teacherPodiumItem{display:grid;grid-template-columns:54px minmax(0,1fr) 70px;align-items:center;gap:10px;padding:12px;border-radius:16px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12)}
.teacherPodiumItem.rank1{background:linear-gradient(135deg,rgba(251,191,36,.28),rgba(255,255,255,.08));border-color:rgba(251,191,36,.45)}
.teacherPodiumItem img{width:44px;height:44px;object-fit:contain}.teacherPodiumItem b{display:block;color:#fff;font-size:18px}.teacherPodiumItem span{color:#c7d9f2;font-size:12px}.teacherPodiumItem strong{text-align:right;color:#fbbf24;font-size:20px}.teacherAwardList{display:grid;gap:8px}.teacherAwardList .item{color:#eef7ff}.teacherAwardLog{max-height:190px;overflow:auto}
@media(max-width:720px){.teacherAwardBody{grid-template-columns:1fr}.teacherAwardCard{padding:18px}.teacherAwardCard h1{font-size:28px}}



/* ===== Final ceremony readability + mascot image patch ===== */
#endScreen{
  background:linear-gradient(135deg,#eef7ff 0%,#f8fbff 45%,#e7efff 100%)!important;
  padding:10px!important;
  overflow:auto!important;
}
#endScreen .awardShell{
  width:min(1280px,98vw)!important;
  min-height:calc(100vh - 20px)!important;
  background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(232,242,255,.96))!important;
  border:2px solid rgba(15,31,58,.85)!important;
  border-radius:26px!important;
  color:#0f172a!important;
  padding:18px 24px 20px!important;
  box-shadow:0 24px 70px rgba(15,31,58,.22), inset 0 1px 0 rgba(255,255,255,.98)!important;
}
#endScreen .awardShell::before{
  background:radial-gradient(circle at 18% 0,rgba(56,189,248,.13),transparent 30%),radial-gradient(circle at 82% 3%,rgba(99,102,241,.12),transparent 28%)!important;
}
#endScreen .awardHeader{
  grid-template-columns:150px 1fr 150px!important;
  align-items:center!important;
  margin-bottom:12px!important;
}
#endScreen .brandMascot{
  width:120px!important;
  height:120px!important;
  object-fit:contain!important;
  border-radius:0!important;
  border:0!important;
  background:transparent!important;
  box-shadow:none!important;
  filter:drop-shadow(0 14px 22px rgba(15,31,58,.30)) drop-shadow(0 0 8px rgba(56,189,248,.18))!important;
  animation:mascotFloat 2.4s ease-in-out infinite!important;
}
#endScreen .awardKicker{
  background:linear-gradient(180deg,#13233f,#061326)!important;
  border:2px solid #fbbf24!important;
  color:#ffd84d!important;
  box-shadow:0 8px 18px rgba(15,31,58,.18)!important;
  font-size:15px!important;
  letter-spacing:.3px!important;
  padding:8px 22px!important;
}
#endScreen .awardMainTitle{
  color:#071b33!important;
  background:none!important;
  -webkit-background-clip:initial!important;
  text-shadow:0 4px 0 rgba(255,255,255,.55),0 16px 34px rgba(15,31,58,.16)!important;
}
#endScreen .awardSubtitle{color:#263b59!important;font-weight:900!important;}
#endScreen .teamAwardBanner{
  background:linear-gradient(135deg,rgba(191,226,255,.96),rgba(202,214,255,.92))!important;
  border:1px solid rgba(96,165,250,.42)!important;
  color:#0f172a!important;
}
#endScreen .teamAwardText strong,
#endScreen .teamAwardText span,
#endScreen .teamAwardText .mini{color:#0f172a!important;}
#endScreen .awardGrid{grid-template-columns:minmax(0,1.7fr) minmax(330px,.9fr)!important;gap:18px!important;}
#endScreen .awardCard{
  background:linear-gradient(145deg,#14243d,#071a31)!important;
  border:1px solid rgba(191,219,254,.28)!important;
  color:#f8fbff!important;
  box-shadow:0 18px 46px rgba(15,31,58,.24)!important;
}
#endScreen .awardSectionTitle{color:#ffffff!important;font-weight:1000!important;}
#endScreen .podiumStage{height:440px!important;}
#endScreen .podiumStage::before{
  background:linear-gradient(180deg,rgba(255,255,255,.10),rgba(255,255,255,.05))!important;
  border:1px solid rgba(255,255,255,.14)!important;
}
#endScreen .podiumCharacter{
  background:transparent!important;
  border:0!important;
  overflow:visible!important;
  box-shadow:none!important;
  display:flex!important;
  align-items:center!important;
  justify-content:center!important;
  width:90px!important;
  height:90px!important;
}
#endScreen .rank1 .podiumCharacter{
  background:transparent!important;
  border:0!important;
  box-shadow:none!important;
  width:120px!important;
  height:120px!important;
}
#endScreen .podiumCharacterImg{
  width:100%!important;
  height:100%!important;
  object-fit:contain!important;
  display:block!important;
  filter:drop-shadow(0 14px 24px rgba(0,0,0,.32)) drop-shadow(0 0 6px rgba(56,189,248,.14))!important;
}
#endScreen .rank1 .podiumCharacterImg{
  filter:drop-shadow(0 20px 38px rgba(250,204,21,.38)) drop-shadow(0 0 16px rgba(250,204,21,.28)) drop-shadow(0 14px 24px rgba(0,0,0,.28))!important;
}
#endScreen .podiumName,
#endScreen .podiumScore,
#endScreen .mvpText b,
#endScreen .mvpText span,
#endScreen .rankItem,
#endScreen .rankItem span,
#endScreen .rankItem strong,
#endScreen .logItem,
#endScreen .logItem span,
#endScreen .mini{
  color:#ffffff!important;
}
#endScreen .podiumMeta{color:#dbeafe!important;}
#endScreen .rank1 .podiumScore{color:#fbbf24!important;}
#endScreen .medalBadge{color:#111827!important;}
#endScreen .mvpCard,
#endScreen .rankItem,
#endScreen .logItem{
  background:linear-gradient(180deg,rgba(255,255,255,.10),rgba(255,255,255,.045))!important;
  border:1px solid rgba(255,255,255,.16)!important;
  border-radius:16px!important;
}
#endScreen .mvpIcon{
  color:#111827!important;
  background:linear-gradient(135deg,#dbeafe,#93c5fd)!important;
  border:1px solid rgba(255,255,255,.55)!important;
}
#endScreen .podiumBase{color:#ffffff!important;}
@media (max-width:880px){#endScreen .awardHeader{grid-template-columns:1fr!important}#endScreen .brandMascotWrap:last-child{display:none!important}#endScreen .podiumStage{height:auto!important}}


/* 24-color personal palette: 4 rows x 6 columns - fit inside room info panel */
.colorWrap{
  display:grid!important;
  grid-template-columns:repeat(6, minmax(0, 1fr))!important;
  grid-auto-rows:42px!important;
  gap:8px 6px!important;
  align-items:center!important;
  justify-items:center!important;
  align-content:start!important;
  width:100%!important;
  max-width:100%!important;
  box-sizing:border-box!important;
  margin-top:10px!important;
  overflow:visible!important;
}
.colorBtn{
  width:28px!important;
  height:28px!important;
  min-width:28px!important;
  min-height:28px!important;
  padding:0!important;
  border-radius:999px!important;
  border:2px solid rgba(255,255,255,.78)!important;
  box-shadow:0 4px 10px rgba(15,31,58,.18)!important;
}
.colorBtn[style*="#ffffff"]{
  border-color:rgba(15,23,42,.65)!important;
}
.colorBtn[style*="#111827"]{
  border-color:rgba(255,255,255,.9)!important;
}


.colorBtn.active{
  outline:3px solid rgba(125,211,252,.9)!important;
  outline-offset:3px!important;
}


.colorHelp{font-size:16px!important;font-weight:800!important;color:#e4efff!important;line-height:1.35;margin:8px 0 10px!important}

.questionChoiceRow{display:flex;gap:8px;align-items:center;margin-top:8px}
.choiceBadge{display:inline-flex;align-items:center;justify-content:center;width:38px;height:34px;min-width:38px;border-radius:999px;background:linear-gradient(180deg,#e0faff,#caf4ff);border:1px solid rgba(34,211,238,.36);color:#115e75;font-weight:1000;font-size:20px;line-height:1;box-shadow:0 4px 10px rgba(15,31,58,.08)}
.answerRadioLabel{display:inline-flex;align-items:center;gap:8px;white-space:nowrap;font-weight:800;color:#183153;line-height:1;border:none!important;background:transparent!important;box-shadow:none!important;text-decoration:none!important}
.answerRadioLabel input[type='radio']{appearance:none;-webkit-appearance:none;width:18px;height:18px;margin:0;border-radius:999px;border:2px solid #7b8798;background:#ffffff;display:inline-block;position:relative;box-shadow:none;outline:none;vertical-align:middle;flex:0 0 auto}
.answerRadioLabel input[type='radio']:checked{border-color:#0ea5e9;background:#ffffff}
.answerRadioLabel input[type='radio']::after{content:'';position:absolute;top:50%;left:50%;width:8px;height:8px;border-radius:999px;background:transparent;transform:translate(-50%,-50%)}
.answerRadioLabel input[type='radio']:checked::after{background:#0ea5e9}
.answerRadioLabel span{display:inline-block}


/* ===== Student mobile prep screen fix =====
   Galaxy / phone browsers were rendering the desktop prep card too wide.
   Keep PC layout, but make the student join/prep flow responsive first. */
@media (pointer:coarse), (max-width:900px){
  html,body{width:100%!important;max-width:100%!important;overflow-x:hidden!important;}
  body{min-height:100dvh!important;overflow:hidden!important;}
  .topbar{height:56px!important;padding:0 max(10px, env(safe-area-inset-left)) 0 10px!important;gap:8px!important;}
  .brand{font-size:18px!important;flex:0 0 auto!important;}
  .roomTitle{font-size:clamp(18px,5vw,28px)!important;width:46vw!important;max-width:46vw!important;letter-spacing:0!important;line-height:1.05!important;}
  .statusBadge{max-width:34vw!important;padding:6px 8px!important;font-size:10px!important;gap:5px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;}
  .statusBadge::before{width:7px!important;height:7px!important;flex:0 0 auto!important;}
  body.show-top-leave .topLeaveBtn{display:inline-flex!important;flex:0 0 auto!important;}

  #joinScreen,#prepScreen{width:100vw!important;height:calc(100dvh - 56px)!important;min-height:calc(100dvh - 56px)!important;max-width:100vw!important;padding:12px!important;overflow-y:auto!important;overflow-x:hidden!important;-webkit-overflow-scrolling:touch!important;align-items:flex-start!important;justify-content:center!important;}
  #joinScreen{align-items:center!important;}
  #joinScreen .card,#prepScreen .card{width:100%!important;max-width:100%!important;border-radius:22px!important;padding:18px!important;overflow:visible!important;}
  #prepScreen .card h1{font-size:clamp(30px,8vw,44px)!important;line-height:1.12!important;margin:0 0 10px!important;padding-top:0!important;}
  #prepScreen .card > .mini{font-size:13px!important;line-height:1.5!important;}

  #prepScreen .prepLayout{display:grid!important;grid-template-columns:1fr!important;gap:14px!important;margin-top:14px!important;align-items:start!important;flex-direction:unset!important;}
  #prepScreen .prepInfoPanel,#prepScreen .prepQuestionPanel{width:100%!important;min-width:0!important;max-width:100%!important;padding:14px!important;border-radius:20px!important;margin:0!important;}
  #prepScreen .panelTitle,#prepScreen .prepQuestionPanel > div:first-child{gap:10px!important;}
  #prepScreen .panel h3{font-size:19px!important;margin-bottom:10px!important;}
  #prepScreen .infoGrid{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:8px!important;}
  #prepScreen .infoCard{min-width:0!important;padding:10px!important;border-radius:16px!important;}
  #prepScreen .label{font-size:12px!important;line-height:1.25!important;}
  #prepScreen .value{font-size:18px!important;line-height:1.22!important;overflow-wrap:anywhere!important;word-break:keep-all!important;}
  #prepScreen #infoTitle{font-size:16px!important;}
  #prepScreen #infoNickname{font-size:18px!important;}

  #prepScreen .colorHelp{font-size:14px!important;margin:10px 0 8px!important;}
  #prepScreen .colorWrap{grid-template-columns:repeat(6,minmax(0,1fr))!important;grid-auto-rows:38px!important;gap:8px 5px!important;width:100%!important;max-width:100%!important;}
  #prepScreen .colorBtn{width:30px!important;height:30px!important;min-width:30px!important;min-height:30px!important;}
  #prepScreen .teamWrap{display:grid!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:8px!important;}
  #prepScreen .teamBtn{width:100%!important;text-align:center!important;}

  #prepScreen #questionsList{max-height:none!important;overflow:visible!important;padding-right:0!important;}
  #prepScreen #questionsList .qitem{padding:12px!important;margin:10px 0!important;border-radius:18px!important;max-width:100%!important;overflow:hidden!important;}
  #prepScreen #questionsList .qitem textarea,
  #prepScreen #questionsList .qitem input{width:100%!important;max-width:100%!important;min-width:0!important;}
  #prepScreen #questionsList .qitem textarea{min-height:82px!important;}
  #prepScreen .questionChoiceRow{display:grid!important;grid-template-columns:auto minmax(0,1fr) auto!important;gap:8px!important;align-items:center!important;}
  #prepScreen .questionChoiceRow .badge{white-space:nowrap!important;}
  #prepScreen .answerRadioLabel{font-size:14px!important;gap:6px!important;white-space:nowrap!important;justify-self:end!important;}
  #prepScreen .answerRadioLabel input[type='radio']{width:18px!important;height:18px!important;}

  #prepScreen .prepActions{display:grid!important;grid-template-columns:1fr 1fr!important;gap:10px!important;margin-top:14px!important;justify-content:stretch!important;}
  #prepScreen .prepActions button{width:100%!important;min-height:52px!important;}
}
@media (orientation:portrait) and (pointer:coarse){
  body.show-top-leave .statusBadge{display:none!important;}
  body.show-top-leave .roomTitle{width:42vw!important;max-width:42vw!important;}
  body.show-top-leave .topLeaveBtn{height:34px!important;padding:0 13px!important;font-size:12px!important;}
}

@media (max-width:520px){
  .topbar{height:54px!important;padding:0 8px!important;}
  .brand{font-size:16px!important;}
  .roomTitle{width:42vw!important;max-width:42vw!important;font-size:20px!important;}
  .statusBadge{max-width:31vw!important;font-size:9px!important;padding:5px 7px!important;}
  #joinScreen,#prepScreen{height:calc(100dvh - 54px)!important;min-height:calc(100dvh - 54px)!important;padding:10px!important;}
  #joinScreen .card,#prepScreen .card{padding:14px!important;border-radius:20px!important;}
  #prepScreen .infoGrid{grid-template-columns:1fr!important;}
  #prepScreen .colorWrap{grid-auto-rows:36px!important;gap:7px 4px!important;}
  #prepScreen .colorBtn{width:28px!important;height:28px!important;min-width:28px!important;min-height:28px!important;}
  #prepScreen .questionChoiceRow{grid-template-columns:auto minmax(0,1fr)!important;grid-template-areas:"choice input" "radio radio"!important;}
  #prepScreen .questionChoiceRow .badge{grid-area:choice!important;}
  #prepScreen .questionChoiceRow input[data-field='choice']{grid-area:input!important;}
  #prepScreen .questionChoiceRow .answerRadioLabel{grid-area:radio!important;justify-self:end!important;margin-top:2px!important;}
  #prepScreen .prepActions{grid-template-columns:1fr!important;}
}

@media (orientation:landscape) and (pointer:coarse) and (max-height:560px){
  .topbar{height:52px!important;}
  .brand{font-size:18px!important;}
  .roomTitle{font-size:25px!important;width:48vw!important;max-width:48vw!important;}
  .statusBadge{max-width:28vw!important;font-size:10px!important;}
  #joinScreen,#prepScreen{height:calc(100dvh - 52px)!important;min-height:calc(100dvh - 52px)!important;padding:10px!important;}
  #prepScreen .card{padding:16px!important;border-radius:20px!important;}
  #prepScreen .card h1{font-size:30px!important;margin-bottom:6px!important;}
  #prepScreen .card > .mini{font-size:13px!important;}
  #prepScreen .prepLayout{grid-template-columns:minmax(250px,36%) minmax(0,1fr)!important;gap:12px!important;}
  #prepScreen .prepInfoPanel,#prepScreen .prepQuestionPanel{padding:12px!important;border-radius:18px!important;}
  #prepScreen .panel h3{font-size:17px!important;margin-bottom:8px!important;}
  #prepScreen .infoGrid{grid-template-columns:1fr!important;gap:6px!important;}
  #prepScreen .infoCard{padding:8px 10px!important;border-radius:14px!important;}
  #prepScreen .label{font-size:11px!important;}
  #prepScreen .value{font-size:16px!important;}
  #prepScreen .colorWrap{grid-auto-rows:31px!important;gap:5px!important;}
  #prepScreen .colorBtn{width:24px!important;height:24px!important;min-width:24px!important;min-height:24px!important;}
  #prepScreen #questionsList{max-height:calc(100dvh - 225px)!important;overflow:auto!important;padding-right:4px!important;}
  #prepScreen #questionsList .qitem{padding:10px!important;margin:8px 0!important;}
  #prepScreen #questionsList .qitem textarea{min-height:58px!important;}
  #prepScreen .questionChoiceRow{grid-template-columns:auto minmax(0,1fr) auto!important;gap:7px!important;}
  #prepScreen .prepActions{position:sticky!important;bottom:0!important;z-index:3!important;background:linear-gradient(180deg,rgba(238,247,255,.88),rgba(238,247,255,.98))!important;border-radius:16px!important;padding:10px 0 0!important;margin-top:10px!important;}
  #prepScreen .prepActions button{min-height:44px!important;}
}


/* ===== v2.0 Galaxy student compact prep + full-map game patch ===== */
#touchMovePad{display:none;position:fixed;left:0;top:0;width:132px;height:132px;margin:-66px 0 0 -66px;z-index:40;pointer-events:none;touch-action:none;user-select:none;-webkit-user-select:none;opacity:0;transform:scale(.92);transition:opacity .08s ease,transform .08s ease;}
#touchMovePad.active{display:block;opacity:1;transform:scale(1);}
#touchMovePad .joyBase{position:absolute;inset:0;border-radius:999px;background:rgba(8,22,43,.30);border:2px solid rgba(56,189,248,.58);box-shadow:0 16px 36px rgba(0,0,0,.22),inset 0 0 0 10px rgba(125,211,252,.10);backdrop-filter:blur(6px);}
#touchMovePad .joyKnob{position:absolute;left:50%;top:50%;width:58px;height:58px;margin:-29px 0 0 -29px;border-radius:999px;background:radial-gradient(circle at 35% 30%,rgba(255,255,255,.50),rgba(56,189,248,.42) 35%,rgba(37,99,235,.50));border:2px solid rgba(224,242,254,.78);box-shadow:0 10px 24px rgba(15,31,58,.24);transform:translate(0,0);}
#touchMovePad .joyHint{position:absolute;left:50%;top:50%;width:12px;height:12px;margin:-6px 0 0 -6px;border-radius:999px;background:rgba(255,255,255,.45);}

@media (pointer:coarse), (max-width:900px){
  body.remap-game-active{overflow:hidden!important;}

  /* Student game: show the whole map first, not the PC side columns. */
  #gameScreen{width:100vw!important;max-width:100vw!important;height:calc(100dvh - 54px)!important;min-height:calc(100dvh - 54px)!important;grid-template-columns:1fr!important;grid-template-rows:1fr!important;gap:0!important;padding:6px!important;overflow:hidden!important;}
  #gameScreen .studentCol{display:none!important;}
  #gameScreen .studentMain{display:flex!important;width:100%!important;height:100%!important;min-width:0!important;min-height:0!important;overflow:hidden!important;}
  #gameScreen .studentMapPanel{width:100%!important;height:100%!important;min-width:0!important;min-height:0!important;margin:0!important;padding:8px!important;border-radius:18px!important;display:flex!important;flex-direction:column!important;overflow:hidden!important;}
  #gameScreen .mapHeader{flex:0 0 auto!important;min-height:32px!important;margin:0 0 6px!important;gap:8px!important;}
  #gameScreen .mapTitleWrap h3{font-size:18px!important;line-height:1!important;margin:0!important;}
  #gameScreen .mapTitleWrap .mini{display:none!important;}
  #gameScreen .teamPill{font-size:11px!important;line-height:1!important;padding:7px 10px!important;white-space:nowrap!important;}
  #gameScreen #mapWrap{flex:1 1 auto!important;min-height:0!important;width:100%!important;padding:6px!important;border-radius:16px!important;display:flex!important;align-items:center!important;justify-content:center!important;overflow:hidden!important;touch-action:none!important;}
  #gameScreen canvas{display:block!important;width:auto!important;height:auto!important;max-width:100%!important;max-height:100%!important;border-radius:14px!important;border-width:2px!important;box-shadow:0 14px 34px rgba(15,31,58,.24)!important;}
}

@media (orientation:landscape) and (pointer:coarse){
  /* Top bar + game fit for phone landscape.
     Keep the centered REMAP title clear by pinning status right next to the leave button. */
  .topbar{height:48px!important;padding:0 8px!important;justify-content:flex-start!important;position:relative!important;}
  .brand{font-size:17px!important;flex:0 0 auto!important;margin-right:auto!important;}
  .roomTitle{font-size:clamp(20px,4.2vw,30px)!important;width:30vw!important;max-width:30vw!important;z-index:1!important;}
  .statusBadge{max-width:24vw!important;font-size:9px!important;padding:5px 7px!important;}
  body.show-top-leave .topLeaveBtn{display:none!important;}
  body.show-top-leave .statusBadge{display:inline-flex!important;position:absolute!important;right:8px!important;top:50%!important;transform:translateY(-50%)!important;max-width:30vw!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;z-index:7!important;}
  #gameScreen{height:calc(100dvh - 48px)!important;min-height:calc(100dvh - 48px)!important;padding:5px!important;}
  #gameScreen .studentMapPanel{padding:7px!important;}
  #gameScreen .mapHeader{min-height:28px!important;margin-bottom:5px!important;}
  #gameScreen .mapTitleWrap h3{font-size:17px!important;}
  #gameScreen #mapWrap{padding:5px!important;}
  /* Student prep: keep it inside one landscape viewport and give the question editor priority. */
  #prepScreen{height:calc(100dvh - 48px)!important;min-height:calc(100dvh - 48px)!important;padding:8px!important;overflow:hidden!important;align-items:stretch!important;justify-content:center!important;}
  #prepScreen .card{height:100%!important;max-height:100%!important;width:100%!important;max-width:100%!important;padding:12px!important;border-radius:18px!important;display:flex!important;flex-direction:column!important;overflow:hidden!important;}
  #prepScreen .card h1{display:none!important;}
  #prepScreen .card > .mini{display:none!important;}
  #prepScreen .prepLayout{flex:1 1 auto!important;min-height:0!important;display:grid!important;grid-template-columns:minmax(210px,31%) minmax(0,1fr)!important;gap:10px!important;margin-top:0!important;overflow:hidden!important;align-items:stretch!important;}
  #prepScreen .prepInfoPanel,#prepScreen .prepQuestionPanel{height:100%!important;min-height:0!important;max-height:100%!important;margin:0!important;padding:10px!important;border-radius:17px!important;overflow:hidden!important;}
  #prepScreen .prepInfoPanel{display:flex!important;flex-direction:column!important;}
  #prepScreen .prepQuestionPanel{display:flex!important;flex-direction:column!important;}
  #prepScreen .prepQuestionPanel > div:first-child{flex:0 0 auto!important;}
  #prepScreen .prepQuestionPanel > .mini{display:none!important;}
  #prepScreen .panel h3{font-size:17px!important;margin:0 0 6px!important;}
  #prepScreen .infoGrid{grid-template-columns:1fr!important;gap:5px!important;overflow:auto!important;min-height:0!important;}
  #prepScreen .infoCard{padding:7px 9px!important;border-radius:13px!important;}
  #prepScreen .label{font-size:10px!important;margin-bottom:2px!important;}
  #prepScreen .value{font-size:14px!important;line-height:1.15!important;}
  #prepScreen #infoTitle,#prepScreen #infoNickname{font-size:14px!important;}
  #prepScreen #teamSection{margin-top:6px!important;}
  #prepScreen #colorHelp{display:none!important;}
  #prepScreen .colorWrap{flex:0 0 auto!important;grid-template-columns:repeat(6,minmax(0,1fr))!important;grid-auto-rows:27px!important;gap:4px!important;margin-top:6px!important;}
  #prepScreen .colorBtn{width:22px!important;height:22px!important;min-width:22px!important;min-height:22px!important;}
  #prepScreen #questionsList{flex:1 1 auto!important;min-height:0!important;max-height:none!important;overflow:auto!important;padding:0 4px 0 0!important;}
  #prepScreen #questionsList .qitem{margin:0!important;padding:10px!important;border-radius:16px!important;min-height:100%!important;display:flex!important;flex-direction:column!important;gap:6px!important;}
  #prepScreen #questionsList .qitem textarea{min-height:84px!important;font-size:19px!important;line-height:1.35!important;margin-top:4px!important;resize:none!important;}
  #prepScreen .questionChoiceRow{display:grid!important;grid-template-columns:54px minmax(0,1fr) 54px!important;gap:6px!important;margin-top:2px!important;align-items:center!important;}
  #prepScreen .questionChoiceRow .badge{font-size:11px!important;padding:6px 5px!important;text-align:center!important;}
  #prepScreen .questionChoiceRow input[data-field='choice']{height:39px!important;font-size:17px!important;padding:7px 9px!important;}
  #prepScreen .answerRadioLabel{font-size:12px!important;gap:4px!important;justify-self:end!important;}
  #prepScreen .answerRadioLabel input[type='radio']{width:17px!important;height:17px!important;}
  #prepScreen .prepActions{flex:0 0 auto!important;position:static!important;display:grid!important;grid-template-columns:1fr 1fr!important;gap:8px!important;margin-top:8px!important;padding:0!important;background:transparent!important;border-radius:0!important;}
  #prepScreen .prepActions button{min-height:42px!important;height:42px!important;font-size:16px!important;}
}

@media (orientation:portrait) and (pointer:coarse){
  /* Portrait phones: larger question editor without horizontal layout. */
  #prepScreen .prepQuestionPanel{padding:14px!important;}
  #prepScreen #questionsList .qitem textarea{min-height:120px!important;font-size:18px!important;}
  #prepScreen .questionChoiceRow input[data-field='choice']{min-height:44px!important;font-size:16px!important;}
}


/* ===== v2.2 mobile landscape side information panels =====
   Keep the mobile map centered, but reuse the desktop side information in the
   empty left/right spaces on phone landscape screens. */
@media (orientation:landscape) and (pointer:coarse){
  #gameScreen{
    grid-template-columns:clamp(118px,17vw,178px) minmax(0,1fr) clamp(118px,17vw,178px)!important;
    grid-template-rows:1fr!important;
    gap:6px!important;
    padding:5px!important;
  }
  #gameScreen .studentCol{
    display:flex!important;
    flex-direction:column!important;
    min-width:0!important;
    min-height:0!important;
    width:auto!important;
    height:100%!important;
    overflow:hidden!important;
    padding:0!important;
    gap:6px!important;
  }
  #gameScreen .studentMain{
    min-width:0!important;
    min-height:0!important;
    height:100%!important;
    overflow:hidden!important;
  }
  #gameScreen .studentMapPanel{
    padding:7px!important;
  }
  #gameScreen .panel{
    margin:0!important;
    padding:8px!important;
    border-radius:14px!important;
    overflow:hidden!important;
  }
  #gameScreen .panel h3{
    font-size:12px!important;
    line-height:1.1!important;
    margin:0 0 5px!important;
    white-space:nowrap!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
  }
  #studentLeft .panel:nth-of-type(1){flex:1 1 48%!important;min-height:0!important;}
  #studentLeft .panel:nth-of-type(2){flex:0 1 26%!important;min-height:54px!important;}
  #studentLeft .leaveWrap{flex:0 0 38px!important;padding:0!important;}
  #studentLeft #leaveBtn{height:38px!important;min-height:38px!important;font-size:13px!important;border-radius:12px!important;}
  #studentRight .panel:nth-of-type(1){flex:0 1 32%!important;min-height:72px!important;}
  #studentRight .panel:nth-of-type(2){flex:0 1 22%!important;min-height:52px!important;}
  #studentRight .panel:nth-of-type(3){flex:1 1 auto!important;min-height:0!important;}
  #topInfo,
  #rankingList,
  #teamRankingList,
  #battleList,
  #logList{
    min-height:0!important;
    max-height:100%!important;
    overflow:auto!important;
    -webkit-overflow-scrolling:touch!important;
    scrollbar-width:none!important;
  }
  #topInfo::-webkit-scrollbar,
  #rankingList::-webkit-scrollbar,
  #teamRankingList::-webkit-scrollbar,
  #battleList::-webkit-scrollbar,
  #logList::-webkit-scrollbar{display:none!important;}
  #gameScreen .statHero{
    padding:6px 7px!important;
    border-radius:11px!important;
    margin-bottom:4px!important;
  }
  #gameScreen .statHero .mini{font-size:9px!important;line-height:1.15!important;}
  #gameScreen .statHero strong{font-size:12px!important;line-height:1.15!important;margin-top:2px!important;}
  #gameScreen .rankItem,
  #gameScreen .battleItem,
  #gameScreen .logItem{
    gap:5px!important;
    padding:4px 0!important;
    font-size:10px!important;
    line-height:1.2!important;
  }
  #gameScreen .rankItem span,
  #gameScreen .battleItem span,
  #gameScreen .logItem span{
    min-width:0!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
  }
  #gameScreen .rankItem span:first-child,
  #gameScreen .battleItem span:first-child,
  #gameScreen .logItem span:first-child{white-space:nowrap!important;}
  #gameScreen .rankItem strong,
  #gameScreen .battleItem strong{font-size:10px!important;white-space:nowrap!important;}
  #gameScreen .mapHeader{min-height:27px!important;margin-bottom:5px!important;}
  #gameScreen .mapTitleWrap h3{font-size:16px!important;}
  #gameScreen .teamPill{font-size:10px!important;padding:6px 9px!important;}
  #gameScreen #mapWrap{padding:5px!important;}
  #touchMovePad{width:118px!important;height:118px!important;margin:-59px 0 0 -59px!important;}
  #touchMovePad .joyKnob{width:52px!important;height:52px!important;margin:-26px 0 0 -26px!important;}
}

@media (orientation:portrait) and (pointer:coarse){
  /* Portrait stays focused on the map; side panels remain hidden to avoid crowding. */
  #gameScreen .studentCol{display:none!important;}
}



/* ===== v2.3 mobile landscape cleanup =====
   Left side is a touch-only joystick zone again. Right side keeps rankings and
   uses the former event-log area for the leave button. */
@media (orientation:landscape) and (pointer:coarse){
  #gameScreen{
    grid-template-columns:clamp(102px,15vw,166px) minmax(0,1fr) clamp(118px,16vw,174px)!important;
    gap:6px!important;
  }
  #studentLeft{
    display:block!important;
    min-width:0!important;
    min-height:0!important;
    overflow:visible!important;
    padding:0!important;
    pointer-events:none!important;
    background:transparent!important;
  }
  #studentLeft > .panel,
  #studentLeft > .leaveWrap{
    display:none!important;
  }
  #studentRight{
    display:flex!important;
    flex-direction:column!important;
    gap:6px!important;
    min-width:0!important;
    min-height:0!important;
    overflow:hidden!important;
    padding:0!important;
  }
  #studentRight .panel:nth-of-type(1){flex:1 1 42%!important;min-height:84px!important;}
  #studentRight .panel:nth-of-type(2){flex:0 1 30%!important;min-height:62px!important;}
  #studentRight .panel:nth-of-type(3){display:none!important;}
  #studentRight .leaveWrap.mobileLeaveDock{
    display:block!important;
    flex:0 0 42px!important;
    min-height:42px!important;
    padding:0!important;
    margin:0!important;
  }
  #studentRight .leaveWrap.mobileLeaveDock #leaveBtn{
    width:100%!important;
    height:42px!important;
    min-height:42px!important;
    border-radius:14px!important;
    font-size:14px!important;
    font-weight:1000!important;
  }
  #touchMovePad{z-index:55!important;}
  .mapZoomControls{display:inline-flex!important;}
}


/* ===== v2.4 mobile landscape map-left expansion =====
   Remove the empty left joystick column. The map now starts at the left edge
   so the joystick is detected on the actual map/card area instead of a black gutter. */
@media (orientation:landscape) and (pointer:coarse){
  #gameScreen{
    grid-template-columns:minmax(0,1fr) clamp(118px,16vw,174px)!important;
    grid-template-rows:1fr!important;
    gap:6px!important;
    padding:5px 5px 5px 0!important;
  }
  #studentLeft{
    display:none!important;
  }
  #studentCenter{
    grid-column:1!important;
    min-width:0!important;
    width:100%!important;
    height:100%!important;
  }
  #studentRight{
    grid-column:2!important;
  }
  #gameScreen .studentMapPanel{
    border-top-left-radius:0!important;
    border-bottom-left-radius:0!important;
    padding-left:7px!important;
  }
  #gameScreen #mapWrap{
    width:100%!important;
  }
}



/* ===== v3.2 mobile portrait answer-label visibility fix =====
   On narrow phones the generic qitem input width rule could squeeze the radio
   column, making the word "정답" look clipped. Keep the answer selector compact
   but give it a real column width in portrait mode. */
@media (orientation:portrait) and (pointer:coarse){
  #prepScreen .questionChoiceRow{
    display:grid!important;
    grid-template-columns:52px minmax(0,1fr) 74px!important;
    grid-template-areas:none!important;
    column-gap:6px!important;
    row-gap:6px!important;
    align-items:center!important;
    width:100%!important;
  }
  #prepScreen .questionChoiceRow .badge{
    grid-area:auto!important;
    min-width:48px!important;
    padding:7px 5px!important;
    text-align:center!important;
    white-space:nowrap!important;
  }
  #prepScreen .questionChoiceRow input[data-field='choice']{
    grid-area:auto!important;
    min-width:0!important;
    width:100%!important;
    max-width:100%!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel{
    grid-area:auto!important;
    justify-self:end!important;
    justify-content:flex-end!important;
    width:70px!important;
    min-width:70px!important;
    max-width:70px!important;
    overflow:visible!important;
    gap:5px!important;
    font-size:13px!important;
    white-space:nowrap!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel input[type='radio']{
    width:18px!important;
    height:18px!important;
    min-width:18px!important;
    max-width:18px!important;
    flex:0 0 18px!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel span{
    display:inline-block!important;
    min-width:28px!important;
    overflow:visible!important;
  }
}

@media (orientation:portrait) and (pointer:coarse) and (max-width:360px){
  #prepScreen .questionChoiceRow{grid-template-columns:48px minmax(0,1fr) 68px!important;gap:5px!important;}
  #prepScreen .questionChoiceRow .answerRadioLabel{width:66px!important;min-width:66px!important;max-width:66px!important;font-size:12px!important;gap:4px!important;}
}


/* ===== v3.3 mobile ceremony layout fix =====
   Phone portrait should read as a simple ranking list: 1st -> 2nd -> 3rd.
   Hide decorative header mascots because they take too much vertical space. */
@media (max-width: 640px), (max-width: 880px) and (orientation: portrait){
  #endScreen{
    padding:8px 6px 14px!important;
    height:auto!important;
    min-height:calc(100vh - 64px)!important;
    overflow:auto!important;
  }
  #endScreen .awardShell{
    width:calc(100vw - 12px)!important;
    padding:12px!important;
    border-radius:18px!important;
  }
  #endScreen .awardHeader{
    display:block!important;
    grid-template-columns:1fr!important;
    margin-bottom:8px!important;
  }
  #endScreen .awardHeader .brandMascotWrap{
    display:none!important;
  }
  #endScreen .awardKicker{
    font-size:11px!important;
    padding:5px 12px!important;
  }
  #endScreen .awardMainTitle{
    font-size:26px!important;
    line-height:1.08!important;
    margin:6px 0 1px!important;
  }
  #endScreen .awardSubtitle{
    font-size:11px!important;
  }
  #endScreen .teamAwardBanner{
    padding:10px!important;
    border-radius:16px!important;
    gap:9px!important;
    margin-bottom:8px!important;
  }
  #endScreen .teamAwardIcon{
    font-size:24px!important;
  }
  #endScreen .teamAwardText strong{
    font-size:16px!important;
  }
  #endScreen .awardGrid{
    display:grid!important;
    grid-template-columns:1fr!important;
    gap:10px!important;
    margin-top:8px!important;
  }
  #endScreen .awardCard{
    padding:10px!important;
    border-radius:18px!important;
  }
  #endScreen .awardSectionTitle{
    font-size:15px!important;
    margin-bottom:8px!important;
  }
  #endScreen .podiumStage{
    display:flex!important;
    flex-direction:column!important;
    height:auto!important;
    gap:9px!important;
    padding:0!important;
  }
  #endScreen .podiumStage::before,
  #endScreen .winnerGlow,
  #endScreen .winnerConfetti{
    display:none!important;
  }
  #endScreen .podiumSpot.rank1{order:1!important;}
  #endScreen .podiumSpot.rank2{order:2!important;}
  #endScreen .podiumSpot.rank3{order:3!important;}
  #endScreen .podiumSpot{
    width:100%!important;
    min-height:84px!important;
    margin:0!important;
    padding:10px!important;
    display:grid!important;
    grid-template-columns:58px minmax(0,1fr) 56px!important;
    grid-template-areas:
      "char medal base"
      "char name base"
      "char meta score"!important;
    align-items:center!important;
    column-gap:10px!important;
    border-radius:16px!important;
    background:linear-gradient(180deg,rgba(255,255,255,.10),rgba(255,255,255,.045))!important;
    border:1px solid rgba(255,255,255,.16)!important;
  }
  #endScreen .podiumCharacter,
  #endScreen .rank1 .podiumCharacter{
    grid-area:char!important;
    width:52px!important;
    height:52px!important;
    margin:0!important;
  }
  #endScreen .podiumCharacterImg,
  #endScreen .rank1 .podiumCharacterImg{
    width:52px!important;
    height:52px!important;
    object-fit:contain!important;
  }
  #endScreen .medalBadge{
    grid-area:medal!important;
    justify-self:start!important;
    margin:0!important;
    padding:4px 8px!important;
    font-size:11px!important;
    line-height:1!important;
  }
  #endScreen .podiumName{
    grid-area:name!important;
    margin:0!important;
    text-align:left!important;
    font-size:15px!important;
    line-height:1.15!important;
    width:100%!important;
  }
  #endScreen .podiumMeta{
    grid-area:meta!important;
    margin:0!important;
    text-align:left!important;
    font-size:10px!important;
    line-height:1.25!important;
  }
  #endScreen .podiumScore,
  #endScreen .rank1 .podiumScore{
    grid-area:score!important;
    justify-self:end!important;
    margin:0!important;
    font-size:16px!important;
    line-height:1!important;
  }
  #endScreen .podiumBase,
  #endScreen .rank1 .podiumBase,
  #endScreen .rank2 .podiumBase,
  #endScreen .rank3 .podiumBase{
    grid-area:base!important;
    justify-self:end!important;
    width:50px!important;
    height:50px!important;
    min-height:0!important;
    margin:0!important;
    border-radius:14px!important;
    font-size:28px!important;
  }
  #endScreen .emptyPodium{
    grid-column:1 / -1!important;
    height:auto!important;
    min-height:68px!important;
    padding:14px!important;
    border-radius:14px!important;
  }
  #endScreen .fullRankList,
  #endScreen .awardLogList{
    max-height:none!important;
  }
  #endScreen .mvpGrid{
    gap:7px!important;
  }
  #endScreen .mvpCard,
  #endScreen .rankItem,
  #endScreen .logItem{
    padding:8px!important;
    border-radius:13px!important;
  }
  #endScreen .awardFooter{
    margin-top:10px!important;
  }
}

/* ===== v3.17 mobile landscape ceremony order fix =====
   On phone landscape, keep results as 1st -> 2nd -> 3rd instead of podium-style visual order. */
@media (pointer:coarse) and (orientation:landscape){
  #endScreen .brandMascotWrap{display:none!important;}
  #endScreen .awardHeader{display:block!important;margin-bottom:8px!important;}
  #endScreen .awardMainTitle{font-size:30px!important;line-height:1.08!important;}
  #endScreen .awardGrid{grid-template-columns:1fr 1fr!important;gap:10px!important;}
  #endScreen .podiumStage{display:flex!important;flex-direction:column!important;height:auto!important;gap:8px!important;padding:0!important;}
  #endScreen .podiumStage::before,#endScreen .winnerGlow,#endScreen .winnerConfetti{display:none!important;}
  #endScreen .podiumSpot.rank1{order:1!important;}
  #endScreen .podiumSpot.rank2{order:2!important;}
  #endScreen .podiumSpot.rank3{order:3!important;}
  #endScreen .podiumSpot{width:100%!important;min-height:72px!important;margin:0!important;padding:8px!important;display:grid!important;grid-template-columns:52px minmax(0,1fr) 52px!important;grid-template-areas:"char medal base" "char name base" "char meta score"!important;align-items:center!important;column-gap:8px!important;border-radius:15px!important;background:linear-gradient(180deg,rgba(255,255,255,.10),rgba(255,255,255,.045))!important;border:1px solid rgba(255,255,255,.16)!important;}
  #endScreen .podiumCharacter,#endScreen .rank1 .podiumCharacter{grid-area:char!important;width:48px!important;height:48px!important;margin:0!important;}
  #endScreen .podiumCharacterImg,#endScreen .rank1 .podiumCharacterImg{width:48px!important;height:48px!important;object-fit:contain!important;}
  #endScreen .medalBadge{grid-area:medal!important;justify-self:start!important;margin:0!important;padding:4px 8px!important;font-size:11px!important;}
  #endScreen .podiumName{grid-area:name!important;text-align:left!important;margin:0!important;font-size:14px!important;}
  #endScreen .podiumMeta{grid-area:meta!important;text-align:left!important;margin:0!important;font-size:10px!important;}
  #endScreen .podiumScore,#endScreen .rank1 .podiumScore{grid-area:score!important;justify-self:end!important;margin:0!important;font-size:15px!important;}
  #endScreen .podiumBase,#endScreen .rank1 .podiumBase,#endScreen .rank2 .podiumBase,#endScreen .rank3 .podiumBase{grid-area:base!important;justify-self:end!important;width:46px!important;height:46px!important;min-height:0!important;margin:0!important;border-radius:13px!important;font-size:26px!important;}
}


/* ===== v3.28 desktop ceremony podium order =====
   PC: 2nd - 1st - 3rd visual podium. Mobile rules below/above keep 1st - 2nd - 3rd. */
@media (min-width:881px) and (pointer:fine){
  #endScreen .podiumStage .podiumSpot.rank2{order:1!important;}
  #endScreen .podiumStage .podiumSpot.rank1{order:2!important;}
  #endScreen .podiumStage .podiumSpot.rank3{order:3!important;}
}

/* Make student question choices look like printed exam numbers. */
#prepScreen .choiceBadge{font-size:17px!important;font-weight:1000!important;color:#0f355d!important;background:linear-gradient(180deg,#e0fbff,#c8f4ff)!important;border:1px solid rgba(34,211,238,.38)!important;}

/* ===== v3.20 mobile landscape answer-label visibility fix =====
   On phone landscape screens the answer column could be too narrow, so the word
   "정답" was clipped. Give the radio label a real width and shrink only the
   circled choice number column. */
@media (orientation:landscape) and (pointer:coarse) and (max-height:560px){
  #prepScreen .questionChoiceRow{
    display:grid!important;
    grid-template-columns:42px minmax(0,1fr) 78px!important;
    column-gap:6px!important;
    align-items:center!important;
    overflow:visible!important;
  }
  #prepScreen .questionChoiceRow .choiceBadge{
    width:34px!important;
    min-width:34px!important;
    max-width:34px!important;
    height:30px!important;
    min-height:30px!important;
    padding:0!important;
    justify-self:center!important;
    font-size:16px!important;
    line-height:1!important;
  }
  #prepScreen .questionChoiceRow input[data-field='choice']{
    min-width:0!important;
    width:100%!important;
    max-width:100%!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel{
    width:76px!important;
    min-width:76px!important;
    max-width:76px!important;
    justify-self:end!important;
    justify-content:flex-end!important;
    overflow:visible!important;
    white-space:nowrap!important;
    gap:5px!important;
    font-size:13px!important;
    line-height:1!important;
    box-sizing:border-box!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel input[type='radio']{
    width:18px!important;
    height:18px!important;
    min-width:18px!important;
    max-width:18px!important;
    flex:0 0 18px!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel span{
    display:inline-block!important;
    min-width:30px!important;
    overflow:visible!important;
  }
}


/* ===== v3.15 mobile end-screen safe-area fix =====
   Keep the Home button above Android/iOS browser navigation bars. */
@media (max-width: 640px), (max-width: 880px) and (orientation: portrait){
  #endScreen{
    min-height:calc(100dvh - 64px)!important;
    padding-bottom:max(92px, calc(env(safe-area-inset-bottom, 0px) + 92px))!important;
    scroll-padding-bottom:max(92px, calc(env(safe-area-inset-bottom, 0px) + 92px))!important;
  }
  #endScreen .awardShell{
    margin-bottom:max(74px, calc(env(safe-area-inset-bottom, 0px) + 74px))!important;
  }
  #endScreen .awardFooter{
    position:sticky!important;
    bottom:max(14px, calc(env(safe-area-inset-bottom, 0px) + 14px))!important;
    z-index:8!important;
    margin-top:14px!important;
    padding:8px 0 max(8px, env(safe-area-inset-bottom, 0px))!important;
    background:linear-gradient(180deg,rgba(235,244,255,0),rgba(235,244,255,.94) 35%,rgba(235,244,255,.98))!important;
    border-radius:18px!important;
  }
  #endHomeBtn{
    min-height:50px!important;
    width:min(240px,82vw)!important;
    padding:0 24px!important;
    font-size:15px!important;
    font-weight:1000!important;
    line-height:50px!important;
    white-space:nowrap!important;
  }
}


/* ===== v3.5 PC prep question editor expansion =====
   On desktop, use the empty vertical space on the right side so the question
   editor feels like a full work area instead of a small inner scroll box. */
@media (min-width:881px) and (pointer:fine){
  #prepScreen .card{
    width:min(1080px,95vw)!important;
  }
  #prepScreen .prepLayout{
    display:grid!important;
    grid-template-columns:minmax(320px,36%) minmax(430px,1fr)!important;
    gap:14px!important;
    align-items:stretch!important;
    flex-wrap:nowrap!important;
  }
  #prepScreen .prepInfoPanel,
  #prepScreen .prepQuestionPanel{
    min-height:min(590px,calc(100vh - 220px))!important;
  }
  #prepScreen .prepQuestionPanel{
    display:flex!important;
    flex-direction:column!important;
  }
  #prepScreen .prepQuestionPanel > div:first-child,
  #prepScreen .prepQuestionPanel > .mini{
    flex:0 0 auto!important;
  }
  #prepScreen #questionsList{
    flex:1 1 auto!important;
    min-height:0!important;
    max-height:none!important;
    overflow:auto!important;
    padding-right:8px!important;
  }
  #prepScreen #questionsList .qitem{
    min-height:100%!important;
    display:flex!important;
    flex-direction:column!important;
    gap:8px!important;
  }
  #prepScreen #questionsList .qitem textarea{
    min-height:118px!important;
  }
}


/* ===== v3.21 mobile landscape answer-label hard fix =====
   The previous landscape rule was limited by viewport height, so some mobile
   browsers still clipped the Korean word "정답". This final override applies to
   all phone/tablet landscape layouts and reserves a dedicated answer column. */
@media (orientation: landscape) and (pointer: coarse),
       (orientation: landscape) and (max-width: 980px){
  #prepScreen .questionChoiceRow{
    display:grid!important;
    grid-template-columns:34px minmax(0,1fr) 92px!important;
    grid-template-areas:"num choice answer"!important;
    column-gap:6px!important;
    align-items:center!important;
    width:100%!important;
    max-width:100%!important;
    box-sizing:border-box!important;
    overflow:visible!important;
    padding-right:4px!important;
  }
  #prepScreen .questionChoiceRow .choiceBadge,
  #prepScreen .questionChoiceRow .badge{
    grid-area:num!important;
    width:30px!important;
    min-width:30px!important;
    max-width:30px!important;
    height:30px!important;
    min-height:30px!important;
    padding:0!important;
    justify-self:center!important;
    font-size:15px!important;
    line-height:1!important;
    white-space:nowrap!important;
  }
  #prepScreen .questionChoiceRow input[data-field='choice']{
    grid-area:choice!important;
    min-width:0!important;
    width:100%!important;
    max-width:100%!important;
    box-sizing:border-box!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel{
    grid-area:answer!important;
    display:inline-flex!important;
    align-items:center!important;
    justify-content:flex-end!important;
    width:90px!important;
    min-width:90px!important;
    max-width:90px!important;
    flex:0 0 90px!important;
    gap:5px!important;
    padding:0 2px!important;
    margin:0!important;
    overflow:visible!important;
    white-space:nowrap!important;
    box-sizing:border-box!important;
    font-size:12px!important;
    line-height:1!important;
    letter-spacing:-0.2px!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel input[type='radio']{
    width:18px!important;
    height:18px!important;
    min-width:18px!important;
    max-width:18px!important;
    flex:0 0 18px!important;
    margin:0!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel span{
    display:inline-block!important;
    width:auto!important;
    min-width:28px!important;
    max-width:none!important;
    flex:0 0 auto!important;
    overflow:visible!important;
    white-space:nowrap!important;
    text-overflow:clip!important;
  }
}
@media (orientation: landscape) and (pointer: coarse) and (max-width: 700px){
  #prepScreen .questionChoiceRow{
    grid-template-columns:30px minmax(0,1fr) 82px!important;
    column-gap:5px!important;
  }
  #prepScreen .questionChoiceRow .choiceBadge,
  #prepScreen .questionChoiceRow .badge{
    width:28px!important;
    min-width:28px!important;
    max-width:28px!important;
    height:28px!important;
    min-height:28px!important;
    font-size:14px!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel{
    width:80px!important;
    min-width:80px!important;
    max-width:80px!important;
    flex-basis:80px!important;
    gap:4px!important;
    font-size:11px!important;
  }
  #prepScreen .questionChoiceRow .answerRadioLabel input[type='radio']{
    width:17px!important;
    height:17px!important;
    min-width:17px!important;
    max-width:17px!important;
    flex-basis:17px!important;
  }
}


/* ===== v3.26 teacher credit footer patch ===== */
@media(max-width:1180px){
  #createScreen .teacherCredit{
    text-align:center!important;
    margin-top:12px!important;
    padding:10px 12px!important;
    border-radius:14px!important;
    background:rgba(255,255,255,.50)!important;
    border:1px solid rgba(91,141,204,.18)!important;
    color:#35516f!important;
    font-size:12px!important;
  }
  #createScreen .teacherCredit div{white-space:normal!important;}
}
@media(min-width:1181px){
  #createScreen .teacherCredit{
    max-width:520px;
    margin-left:auto;
    padding-right:4px;
  }
}

/* ===== v3.27 multi-teacher room owner layout ===== */
.teacherRoomMetaRow{display:grid!important;grid-template-columns:minmax(0,1.4fr) minmax(180px,.6fr)!important;gap:10px!important;}
@media(max-width:640px){.teacherRoomMetaRow{grid-template-columns:1fr!important;}}

.aiReviewPanel{border-color:rgba(34,211,238,.30)!important;background:linear-gradient(180deg,rgba(14,165,233,.11),rgba(255,255,255,.052))!important}
.aiReviewActions{display:flex;gap:8px;align-items:center;margin:8px 0 10px}.aiReviewActions button{width:100%;min-height:40px;padding:8px 10px;font-size:14px}.aiReviewBox{display:grid;gap:8px;max-height:220px;overflow:auto;padding-right:3px}.aiReviewSummary{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:4px}.aiPill{display:inline-flex;align-items:center;gap:4px;border-radius:999px;padding:5px 8px;font-size:11px;font-weight:1000;border:1px solid rgba(255,255,255,.14);background:rgba(255,255,255,.08)}.aiPill.ok{color:#dcfce7;background:rgba(22,163,74,.20)}.aiPill.warn{color:#fef3c7;background:rgba(245,158,11,.22)}.aiPill.bad{color:#fee2e2;background:rgba(239,68,68,.20)}.aiReviewItem{border:1px solid rgba(255,255,255,.10);background:rgba(255,255,255,.065);border-radius:13px;padding:8px}.aiReviewItem b{display:block;color:#0f2544;font-size:13px;margin-bottom:4px}.aiReviewItem .mini{font-size:11px}.aiStatus{display:inline-block;border-radius:999px;padding:3px 7px;margin-right:5px;font-size:11px;font-weight:1000}.aiStatus.ok{background:rgba(34,197,94,.20);color:#dcfce7}.aiStatus.warn{background:rgba(245,158,11,.20);color:#fef3c7}.aiStatus.bad{background:rgba(239,68,68,.22);color:#fee2e2}.aiStatus.muted{background:rgba(148,163,184,.18);color:#e2e8f0}.aiReviewWarning{border:1px solid rgba(245,158,11,.28);border-radius:12px;padding:7px 8px;background:rgba(245,158,11,.12);color:#78350f;font-weight:900}
#aiReviewBtn{background:linear-gradient(135deg,#22c55e,#0891b2 60%,#2563eb)!important;color:#fff!important;box-shadow:0 12px 24px rgba(14,165,233,.20)!important}
@media(max-width:1180px){.aiReviewBox{max-height:180px}.aiReviewActions{margin-top:6px}}


.winnerCrown{position:absolute;left:50%;top:-20px;transform:translateX(-50%) rotate(-7deg);z-index:6;font-size:34px;filter:drop-shadow(0 7px 9px rgba(0,0,0,.28)) drop-shadow(0 0 10px rgba(251,191,36,.52));animation:crownFloat 1.8s ease-in-out infinite alternate;pointer-events:none}
@keyframes crownFloat{from{transform:translateX(-50%) translateY(0) rotate(-7deg)}to{transform:translateX(-50%) translateY(-3px) rotate(6deg)}}
@media(max-width:760px){.winnerCrown{font-size:26px;top:-15px}}


/* v3.32 teacher mobile AI review visibility */
@media(max-width:1180px){
  #operateScreen .rightCol{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;overflow:visible!important;padding-right:0!important;}
  #operateScreen .rightCol .aiReviewPanel{order:99!important;width:100%!important;display:block!important;margin-bottom:10px!important;}
  #operateScreen .aiReviewBox{max-height:none!important;overflow:visible!important;}
}
@media(max-width:820px) and (orientation:landscape){
  #operateScreen .rightCol{grid-template-columns:1fr 1fr!important;align-items:start!important;}
  #operateScreen .rightCol .aiReviewPanel{grid-column:1 / -1!important;}
  #operateScreen .rightCol .aiReviewBox{max-height:160px!important;overflow:auto!important;}
}


/* ===== v3.34 student info compact grid + teacher setup 3-row patch ===== */
/* Student prep: title stays wide, the six small room/nickname fields form 3 columns x 2 rows on phones. */
#prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:1 / -1;}
#prepScreen .studentRoomInfoGrid .compactInfoCard .value{word-break:keep-all;}
@media (orientation:portrait) and (pointer:coarse), (max-width:760px){
  #prepScreen .studentRoomInfoGrid{
    grid-template-columns:repeat(3,minmax(0,1fr))!important;
    gap:8px!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard{
    grid-column:1 / -1!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard{
    min-height:76px!important;
    padding:10px 8px!important;
    display:flex!important;
    flex-direction:column!important;
    justify-content:center!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .label{
    font-size:clamp(10px,2.9vw,12px)!important;
    line-height:1.15!important;
    margin-bottom:6px!important;
    white-space:normal!important;
    letter-spacing:-.45px!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{
    font-size:clamp(18px,5vw,28px)!important;
    line-height:1.05!important;
    letter-spacing:-1.2px!important;
    white-space:nowrap!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard .value{
    font-size:clamp(21px,6.2vw,34px)!important;
  }
}
@media (orientation:portrait) and (pointer:coarse) and (max-width:380px){
  #prepScreen .studentRoomInfoGrid{gap:6px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard{padding:8px 6px!important;min-height:68px!important;border-radius:14px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{font-size:clamp(16px,4.7vw,22px)!important;}
}
/* Student landscape keeps the previous compact side panel to avoid squeezing the question editor. */
@media (orientation:landscape) and (pointer:coarse){
  #prepScreen .studentRoomInfoGrid{grid-template-columns:1fr!important;}
  #prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:auto!important;}
}
/* Teacher PC setup: compact numeric settings become 4 columns, so 10 fields render as 3 rows instead of 5. */
@media (min-width:821px){
  #createScreen .teacherSettingsGrid{
    grid-template-columns:repeat(4,minmax(0,1fr))!important;
    gap:10px!important;
  }
  #createScreen .teacherSettingsGrid .field input{
    min-height:40px!important;
    padding:8px 10px!important;
  }
}

</style>
</head>
<body>
<div class='topbar'><div class='brand'>ReMap</div><div id='roomTitleBar' class='roomTitle'>REMAP</div><div id='statusBar' class='statusBadge'>[현재상황: 준비 중]</div><button id='topLeaveBtn' class='topLeaveBtn' type='button'>나가기</button></div>
<div id='joinScreen'>
  <div class='card' style='max-width:560px'>
    <h1 style='margin-top:0;color:#173b7a'>방 입장</h1>
    <div class='mini'>방 코드와 닉네임을 먼저 입력하세요. 닉네임은 방장에게 즉시 표시됩니다.</div>
    <div class='stack' style='margin-top:16px'>
      <input id='roomCode' placeholder='방 코드' maxlength='4' style='text-transform:uppercase' />
      <input id='nickname' placeholder='닉네임' maxlength='12' />
      <button id='checkRoomBtn'>다음</button>
    </div>
    <div id='roomInfo' class='mini' style='margin-top:10px'></div>
  </div>
</div>
<div id='prepScreen'>
  <div class='card'>
    <h1 style='margin-top:0;color:#173b7a'>참가 준비</h1>
    <div class='mini'>방 정보를 확인한 뒤 팀 선택과 문제 입력을 완료하세요.</div>
    <div class='row prepLayout'>
      <div class='panel prepInfoPanel'>
        <h3>방 정보</h3>
        <div class='infoGrid studentRoomInfoGrid'>
          <div class='infoCard infoTitleCard'><div class='label'>방 제목</div><div class='value' id='infoTitle'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>방 코드</div><div class='value' id='infoCode'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>게임 모드</div><div class='value' id='infoMode'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>배틀당 문제 수</div><div class='value' id='infoQuestions'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>문제당 제한시간</div><div class='value' id='infoTime'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>맵 종류</div><div class='value' id='infoMap'>-</div></div>
          <div class='infoCard compactInfoCard'><div class='label'>닉네임</div><div class='value' id='infoNickname'>-</div></div>
        </div>
        <div id='teamSection' style='margin-top:12px'><div style='font-weight:700;color:#173b7a'>팀 선택</div><div id='teamWrap' class='teamWrap'></div></div>
        <div id='colorHelp' style='margin-top:12px' class='mini'>개인전에서는 색상을 직접 고를 수 있습니다. 팀전에서는 팀 색상이 고정됩니다.</div>
        <div id='colorWrap' class='colorWrap'></div>
      </div>
      <div class='panel prepQuestionPanel'>
        <div style='display:flex;justify-content:space-between;align-items:center'><h3>문제 입력</h3><span class='badge' id='requiredCount'>문제 0개 필수</span></div>
        <div class='mini'>이 방은 설정된 문제 수만큼 정확히 입력해야 입장할 수 있습니다.</div>
        <div id='questionsList'></div>
      </div>
    </div>
    <div class='row prepActions'>
      <button id='backBtn' class='ghost'>나가기</button>
      <button id='joinBtn'>입장 완료</button>
    </div>
  </div>
</div>
<div id='gameScreen'>
  <aside id='studentLeft' class='studentCol'>
    <div class='panel'><h3>내 정보</h3><div id='topInfo'></div></div>
    <div class='panel'><h3>현재 배틀 진행</h3><div id='battleList'></div></div>
    <div class='leaveWrap'><button id='leaveBtn' class='ghost'>나가기</button></div>
  </aside>
  <main id='studentCenter' class='studentMain'>
    <div class='panel studentMapPanel'>
      <div class='mapHeader'><div class='mapTitleWrap'><h3>게임 맵</h3><div class='mini'>화살표 키 또는 WASD로 이동하며 친구를 만나 퀴즈 배틀을 시작합니다.</div></div><span class='teamPill'>실시간 이동</span></div>
      <div id='mapWrap'><div id='mapZoomControls' class='mapZoomControls' aria-label='맵 크기 조절'><button id='mapZoomOutBtn' type='button'>−</button><span id='mapZoomValue'>100%</span><button id='mapZoomInBtn' type='button'>＋</button></div><canvas id='gameCanvas' width='1060' height='612'></canvas></div>
    </div>
  </main>
  <aside id='studentRight' class='studentCol'>
    <div class='panel'><h3>개인 순위</h3><div id='rankingList'></div></div>
    <div class='panel'><h3>팀 순위</h3><div id='teamRankingList'></div></div>
    <div class='panel'><h3>이벤트 로그</h3><div id='logList'></div></div>
  </aside>
</div>
<div id='touchMovePad' aria-label='모바일 이동 조이스틱'>
  <div class='joyBase'></div>
  <div class='joyHint'></div>
  <div class='joyKnob'></div>
</div>
<div id='battleOverlay'><div id='battleCard'><div class='battleHeader'><div><h2 id='battleTitle' class='battleTitleText'>배틀 시작</h2><div class='mini' id='battleMeta'></div></div><div class='battleTimerPill' id='battleTimer'>제한시간</div></div><div class='battleProgressBar'><span id='battleProgressFill' style='width:0%'></span></div><div class='battleQuestionBox' id='battleQuestion'></div><div id='battleChoices'></div></div></div>
<div id='battleEffect'></div>
<div id='resultOverlay'><div class='resultCard' id='resultCard'><h2 id='resultTitle'>배틀 결과</h2><div id='resultText' style='margin:12px 0'></div><button id='resultConfirmBtn'>확인</button></div></div>
<div id='countdownOverlay'><div id='countdownText'>3</div></div>
<div id='toast' class='bottom'></div>
<div id='endScreen'>
  <div class='card awardShell'>
    <div class='awardHeader'>
      <div class='brandMascotWrap'><img class='brandMascot' id='headerMascotLeft' src='/character/60a5fa.svg' alt='ReMap 캐릭터'></div>
      <div class='awardTitleBox'>
        <div class='awardKicker'><span>🏆</span><span>FINAL CEREMONY</span></div>
        <h1 class='awardMainTitle'>최종 결과 시상식</h1>
        <div class='awardSubtitle'>오늘의 배틀 결과를 확인해요</div>
      </div>
      <div class='brandMascotWrap'><img class='brandMascot' id='headerMascotRight' src='/character/60a5fa.svg' alt='ReMap 캐릭터'></div>
    </div>
    <div id='finalTeamRanks'></div>
    <div class='awardGrid'>
      <div class='awardCard'><div id='finalRanks'></div></div>
      <div class='awardCard'><div id='finalMvp'></div><h3 class='awardSectionTitle' style='margin-top:14px'>최근 로그</h3><div id='finalLogs' class='awardLogList'></div></div>
    </div>
    <div class='awardFooter'><button id='endHomeBtn' class='ghost'>홈으로</button></div>
  </div>
</div>
<script>
const CHAR_COLORS = __CHAR_COLORS__;
const TEAM_COLORS = {'A':'#3b82f6','B':'#ef4444','C':'#facc15','D':'#22c55e','E':'#94a3b8'};
const state={socket:null,playerId:null,reconnectToken:null,myPlayer:null,players:[],rankings:[],teamRankings:[],battles:[],settings:null,gameStatus:'idle',remainingTime:0,logs:[],keys:{},questions:[],currentBattle:null,timerHandle:null,bgImage:null,selectedTeam:'A',selectedColor:CHAR_COLORS[0],roomCode:'',roomTitle:'',nickname:'',autoResultHide:null,roomInfo:null,pendingId:null,mapWalls:[],countdownActive:false,countdownTimer:null,renderPlayers:{},lastFrameTs:0,lastMoveSentAt:0,animationFrameId:null,gameLoopRunning:false,intentionalClose:false,reconnectAttempts:0,reconnectTimer:null};
const roomCodeInput=document.getElementById('roomCode'), nicknameInput=document.getElementById('nickname');
const questionsList=document.getElementById('questionsList'), joinBtn=document.getElementById('joinBtn'), checkRoomBtn=document.getElementById('checkRoomBtn');
const joinScreen=document.getElementById('joinScreen'), prepScreen=document.getElementById('prepScreen'), gameScreen=document.getElementById('gameScreen'), endScreen=document.getElementById('endScreen');
const canvas=document.getElementById('gameCanvas'), ctx=canvas.getContext('2d'), topInfo=document.getElementById('topInfo'), rankingList=document.getElementById('rankingList'), teamRankingList=document.getElementById('teamRankingList'), battleList=document.getElementById('battleList'), logList=document.getElementById('logList');
const battleOverlay=document.getElementById('battleOverlay'), battleCard=document.getElementById('battleCard'), battleTitle=document.getElementById('battleTitle'), battleMeta=document.getElementById('battleMeta'), battleTimer=document.getElementById('battleTimer'), battleQuestion=document.getElementById('battleQuestion'), battleChoices=document.getElementById('battleChoices'), battleProgressFill=document.getElementById('battleProgressFill'), battleEffect=document.getElementById('battleEffect');
const toast=document.getElementById('toast'), roomInfo=document.getElementById('roomInfo'), roomTitleBar=document.getElementById('roomTitleBar'), statusBar=document.getElementById('statusBar');
const initialRoomCode=new URLSearchParams(location.search).get('code');if(initialRoomCode){roomCodeInput.value=initialRoomCode.trim().toUpperCase().slice(0,4);}
const infoNickname=document.getElementById('infoNickname');
const countdownOverlay=document.getElementById('countdownOverlay'), countdownText=document.getElementById('countdownText');
const teamWrap=document.getElementById('teamWrap'), teamSection=document.getElementById('teamSection'), colorWrap=document.getElementById('colorWrap'), colorHelp=document.getElementById('colorHelp'), leaveBtn=document.getElementById('leaveBtn'), topLeaveBtn=document.getElementById('topLeaveBtn');
const resultOverlay=document.getElementById('resultOverlay'), resultCard=document.getElementById('resultCard'), resultTitle=document.getElementById('resultTitle'), resultText=document.getElementById('resultText'), resultConfirmBtn=document.getElementById('resultConfirmBtn');
const backBtn=document.getElementById('backBtn');
const touchMovePad=document.getElementById('touchMovePad');
const mapZoomOutBtn=document.getElementById('mapZoomOutBtn'), mapZoomInBtn=document.getElementById('mapZoomInBtn'), mapZoomValue=document.getElementById('mapZoomValue');
let studentMapScale=1;
function applyStudentMapScale(){document.documentElement.style.setProperty('--student-map-scale', String(studentMapScale));if(mapZoomValue)mapZoomValue.textContent=Math.round(studentMapScale*100)+'%';}
function changeStudentMapScale(delta){studentMapScale=Math.max(0.85,Math.min(1.35,Number((studentMapScale+delta).toFixed(2))));applyStudentMapScale();}
if(mapZoomOutBtn)mapZoomOutBtn.onclick=(ev)=>{ev.stopPropagation();changeStudentMapScale(-0.1);};
if(mapZoomInBtn)mapZoomInBtn.onclick=(ev)=>{ev.stopPropagation();changeStudentMapScale(0.1);};
applyStudentMapScale();
let toastTimer=null;
function setGameActive(active){document.body.classList.toggle('remap-game-active', !!active);}
function setupTouchMovePad(){
  if(!touchMovePad)return;
  const knob=touchMovePad.querySelector('.joyKnob');
  const radius=58;
  const deadZone=7;
  let activeId=null;
  let originX=0, originY=0;
  const isTouchLike=()=>window.matchMedia('(pointer: coarse)').matches || window.innerWidth<=900;
  const clear=()=>{
    activeId=null;
    state.touchDx=0;
    state.touchDy=0;
    touchMovePad.classList.remove('active');
    if(knob)knob.style.transform='translate(0px,0px)';
  };
  const update=(clientX,clientY)=>{
    const rawX=clientX-originX;
    const rawY=clientY-originY;
    const dist=Math.hypot(rawX,rawY);
    if(dist<deadZone){
      state.touchDx=0;
      state.touchDy=0;
      if(knob)knob.style.transform='translate(0px,0px)';
      return;
    }
    const clamped=Math.min(radius,dist);
    const nx=rawX/dist;
    const ny=rawY/dist;
    const knobX=nx*clamped;
    const knobY=ny*clamped;
    // 모바일 조이스틱은 버튼 조작처럼 일정한 속도로 움직이게 한다.
    // 손가락이 중심에서 조금만 벗어나도 방향은 유지하고, 속도는 기본 이동속도와 맞춘다.
    state.touchDx=Number(nx.toFixed(3));
    state.touchDy=Number(ny.toFixed(3));
    if(knob)knob.style.transform=`translate(${knobX}px,${knobY}px)`;
  };
  const canStart=(ev)=>{
    if(!isTouchLike())return false;
    if(!document.body.classList.contains('remap-game-active'))return false;
    if(gameScreen.style.display!=='grid')return false;
    if(state.gameStatus==='finished'||state.gameStatus==='countdown')return false;
    if(battleOverlay.style.display==='flex'||resultOverlay.style.display==='flex'||countdownOverlay.style.display==='flex')return false;
    return ev.clientX <= window.innerWidth*0.78;
  };
  const start=(ev)=>{
    if(!canStart(ev))return;
    ev.preventDefault();
    activeId=ev.pointerId;
    originX=ev.clientX;
    originY=ev.clientY;
    touchMovePad.style.left=originX+'px';
    touchMovePad.style.top=originY+'px';
    touchMovePad.classList.add('active');
    if(knob)knob.style.transform='translate(0px,0px)';
    update(ev.clientX,ev.clientY);
    try{gameScreen.setPointerCapture(ev.pointerId);}catch(e){}
  };
  const move=(ev)=>{
    if(activeId!==ev.pointerId)return;
    ev.preventDefault();
    update(ev.clientX,ev.clientY);
  };
  const end=(ev)=>{
    if(activeId!==null&&activeId!==ev.pointerId)return;
    clear();
  };
  gameScreen.addEventListener('pointerdown',start,{passive:false});
  gameScreen.addEventListener('pointermove',move,{passive:false});
  gameScreen.addEventListener('pointerup',end,{passive:false});
  gameScreen.addEventListener('pointercancel',end,{passive:false});
  window.addEventListener('blur',clear);
  window.addEventListener('orientationchange',clear);
}

function setupMobileLeaveDock(){
  const left=document.getElementById('studentLeft');
  const right=document.getElementById('studentRight');
  const leaveWrap=document.querySelector('#studentLeft .leaveWrap, #studentRight .leaveWrap');
  if(!left||!right||!leaveWrap)return;
  const marker=document.createComment('leave-wrap-home');
  if(!leaveWrap.__homeMarker){
    leaveWrap.parentNode.insertBefore(marker, leaveWrap);
    leaveWrap.__homeMarker=marker;
  }
  const update=()=>{
    const mobileLandscape=window.matchMedia('(orientation: landscape) and (pointer: coarse)').matches;
    if(mobileLandscape){
      if(leaveWrap.parentNode!==right)right.appendChild(leaveWrap);
      leaveWrap.classList.add('mobileLeaveDock');
    }else{
      const home=leaveWrap.__homeMarker;
      if(home&&home.parentNode&&leaveWrap.parentNode!==home.parentNode){
        home.parentNode.insertBefore(leaveWrap, home.nextSibling);
      }
      leaveWrap.classList.remove('mobileLeaveDock');
    }
  };
  update();
  window.addEventListener('resize',update);
  window.addEventListener('orientationchange',()=>setTimeout(update,100));
}
setupMobileLeaveDock();
setupTouchMovePad();
function showToast(msg, position='bottom', tone='neutral'){const posClass=position==='center'?'center':position==='feedback'?'feedback':'bottom';toast.textContent=msg;toast.className=posClass + (tone==='success'?' success':tone==='error'?' error':'');toast.style.display='block';if(toastTimer)clearTimeout(toastTimer);const duration=(position==='center'||position==='feedback')?1200:2200;toastTimer=setTimeout(()=>toast.style.display='none', duration)}
function pulseCountdownText(value){countdownText.className=value==='START'?'start':'';countdownText.textContent=value;countdownText.style.animation='none';void countdownText.offsetWidth;countdownText.style.animation='countPulse .82s ease-out';}
function startCountdownOverlay(seconds=4){if(state.countdownActive)return;state.countdownActive=true;const seq=[];for(let n=Math.max(1,seconds-1);n>=1;n--){seq.push(String(n));}seq.push('START');let idx=0;countdownOverlay.style.display='flex';pulseCountdownText(seq[0]||'START');if(state.countdownTimer)clearInterval(state.countdownTimer);state.countdownTimer=setInterval(()=>{idx+=1;if(idx>=seq.length){clearInterval(state.countdownTimer);state.countdownTimer=null;state.countdownActive=false;countdownOverlay.style.display='none';countdownText.className='';return;}pulseCountdownText(seq[idx]);},1000);}function stopGameLoop(){state.gameLoopRunning=false;if(state.animationFrameId!==null){cancelAnimationFrame(state.animationFrameId);state.animationFrameId=null;}state.lastFrameTs=0;}function startGameLoop(){stopGameLoop();state.gameLoopRunning=true;state.animationFrameId=requestAnimationFrame(gameLoop);}function clearPlayTimers(){if(state.countdownTimer)clearInterval(state.countdownTimer);state.countdownTimer=null;state.countdownActive=false;if(state.timerHandle)clearInterval(state.timerHandle);state.timerHandle=null;if(state.autoResultHide)clearTimeout(state.autoResultHide);state.autoResultHide=null;}
function buildEmptyQuestions(count){return Array.from({length:count},()=>({text:'',choices:['','','',''],answer:0}))}
function ensureQuestionCount(count){const required=Math.max(0,Number(count)||0);while(state.questions.length<required){state.questions.push({text:'',choices:['','','',''],answer:0});}if(state.questions.length>required){state.questions=state.questions.slice(0,required);}}
function renderTeams(count=4){teamWrap.innerHTML=''; ['A','B','C','D','E'].slice(0,count).forEach(t=>{const b=document.createElement('button');b.className='teamBtn team-'+t+(state.selectedTeam===t?' active':'');b.textContent=t+'팀';b.style.background=TEAM_COLORS[t];b.style.borderColor='rgba(255,255,255,.54)';b.style.color='#fff';b.style.boxShadow=state.selectedTeam===t?'0 12px 24px rgba(15,31,58,.22), 0 0 0 3px rgba(255,255,255,.82)':'0 8px 18px rgba(15,31,58,.14)';b.onclick=()=>{state.selectedTeam=t;renderTeams(count)};teamWrap.appendChild(b);});}
function renderColors(){colorWrap.innerHTML=''; const colorNames=['연빨강','빨강','진빨강','핑크','살구','주황','호박','노랑','라임','초록','민트','청록','하늘','파랑','진파랑','남색','연보라','보라','진보라','마젠타','갈색','흰색','검정']; CHAR_COLORS.forEach((c,idx)=>{const b=document.createElement('button');b.className='colorBtn'+(state.selectedColor===c?' active':'');b.style.background=c;b.title=colorNames[idx]||c;b.onclick=()=>{state.selectedColor=c;renderColors()};colorWrap.appendChild(b);});}
function renderQuestionEditor(){questionsList.innerHTML='';const required=state.roomInfo?.question_count||0;document.getElementById('requiredCount').textContent=`문제 ${required}개 필수`;if(!state.questions.length){questionsList.innerHTML='<div class="mini" style="padding-top:8px">방 정보를 먼저 확인하세요.</div>';return;}state.questions.forEach((q,idx)=>{const div=document.createElement('div');div.className='qitem';const text=escapeHtml(q.text||'');div.innerHTML=`<div style="display:flex;justify-content:space-between;gap:8px;align-items:center"><strong style="color:#f4fbff">문제 ${idx+1}</strong><span class="badge">필수</span></div><textarea data-field="text" data-idx="${idx}" rows="2" style="width:100%;margin-top:8px" placeholder="문제 내용을 입력하세요">${text}</textarea>${['①','②','③','④'].map((n,i)=>`<div class="questionChoiceRow"><span class="choiceBadge">${n}</span><input data-field="choice" data-cidx="${i}" data-idx="${idx}" value="${escapeHtml(q.choices[i]||'')}" style="flex:1" /><label class="answerRadioLabel"><input type="radio" name="ans_${idx}" data-field="answer" data-idx="${idx}" value="${i}" ${q.answer===i?'checked':''}/><span>정답</span></label></div>`).join('')}`;questionsList.appendChild(div);});document.querySelectorAll('[data-field="text"]').forEach(el=>el.oninput=e=>state.questions[Number(el.dataset.idx)].text=e.target.value);document.querySelectorAll('[data-field="choice"]').forEach(el=>el.oninput=e=>state.questions[Number(el.dataset.idx)].choices[Number(el.dataset.cidx)]=e.target.value);document.querySelectorAll('[data-field="answer"]').forEach(el=>el.onchange=e=>state.questions[Number(el.dataset.idx)].answer=Number(e.target.value));}
function savePlayerSession(){try{if(state.playerId&&state.reconnectToken&&state.roomCode){sessionStorage.setItem('remap_player_session',JSON.stringify({playerId:state.playerId,reconnectToken:state.reconnectToken,roomCode:state.roomCode,nickname:state.nickname,ts:Date.now()}));}}catch(e){}}
function clearPlayerSession(){try{sessionStorage.removeItem('remap_player_session');}catch(e){}}
function loadPlayerSession(){try{const raw=sessionStorage.getItem('remap_player_session');if(!raw)return null;const data=JSON.parse(raw);if(!data||!data.playerId||!data.reconnectToken||!data.roomCode)return null;if(Date.now()-Number(data.ts||0)>1000*60*60)return null;return data;}catch(e){return null;}}
function scheduleReconnect(){if(state.intentionalClose||!state.playerId||!state.reconnectToken||!state.roomCode)return;clearTimeout(state.reconnectTimer);const attempt=Math.min(6,Number(state.reconnectAttempts||0)+1);state.reconnectAttempts=attempt;const delay=Math.min(5000,500*attempt);showToast('연결이 잠시 끊겼습니다. 다시 연결 중...', 'bottom');state.reconnectTimer=setTimeout(()=>connectPlayer(state.roomCode,true),delay);}
function setupJoinedSession(msg, ws){state.pendingId=null;state.nickname=state.nickname||nicknameInput.value.trim();state.playerId=msg.player_id;state.reconnectToken=msg.reconnect_token||state.reconnectToken;state.settings=msg.settings;state.roomTitle=msg.room_title;state.roomCode=msg.room_code;state.reconnectAttempts=0;state.intentionalClose=false;savePlayerSession();roomTitleBar.textContent='REMAP';state.renderPlayers={};state.lastFrameTs=0;resizeCanvas();prepScreen.style.display='none';joinScreen.style.display='none';endScreen.style.display='none';gameScreen.style.display='grid';document.body.classList.add('show-top-leave');setGameActive(true);state.socket=ws;render();startGameLoop();}
function setRoomInfo(info){state.roomInfo=info;state.pendingId=info.pending_id||state.pendingId;state.roomCode=roomCodeInput.value.trim().toUpperCase()||state.roomCode;state.roomTitle=info.title||'';state.nickname=(nicknameInput.value.trim()||state.nickname).trim();if(state.nickname&&!nicknameInput.value.trim()){nicknameInput.value=state.nickname;}roomTitleBar.textContent='REMAP';document.getElementById('infoTitle').textContent=info.title||'-';document.getElementById('infoCode').textContent=state.roomCode||'-';document.getElementById('infoMode').textContent=info.game_mode==='team'?`${info.team_count}팀전`:'개인전';document.getElementById('infoQuestions').textContent=`${info.question_count}문제`;document.getElementById('infoTime').textContent=`${info.question_time_limit}초`;document.getElementById('infoMap').textContent=info.map_label||'-';infoNickname.textContent=state.nickname||nicknameInput.value.trim()||'-';state.selectedTeam='A';ensureQuestionCount(info.question_count);if(info.game_mode==='team'){teamSection.style.display='block';renderTeams(info.team_count);colorWrap.style.display='none';colorHelp.textContent='팀전에서는 캐릭터 색상이 팀별로 고정됩니다.';}else{teamSection.style.display='none';teamWrap.innerHTML='';colorWrap.style.display='grid';colorHelp.textContent='캐릭터 색상 선택';renderColors();}renderQuestionEditor();joinScreen.style.display='none';prepScreen.style.display='flex';document.body.classList.add('show-top-leave');}
async function cancelPending(){if(state.pendingId){try{await fetch('/api/room/cancel_prepare',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pending_id:state.pendingId,code:state.roomCode||roomCodeInput.value.trim().toUpperCase()})});}catch(e){}state.pendingId=null;}}
async function checkRoom(){const code=roomCodeInput.value.trim().toUpperCase();const nickname=(nicknameInput.value.trim()||state.nickname).trim();if(!code){showToast('방 코드를 입력하세요.');return;}if(!nickname){showToast('닉네임을 입력하세요.');return;}state.nickname=nickname;nicknameInput.value=nickname;const res=await fetch('/api/room/prepare',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({code,nickname})});const data=await res.json();if(!data.ok){roomInfo.textContent=data.message||'입장할 수 없습니다.';showToast(data.message||'입장할 수 없습니다.');return;}roomInfo.textContent=`${data.title} · ${data.game_mode==='team'?data.team_count+'팀전':'개인전'} · ${data.question_count}문제`;setRoomInfo(data);} 
checkRoomBtn.onclick=checkRoom;
const BLOCKED_PROFANITY_TERMS=['시발','씨발','ㅅㅂ','ㅆㅂ','병신','븅신','개새끼','새끼','지랄','꺼져','좆','존나','졸라','개같','미친놈','미친년','fuck','shit'];
function normalizeProfanityText(v){return String(v||'').toLowerCase().replace(/[^0-9a-zㄱ-ㅎ가-힣]+/g,'');}
function findBlockedProfanity(v){const s=normalizeProfanityText(v);return BLOCKED_PROFANITY_TERMS.find(t=>s.includes(t))||'';}
function findQuestionProfanity(){for(let i=0;i<state.questions.length;i++){const q=state.questions[i]||{};let term=findBlockedProfanity(q.text);if(term)return {index:i+1,field:'문제',term};for(let j=0;j<(q.choices||[]).length;j++){term=findBlockedProfanity(q.choices[j]);if(term)return {index:i+1,field:`선택지 ${j+1}`,term};}}return null;}
setTimeout(()=>{const saved=loadPlayerSession();if(saved&&joinScreen.style.display!=='none'){state.playerId=saved.playerId;state.reconnectToken=saved.reconnectToken;state.roomCode=saved.roomCode;state.nickname=saved.nickname||'';nicknameInput.value=state.nickname||nicknameInput.value;roomCodeInput.value=state.roomCode;showToast('이전 접속을 복구합니다.','bottom');connectPlayer(state.roomCode,true);}},500);
joinBtn.onclick=()=>{if(!state.roomInfo){showToast('먼저 방 코드를 확인하세요.');return;}const nick=(nicknameInput.value.trim()||state.nickname).trim();if(!nick){showToast('닉네임을 입력하세요.');return;}state.nickname=nick;nicknameInput.value=nick;const required=state.roomInfo.question_count;ensureQuestionCount(required);const validQuestions=state.questions.length===required&&!state.questions.some(q=>!q.text.trim()||q.choices.some(c=>!c.trim()));if(!validQuestions){showToast(`문제 ${required}개를 모두 입력하세요.`);return;}const bad=findQuestionProfanity();if(bad){showToast(`${bad.index}번 ${bad.field}에 사용할 수 없는 표현이 포함되어 있습니다.`);return;}connectPlayer(state.roomCode)};
async function goBackToJoin(){await cancelPending();setGameActive(false);prepScreen.style.display='none';joinScreen.style.display='flex';state.roomInfo=null;state.questions=[];document.body.classList.remove('show-top-leave');roomTitleBar.textContent='REMAP';}
backBtn.onclick=goBackToJoin;
function connectPlayer(code,resume=false){const wsUrl=new URL('/ws/player', location.href);wsUrl.protocol=location.protocol==='https:'?'wss:':'ws:';const ws=new WebSocket(wsUrl.href);ws.onopen=()=>{if(resume&&state.playerId&&state.reconnectToken){ws.send(JSON.stringify({type:'resume',room_code:code,player_id:state.playerId,reconnect_token:state.reconnectToken}));}else{ws.send(JSON.stringify({type:'join',pending_id:state.pendingId,room_code:code,team:state.selectedTeam,color:state.selectedColor,nickname:state.nickname||nicknameInput.value.trim(),questions:state.questions}));}};ws.onmessage=(ev)=>{const msg=JSON.parse(ev.data);if(msg.type==='error'){showToast(msg.message);state.intentionalClose=true;try{ws.close();}catch(e){}}else if(msg.type==='joined'||msg.type==='resumed'){setupJoinedSession(msg,ws);if(msg.type==='resumed')showToast('연결이 복구되었습니다.','bottom','success');}else if(msg.type==='state'){const prevStatus=state.gameStatus;state.players=msg.players;state.rankings=msg.rankings;state.teamRankings=msg.team_rankings;state.settings=msg.settings;state.battles=msg.battles;state.gameStatus=msg.game_status;state.remainingTime=msg.remaining_time;state.logs=msg.student_logs||msg.logs||[];state.mapWalls=msg.map_walls||[];state.roomTitle=msg.room.title;roomTitleBar.textContent='REMAP';statusBar.textContent=`[현재상황: ${msg.game_status==='countdown'?'시작 카운트다운':msg.game_status==='running'?'게임 진행 중':msg.game_status==='finished'?'게임 종료':'준비 중'}]`;if(msg.game_status==='countdown'&&prevStatus!=='countdown'){startCountdownOverlay(msg.countdown_remaining||4)}syncRenderPlayers(msg.players);state.myPlayer=state.players.find(p=>p.id===state.playerId)||null;resizeCanvas();if(state.gameLoopRunning&&gameScreen.style.display==='grid'){renderHud();}else{render();}}else if(msg.type==='battle_intro'){showBattleIntro(msg)}else if(msg.type==='battle_question'){startBattleQuestion(msg)}else if(msg.type==='battle_feedback'){showAnswerFeedback(msg)}else if(msg.type==='battle_result'){showBattleResult(msg)}else if(msg.type==='game_end'){showEndScreen(msg)}else if(msg.type==='reset'){clearPlayerSession();showToast(msg.message||'다음 게임 준비');if(msg.target==='home'){leaveToHome(false)}else{leaveToPrep(true)}}};ws.onerror=()=>{};ws.onclose=()=>{if(state.socket===ws){state.socket=null;stopGameLoop();if(!state.intentionalClose&&gameScreen.style.display==='grid'&&state.gameStatus!=='finished'){scheduleReconnect();}}};}
function leaveToPrep(keepQuestions=false){state.intentionalClose=true;clearTimeout(state.reconnectTimer);clearPlayerSession();stopGameLoop();clearPlayTimers();if(state.socket){try{state.socket.close();}catch(e){}}state.socket=null;state.playerId=null;state.reconnectToken=null;state.myPlayer=null;state.players=[];state.rankings=[];state.teamRankings=[];state.battles=[];state.logs=[];state.keys={};state.currentBattle=null;setGameActive(false);gameScreen.style.display='none';endScreen.style.display='none';joinScreen.style.display='none';prepScreen.style.display='flex';document.body.classList.add('show-top-leave');battleOverlay.style.display='none';resultOverlay.style.display='none';countdownOverlay.style.display='none';statusBar.textContent='[현재상황: 준비 중]';roomTitleBar.textContent='REMAP';state.pendingId=null;state.renderPlayers={};state.lastFrameTs=0;if(keepQuestions){ensureQuestionCount(state.roomInfo?.question_count||state.questions.length||0);}else{state.questions=buildEmptyQuestions(state.roomInfo?.question_count||0);}infoNickname.textContent=state.nickname||nicknameInput.value.trim()||'-';renderQuestionEditor();}
function leaveToHome(send=true){state.intentionalClose=true;clearTimeout(state.reconnectTimer);clearPlayerSession();stopGameLoop();clearPlayTimers();if(send&&state.socket&&state.socket.readyState===1){state.socket.send(JSON.stringify({type:'leave'}));}if(state.socket){try{state.socket.close();}catch(e){}}state.socket=null;state.playerId=null;state.reconnectToken=null;state.myPlayer=null;state.players=[];state.rankings=[];state.teamRankings=[];state.battles=[];state.logs=[];state.keys={};state.currentBattle=null;setGameActive(false);gameScreen.style.display='none';prepScreen.style.display='none';endScreen.style.display='none';joinScreen.style.display='flex';document.body.classList.remove('show-top-leave');battleOverlay.style.display='none';resultOverlay.style.display='none';countdownOverlay.style.display='none';state.pendingId=null;state.nickname='';state.renderPlayers={};state.lastFrameTs=0;statusBar.textContent='[현재상황: 준비 중]';roomTitleBar.textContent='REMAP';}
document.getElementById('endHomeBtn').onclick=()=>leaveToHome(true);leaveBtn.onclick=()=>leaveToHome(true);if(topLeaveBtn)topLeaveBtn.onclick=async()=>{if(prepScreen.style.display==='flex'){await goBackToJoin();}else{leaveToHome(true);}};
function resizeCanvas(){if(!state.settings)return;const w=Number(state.settings.map_width)||1060;const h=Number(state.settings.map_height)||612;if(canvas.width!==w)canvas.width=w;if(canvas.height!==h)canvas.height=h;if(state.settings.background_data_url){if(!state.bgImage||state.bgImage.src!==state.settings.background_data_url){const img=new Image();img.onload=()=>{if(!state.gameLoopRunning)render();};img.src=state.settings.background_data_url;state.bgImage=img;}}else if(state.bgImage){state.bgImage=null;}}
function formatTime(sec){const m=Math.floor(sec/60);const s=sec%60;return `${m}:${String(s).padStart(2,'0')}`}
function syncRenderPlayers(players){const seen=new Set();(players||[]).forEach(p=>{seen.add(p.id);const current=state.renderPlayers[p.id];if(current){current.targetX=p.x;current.targetY=p.y;current.nickname=p.nickname;current.color=p.color;current.team=p.team;current.state=p.state;current.direction=p.direction;current.score=p.score;}else{state.renderPlayers[p.id]={x:p.x,y:p.y,targetX:p.x,targetY:p.y,nickname:p.nickname,color:p.color,team:p.team,state:p.state,direction:p.direction,score:p.score,id:p.id};}});Object.keys(state.renderPlayers).forEach(id=>{if(!seen.has(id)){delete state.renderPlayers[id];}})}
function getRenderablePlayers(){return state.players.map(p=>{const ghost=state.renderPlayers[p.id]||{x:p.x,y:p.y,targetX:p.x,targetY:p.y};const follow=0.50;ghost.x += (ghost.targetX-ghost.x)*follow;ghost.y += (ghost.targetY-ghost.y)*follow;state.renderPlayers[p.id]=Object.assign(ghost,{targetX:p.x,targetY:p.y,nickname:p.nickname,color:p.color,team:p.team,state:p.state,direction:p.direction,score:p.score,id:p.id});return {...p,x:ghost.x,y:ghost.y};});}
function drawScene(){if(state.settings){resizeCanvas();}const hasMaze=(state.mapWalls||[]).length>0;ctx.setTransform(1,0,0,1,0,0);ctx.clearRect(0,0,canvas.width,canvas.height);if(state.bgImage&&state.bgImage.complete){ctx.drawImage(state.bgImage,0,0,canvas.width,canvas.height)}else{const bg=ctx.createLinearGradient(0,0,canvas.width,canvas.height);bg.addColorStop(0,'#eaf2ff');bg.addColorStop(.55,'#dbeafe');bg.addColorStop(1,'#cfe2ff');ctx.fillStyle=bg;ctx.fillRect(0,0,canvas.width,canvas.height);ctx.fillStyle='rgba(255,255,255,.30)';ctx.fillRect(14,14,canvas.width-28,canvas.height-28)}drawGrid(hasMaze);drawWalls(state.mapWalls||[]);getRenderablePlayers().forEach(drawPlayer);}
function renderHud(){const me=state.myPlayer;const teamMode=state.settings&&state.settings.game_mode==='team';const statusText=state.gameStatus==='running'?(me&&me.state==='battling'?'배틀 중':'이동 중'):state.gameStatus==='countdown'?'시작 준비':state.gameStatus==='finished'?'게임 종료':'준비 중';const rank=state.rankings.find(r=>r.player_id===state.playerId)?.rank||'-';topInfo.innerHTML=`<div class="statHero"><span class="mini">${escapeHtml(state.roomTitle||'ReMap')}</span><strong>${escapeHtml(me?me.nickname:'-')}</strong></div>${teamMode?`<div class="rankItem"><span>팀</span><strong>${escapeHtml(me&&me.team?me.team+'팀':'-')}</strong></div>`:''}<div class="rankItem"><span>점수</span><strong>${Number(me?me.score:0)}</strong></div><div class="rankItem"><span>현재 순위</span><strong>${escapeHtml(rank)}위</strong></div><div class="rankItem"><span>남은 시간</span><strong>${formatTime(state.remainingTime)}</strong></div><div class="rankItem"><span>상태</span><strong>${escapeHtml(statusText)}</strong></div>`;rankingList.innerHTML=state.rankings.length?state.rankings.map(r=>`<div class="rankItem"><span>${Number(r.rank)||'-'}. ${escapeHtml(r.nickname)}${teamMode&&r.team?` <span class=\"mini\">(${escapeHtml(r.team)})</span>`:''}</span><strong>${Number(r.score||0)}</strong></div>`).join(''):'<div class="mini">대기 중</div>';teamRankingList.innerHTML=teamMode?(state.teamRankings.length?state.teamRankings.map(r=>`<div class="rankItem"><span>${Number(r.rank)||'-'}. ${escapeHtml(r.team)}팀</span><strong>${Number(r.score||0)}</strong></div>`).join(''):'<div class="mini">팀 정보 없음</div>'):'<div class="mini">개인전 모드</div>';battleList.innerHTML=state.battles.length?state.battles.map(b=>`<div class="battleItem"><span>${(b.players||[]).map(escapeHtml).join(' vs ')}</span><strong>${Number(b.progress||0)}/${Number(b.total||0)}</strong></div>`).join(''):'<div class="mini">진행 중인 배틀 없음</div>';logList.innerHTML=state.logs.length?state.logs.slice().reverse().map(l=>`<div class="logItem"><span>${escapeHtml(l.time)}</span><span>${escapeHtml(l.message)}</span></div>`).join(''):'<div class="mini">아직 배틀 기록이 없습니다.</div>';}
function render(){drawScene();renderHud();}
function drawGrid(hasMaze=false){ctx.strokeStyle=hasMaze?'rgba(37,99,235,0.06)':'rgba(37,99,235,0.08)';for(let x=0;x<canvas.width;x+=50){ctx.beginPath();ctx.moveTo(x,0);ctx.lineTo(x,canvas.height);ctx.stroke()}for(let y=0;y<canvas.height;y+=50){ctx.beginPath();ctx.moveTo(0,y);ctx.lineTo(canvas.width,y);ctx.stroke()}}
function drawWalls(walls){(walls||[]).forEach(w=>{const radius=Math.min(14,Math.min(w.w,w.h)*0.22);const g=ctx.createLinearGradient(w.x,w.y,w.x+w.w,w.y+w.h);g.addColorStop(0,'#7fb6ef');g.addColorStop(.55,'#5d8fda');g.addColorStop(1,'#4d78c6');ctx.save();ctx.shadowColor='rgba(59,130,246,0.18)';ctx.shadowBlur=8;ctx.shadowOffsetY=3;ctx.fillStyle=g;roundRect(w.x,w.y,w.w,w.h,radius,true,false);ctx.restore();ctx.strokeStyle='rgba(255,255,255,.42)';ctx.lineWidth=2;roundRect(w.x,w.y,w.w,w.h,radius,false,true);ctx.fillStyle='rgba(255,255,255,.16)';roundRect(w.x+4,w.y+4,Math.max(8,w.w-8),Math.max(6,Math.min(w.h*0.24,16)),Math.max(4,radius*0.5),true,false);});ctx.lineWidth=1;}
function hexToRgb(hex){const clean=(hex||'#60a5fa').replace('#','');const normalized=clean.length===3?clean.split('').map(c=>c+c).join(''):clean;const n=parseInt(normalized,16);return {r:(n>>16)&255,g:(n>>8)&255,b:n&255};}
function darkenColor(hex, factor=0.22){const {r,g,b}=hexToRgb(hex);return `rgb(${Math.max(0,Math.floor(r*(1-factor)))},${Math.max(0,Math.floor(g*(1-factor)))},${Math.max(0,Math.floor(b*(1-factor)))})`;}
function lightenColor(hex, factor=0.18){const {r,g,b}=hexToRgb(hex);return `rgb(${Math.min(255,Math.floor(r+(255-r)*factor))},${Math.min(255,Math.floor(g+(255-g)*factor))},${Math.min(255,Math.floor(b+(255-b)*factor))})`;}
function alphaColor(hex, alpha){const {r,g,b}=hexToRgb(hex);return `rgba(${r},${g},${b},${alpha})`;}
const playerMascotCache={};
function getPlayerMascot(color){const key=normalizeCharacterColor(color);if(!playerMascotCache[key]){const img=new Image();img.onload=()=>render();img.onerror=()=>{playerMascotCache[key]=null;};img.src=buildCharacterSvgUrl(key);playerMascotCache[key]=img;}return playerMascotCache[key];}
function drawPlayer(p){const size=30;const originalColor=p.color||'#60a5fa';const isMe=!!state.playerId&&p.id===state.playerId;const alreadyBattled=!isMe&&Array.isArray(p.battled_ids)&&p.battled_ids.includes(state.playerId);const bodyColor=alreadyBattled?'#94a3b8':originalColor;const mascot=getPlayerMascot(bodyColor);ctx.save();if(p.connected===false){ctx.globalAlpha=0.32}else if(p.connected===false){ctx.globalAlpha=0.32}else if(p.state==='battling'){ctx.globalAlpha=0.45}ctx.translate(p.x,p.y);if(isMe){ctx.save();ctx.globalAlpha=0.96;ctx.shadowColor='rgba(250,204,21,0.86)';ctx.shadowBlur=24;ctx.fillStyle='rgba(250,204,21,0.28)';ctx.beginPath();ctx.arc(0,0,size/2+13,0,Math.PI*2);ctx.fill();ctx.restore();}ctx.shadowColor='rgba(15,23,42,0.18)';ctx.shadowBlur=4;ctx.shadowOffsetY=1;if(mascot&&mascot.complete&&mascot.naturalWidth>0){ctx.drawImage(mascot,-size/2,-size/2,size,size);}else{ctx.fillStyle=bodyColor;ctx.beginPath();ctx.arc(0,0,size/2.4,0,Math.PI*2);ctx.fill();}ctx.shadowColor='transparent';if(alreadyBattled){ctx.save();ctx.lineWidth=1.8;ctx.strokeStyle='rgba(71,85,105,.72)';ctx.beginPath();ctx.arc(0,0,size/2+3,0,Math.PI*2);ctx.stroke();ctx.restore();}if(p.state==='battling'){ctx.beginPath();ctx.arc(0,0,size/2+4.8,0,Math.PI*2);ctx.strokeStyle='rgba(239,68,68,0.85)';ctx.lineWidth=2;ctx.stroke();}ctx.restore();ctx.fillStyle=isMe?'#92400e':'#173b7a';ctx.font=isMe?'bold 12px Arial':'12px Arial';ctx.textAlign='center';const teamLabel=(state.settings&&state.settings.game_mode==='team'&&p.team)?` [${p.team}]`:'';ctx.fillText(`${p.nickname}${teamLabel}`,p.x,p.y-24)}
function roundRect(x,y,w,h,r,fill,stroke){ctx.beginPath();ctx.moveTo(x+r,y);ctx.arcTo(x+w,y,x+w,y+h,r);ctx.arcTo(x+w,y+h,x,y+h,r);ctx.arcTo(x,y+h,x,y,r);ctx.arcTo(x,y,x+w,y,r);ctx.closePath();if(fill)ctx.fill();if(stroke)ctx.stroke()}
function gameLoop(ts=0){if(!state.gameLoopRunning){state.animationFrameId=null;return;}if(!state.lastFrameTs)state.lastFrameTs=ts;state.lastFrameTs=ts;if(gameScreen.style.display==='grid'){drawScene();}if(state.socket&&state.socket.readyState===1&&gameScreen.style.display==='grid'&&state.gameStatus!=='finished'&&state.gameStatus!=='countdown'){let dx=0,dy=0;if(state.keys['arrowleft']||state.keys['a']||state.keys['KeyA'])dx-=1;if(state.keys['arrowright']||state.keys['d']||state.keys['KeyD'])dx+=1;if(state.keys['arrowup']||state.keys['w']||state.keys['KeyW'])dy-=1;if(state.keys['arrowdown']||state.keys['s']||state.keys['KeyS'])dy+=1;if(state.touchDx||state.touchDy){dx=state.touchDx;dy=state.touchDy;}else{const len=Math.hypot(dx,dy);if(len>1){dx/=len;dy/=len;}}if(dx||dy){const now=performance.now?performance.now():Date.now();const moveInterval=1000/30;if(!state.lastMoveSentAt||now-state.lastMoveSentAt>=moveInterval){state.lastMoveSentAt=now;state.socket.send(JSON.stringify({type:'move',dx:Number(dx.toFixed(3)),dy:Number(dy.toFixed(3))}))}}else{state.lastMoveSentAt=0;}}state.animationFrameId=requestAnimationFrame(gameLoop)}
window.addEventListener('keydown',e=>{const k=(e.key||'').toLowerCase();if(k)state.keys[k]=true;if(e.code)state.keys[e.code]=true;});window.addEventListener('keyup',e=>{const k=(e.key||'').toLowerCase();if(k)state.keys[k]=false;if(e.code)state.keys[e.code]=false;});window.addEventListener('blur',()=>{state.keys={};});
function setBattleProgress(index,total){const pct=total?Math.max(0,Math.min(100,((index-1)/total)*100)):0;if(battleProgressFill)battleProgressFill.style.width=pct+'%'}
function showBattleIntro(msg){battleOverlay.style.display='flex';battleCard.className='intro';battleTitle.textContent='배틀 시작';battleMeta.textContent='잠시 후 문제가 나타납니다';battleTimer.textContent='VS';setBattleProgress(1,1);battleQuestion.className='battleVsStage';battleQuestion.innerHTML=`<div class="battleVsText">${escapeHtml(msg.me)} <span style="opacity:.72">VS</span> ${escapeHtml(msg.opponent)}</div><div class="battleVsSub">상대가 낸 문제를 풀어 승부를 겨뤄요</div>`;battleChoices.innerHTML='';clearInterval(state.timerHandle);state.lastBattleIntroTs=Date.now();}
function startBattleQuestion(msg){const waitAfterIntro=Math.max(0,850-(Date.now()-(state.lastBattleIntroTs||0)));const waitAfterFeedback=Math.max(0,620-(Date.now()-(state.lastFeedbackTs||0)));const delay=Math.max(waitAfterIntro,waitAfterFeedback);if(delay>0){setTimeout(()=>startBattleQuestion(msg),delay);return;}state.currentBattle={battleId:msg.battle_id,index:msg.index,total:msg.total,remaining:msg.time_limit,startTs:Date.now(),answered:false};battleOverlay.style.display='flex';battleCard.className='';battleQuestion.className='battleQuestionBox';battleTitle.textContent=`${msg.opponent} 와 퀴즈 배틀`;battleMeta.textContent=`${msg.index}/${msg.total} 문제`;setBattleProgress(msg.index,msg.total);battleQuestion.textContent=msg.question.text;battleChoices.innerHTML='';msg.question.choices.forEach((c,i)=>{const btn=document.createElement('button');btn.className='choice';btn.dataset.index=String(i);btn.textContent=`${i+1}. ${c}`;btn.onclick=()=>submitAnswer(i);battleChoices.appendChild(btn)});updateBattleTimer();clearInterval(state.timerHandle);state.timerHandle=setInterval(()=>{if(!state.currentBattle)return;state.currentBattle.remaining-=1;updateBattleTimer();if(state.currentBattle.remaining<=0){clearInterval(state.timerHandle)}},1000)}
function updateBattleTimer(){const remain=state.currentBattle?.remaining ?? 0;battleTimer.textContent=`⏱ ${remain}초`;if(state.currentBattle&&state.currentBattle.total){setBattleProgress(state.currentBattle.index,state.currentBattle.total)}}
function submitAnswer(selected){if(!state.currentBattle||!state.socket||state.currentBattle.answered)return;state.currentBattle.answered=true;const timeUsed=Math.min(state.settings.question_time_limit,Math.round((Date.now()-state.currentBattle.startTs)/1000));battleChoices.querySelectorAll('.choice').forEach(btn=>{btn.classList.add('locked');btn.disabled=true;if(Number(btn.dataset.index)===selected)btn.classList.add('selected');});state.socket.send(JSON.stringify({type:'answer',battle_id:state.currentBattle.battleId,selected,time_used:timeUsed}));battleMeta.textContent='채점 중...';}
function showAnswerFeedback(msg){state.lastFeedbackTs=Date.now();const selected=msg.selected;battleChoices.querySelectorAll('.choice').forEach(btn=>{const idx=Number(btn.dataset.index);if(selected!==null&&idx===Number(selected)){btn.classList.remove('selected');btn.classList.add(msg.correct?'correct':'wrong');}});battleEffect.textContent=msg.timed_out?'시간 초과!':msg.correct?'정답!':'오답!';battleEffect.className=msg.correct?'success':'error';if(state.effectTimer)clearTimeout(state.effectTimer);state.effectTimer=setTimeout(()=>{battleEffect.style.display='none';battleEffect.className='';},650);}
function showBattleResult(msg){clearTimeout(state.autoResultHide);resultOverlay.style.display='flex';resultCard.className='resultCard '+(msg.result==='win'?'win':msg.result==='lose'?'lose':'draw');resultTitle.textContent=msg.result==='win'?'승리!':msg.result==='lose'?'패배':'무승부';resultText.innerHTML=`상대: <strong>${escapeHtml(msg.opponent)}</strong><br>내 정답: ${Number(msg.my_correct||0)} / 상대 정답: ${Number(msg.opponent_correct||0)}<br>현재 점수: <strong>${Number(msg.my_score||0)}</strong>`;battleOverlay.style.display='none';battleEffect.style.display='none';battleEffect.className='';state.currentBattle=null;resultConfirmBtn.onclick=()=>{resultOverlay.style.display='none'};state.autoResultHide=setTimeout(()=>{resultOverlay.style.display='none'},2800)}
function escapeHtml(value){return String(value??'').replace(/[&<>'"]/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[ch]))}
function ordinalMedal(rank){return rank===1?'🥇':rank===2?'🥈':rank===3?'🥉':'🏅'}
function normalizeCharacterColor(color){const raw=String(color||'').trim();const cleaned=raw.replace(/[^0-9a-fA-F]/g,'');if(cleaned.length===3){return cleaned.split('').map(ch=>ch+ch).join('').toLowerCase();}if(cleaned.length>=6){return cleaned.slice(0,6).toLowerCase();}return '60a5fa';}
function buildCharacterSvgUrl(color){return `/character/${normalizeCharacterColor(color)}.svg`; }
function applyPodiumMascotTint(winnerColor){
  document.querySelectorAll('#endScreen .podiumCharacterImg[data-color]').forEach(img=>{
    img.src=buildCharacterSvgUrl(img.dataset.color||'#60a5fa');
  });
  // 상단 헤더 마스코트를 1위 학생 색상으로 통일
  const mascotColor=winnerColor||'#60a5fa';
  const leftMascot=document.getElementById('headerMascotLeft');
  const rightMascot=document.getElementById('headerMascotRight');
  if(leftMascot) leftMascot.src=buildCharacterSvgUrl(mascotColor);
  if(rightMascot) rightMascot.src=buildCharacterSvgUrl(mascotColor);
}
function podiumSlot(player,rank){
  if(!player){return `<div class="podiumSpot rank${rank}"><div class="emptyPodium">${rank}위<br>참가자 없음</div></div>`}
  const team=player.team?`<div class="podiumMeta">${escapeHtml(player.team)}팀 · 정답 ${player.correct_count??0}개</div>`:`<div class="podiumMeta">정답 ${player.correct_count??0}개 · 배틀 ${player.battles_played??0}회</div>`;
  const bodyColor=player.color||'#60a5fa';
  const darkColor=darkenColor(bodyColor,0.48);
  const lightColor=lightenColor(bodyColor,0.18);
  return `<div class="podiumSpot rank${rank}">
    ${rank===1?'<div class="winnerGlow"></div><div class="winnerCrown" aria-hidden="true">👑</div><div class="winnerConfetti"><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span><span class="confettiPiece"></span></div>':''}
    <div class="podiumCharacter" style="--pc:${escapeHtml(bodyColor)};--pc-dark:${escapeHtml(darkColor)};--pc-light:${escapeHtml(lightColor)}" aria-label="${escapeHtml(player.nickname)} 캐릭터">
      <img class="podiumCharacterImg" src="${escapeHtml(buildCharacterSvgUrl(bodyColor))}" data-color="${escapeHtml(bodyColor)}" alt="${escapeHtml(player.nickname)} 캐릭터">
    </div>
    <div class="medalBadge">${ordinalMedal(rank)} ${rank}위</div>
    <div class="podiumName">${escapeHtml(player.nickname)}</div>
    ${team}
    <div class="podiumScore">${Number(player.score||0)}점</div>
    <div class="podiumBase">${rank}</div>
  </div>`;
}

function showEndScreen(payload){
  const teamRanks=payload.team_rankings||[];
  const rankings=payload.rankings||[];
  const logs=payload.logs||[];
  const playerStats=payload.player_stats||[];
  const winner=payload.winner_team;
  const mvp=rankings[0]||null;
  const best=payload.best_correct||null;
  const most=payload.most_battles||null;
  setGameActive(false);
  gameScreen.style.display='none';
  battleOverlay.style.display='none';
  resultOverlay.style.display='none';
  countDownSafeHide();
  endScreen.style.display='block';document.body.classList.add('show-top-leave');
  const byRank={};
  rankings.slice(0,3).forEach(r=>{byRank[Number(r.rank)]=r});
  const podiumHtml=rankings.length?`<h3 class="awardSectionTitle">개인 시상대</h3><div class="podiumStage">${podiumSlot(byRank[2],2)}${podiumSlot(byRank[1],1)}${podiumSlot(byRank[3],3)}</div>`:`<h3 class="awardSectionTitle">개인 시상대</h3><div class="emptyPodium">아직 결과가 없습니다.</div>`;
  const fullList=playerStats.map((p,idx)=>`<div class="rankItem"><span>${idx+1}. ${escapeHtml(p.nickname)}${p.team?` <span class="mini">(${escapeHtml(p.team)})</span>`:''}<br><span class="mini">정답 ${p.correct_count}/${p.answer_count} · 배틀 ${p.battles_played}회</span></span><strong>${Number(p.score||0)}점</strong></div>`).join('')||'<div class="mini">참가자 기록이 없습니다.</div>';
  document.getElementById('finalRanks').innerHTML=podiumHtml+`<h3 class="awardSectionTitle" style="margin-top:16px">전체 개인 순위</h3><div class="fullRankList">${fullList}</div>`;
  const winnerColor=(byRank[1]&&byRank[1].color)||'#60a5fa';
  applyPodiumMascotTint(winnerColor);
  document.getElementById('finalTeamRanks').innerHTML=winner?`<div class="teamAwardBanner"><div class="teamAwardIcon">🏆</div><div class="teamAwardText"><span class="mini">팀전 우승</span><strong>${escapeHtml(winner.team)}팀</strong><span>${Number(winner.score||0)}점으로 우승!</span></div></div>`:`<div class="teamAwardBanner"><div class="teamAwardIcon">🎮</div><div class="teamAwardText"><span class="mini">게임 모드</span><strong>개인전 결과</strong><span>개인 순위와 MVP를 확인해요.</span></div></div>`;
  const teamList=teamRanks.length?teamRanks.map(r=>`<div class="rankItem"><span>${r.rank}. ${escapeHtml(r.team)}팀</span><strong>${Number(r.score||0)}점</strong></div>`).join(''):'<div class="mini">팀전 순위 없음</div>';
  const mvpHtml=`<h3 class="awardSectionTitle">MVP</h3><div class="mvpGrid">
    <div class="mvpCard"><div class="mvpIcon">👑</div><div class="mvpText"><b>${mvp?escapeHtml(mvp.nickname):'-'}</b><span>최고 점수 MVP · ${mvp?Number(mvp.score||0):0}점</span></div></div>
    <div class="mvpCard"><div class="mvpIcon">✅</div><div class="mvpText"><b>${best?escapeHtml(best.nickname):'-'}</b><span>가장 많이 맞힌 학생 · ${best?Number(best.correct_count||0):0}개</span></div></div>
    <div class="mvpCard"><div class="mvpIcon">⚔️</div><div class="mvpText"><b>${most?escapeHtml(most.nickname):'-'}</b><span>가장 많은 배틀 · ${most?Number(most.battles_played||0):0}회</span></div></div>
  </div><h3 class="awardSectionTitle" style="margin-top:14px">팀 순위</h3><div>${teamList}</div>`;
  document.getElementById('finalMvp').innerHTML=mvpHtml;
  document.getElementById('finalLogs').innerHTML=logs.length?logs.slice().reverse().map(l=>`<div class="logItem"><span>${escapeHtml(l.time)}</span><span>${escapeHtml(l.message)}</span></div>`).join(''):'<div class="mini">배틀 기록 없음</div>';
  statusBar.textContent='[현재상황: 게임 종료]';
}
function countDownSafeHide(){try{countdownOverlay.style.display='none';}catch(e){}if(state.countdownTimer)clearInterval(state.countdownTimer);state.countdownTimer=null;state.countdownActive=false;}
</script>
</body>
</html>
"""


TEACHER_HTML = """
<!doctype html>
<html lang='ko'>
<head>
<meta charset='utf-8'/>
<meta name='viewport' content='width=device-width, initial-scale=1'/>
<title>ReMap 교사</title>
<style>
:root{
  --bg0:#06111f;
  --bg1:#071b35;
  --bg2:#0b2852;
  --card:rgba(15,31,58,.74);
  --card2:rgba(12,26,48,.86);
  --line:rgba(148,197,255,.22);
  --line2:rgba(255,255,255,.12);
  --text:#eef7ff;
  --muted:#9fb7d5;
  --blue:#38bdf8;
  --blue2:#2563eb;
  --cyan:#22d3ee;
  --violet:#8b5cf6;
  --pink:#ec4899;
  --green:#22c55e;
  --red:#ef4444;
  --amber:#f59e0b;
}
html,body{margin:0;height:100%;font-family:Arial,"Malgun Gothic",sans-serif;background:var(--bg0);color:var(--text);overflow:hidden}*{box-sizing:border-box}
body{display:flex;flex-direction:column;background:
  radial-gradient(circle at 18% 8%,rgba(56,189,248,.22),transparent 32%),
  radial-gradient(circle at 82% 16%,rgba(139,92,246,.20),transparent 30%),
  radial-gradient(circle at 55% 92%,rgba(34,211,238,.13),transparent 36%),
  linear-gradient(135deg,var(--bg0),var(--bg1) 48%,var(--bg2));
}
button,input{font:inherit;border-radius:14px;border:1px solid var(--line);padding:10px 12px;background:rgba(255,255,255,.08);color:var(--text);outline:none}
input{width:100%;box-shadow:inset 0 1px 0 rgba(255,255,255,.06)}
input:focus{border-color:rgba(56,189,248,.72);box-shadow:0 0 0 3px rgba(56,189,248,.14)}
button{cursor:pointer;border:none;color:#fff;font-weight:900;background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed);box-shadow:0 14px 30px rgba(37,99,235,.22);transition:.16s ease transform,.16s ease filter,.16s ease box-shadow}
button:hover{transform:translateY(-1px);filter:brightness(1.07);box-shadow:0 18px 36px rgba(56,189,248,.25)}
button:active{transform:translateY(0)}button:disabled{opacity:.58;cursor:not-allowed;transform:none;filter:none}
button.danger{background:linear-gradient(135deg,#fb7185,#ef4444 58%,#b91c1c);box-shadow:0 14px 30px rgba(239,68,68,.20)}
button.ghost{background:rgba(255,255,255,.08);color:#dff4ff;border:1px solid var(--line);box-shadow:none}
button.soft{background:linear-gradient(135deg,rgba(34,211,238,.28),rgba(139,92,246,.26));color:#eaf8ff;border:1px solid var(--line);box-shadow:none}
.topbar{height:64px;background:rgba(6,17,31,.78);border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:0 18px;gap:14px;backdrop-filter:blur(16px);box-shadow:0 12px 34px rgba(0,0,0,.22);z-index:5;position:relative}
.brand{font-weight:1000;font-size:20px;letter-spacing:.2px;background:linear-gradient(135deg,#eff6ff,#38bdf8 45%,#a78bfa);-webkit-background-clip:text;color:transparent;text-shadow:0 0 24px rgba(56,189,248,.14)}
.roomTitle{position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);width:min(720px,58vw);font-weight:1000;text-align:center;font-size:32px;line-height:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:#ffffff!important;text-shadow:0 3px 12px rgba(255,255,255,.38),0 0 22px rgba(255,255,255,.22);letter-spacing:.8px;pointer-events:none;}
.statusBadge{display:inline-flex;align-items:center;gap:6px;padding:8px 13px;border-radius:999px;background:linear-gradient(135deg,rgba(34,211,238,.20),rgba(59,130,246,.22));border:1px solid rgba(125,211,252,.38);color:#dff9ff;font-size:12px;font-weight:1000;box-shadow:0 0 24px rgba(34,211,238,.12)}
.statusBadge::before{content:"";width:8px;height:8px;border-radius:999px;background:#22c55e;box-shadow:0 0 14px rgba(34,197,94,.85)}
.statusBadge.status-running{background:linear-gradient(135deg,#16a34a,#22c55e)!important;border-color:rgba(187,247,208,.88)!important;color:#ffffff!important;box-shadow:0 0 28px rgba(34,197,94,.36),0 10px 24px rgba(22,163,74,.22)!important}
.statusBadge.status-running::before{background:#dcfce7!important;box-shadow:0 0 16px rgba(220,252,231,.95)!important}
.statusBadge.status-finished{background:linear-gradient(135deg,#ef4444,#b91c1c)!important;border-color:rgba(254,202,202,.88)!important;color:#ffffff!important;box-shadow:0 0 28px rgba(239,68,68,.36),0 10px 24px rgba(185,28,28,.22)!important}
.statusBadge.status-finished::before{background:#fee2e2!important;box-shadow:0 0 16px rgba(254,226,226,.95)!important}
.statusBadge.status-waiting{background:linear-gradient(135deg,#facc15,#f59e0b)!important;border-color:rgba(254,240,138,.92)!important;color:#172033!important;box-shadow:0 0 28px rgba(245,158,11,.34),0 10px 24px rgba(180,83,9,.18)!important}
.statusBadge.status-waiting::before{background:#14532d!important;box-shadow:0 0 12px rgba(20,83,45,.55)!important}
.mini{font-size:12px;color:var(--muted);line-height:1.45}.sectionTitle{margin:0 0 10px;color:#f4fbff;font-size:16px}.panel{background:linear-gradient(180deg,rgba(255,255,255,.095),rgba(255,255,255,.052));border:1px solid var(--line);border-radius:22px;padding:14px;box-shadow:0 18px 44px rgba(0,0,0,.20);backdrop-filter:blur(14px)}
.item{display:flex;justify-content:space-between;gap:12px;align-items:center;padding:8px 0;border-bottom:1px solid rgba(255,255,255,.08);color:#dceeff}.item:last-child{border-bottom:none}.item span:first-child{color:#b7c9e2}.item strong{color:#fff}.row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
#createScreen{height:calc(100vh - 64px);display:flex;align-items:flex-start;justify-content:center;padding:24px;overflow:auto}.createCard{width:min(980px,96vw);background:linear-gradient(180deg,rgba(15,31,58,.86),rgba(10,24,45,.88));border:1px solid var(--line);border-radius:28px;padding:24px;box-shadow:0 24px 70px rgba(0,0,0,.30),inset 0 1px 0 rgba(255,255,255,.08);position:relative;overflow:hidden}.createCard::before{content:"";position:absolute;inset:-1px;background:radial-gradient(circle at 12% 0,rgba(56,189,248,.18),transparent 28%),radial-gradient(circle at 88% 18%,rgba(236,72,153,.12),transparent 25%);pointer-events:none}.createCard>*{position:relative}.createCard h1{margin:0 0 8px 0;font-size:30px;color:#fff}.createGrid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:16px}.field{margin-top:0}.field label{display:block;font-size:12px;color:#b8cce8;margin-bottom:6px;font-weight:800}.full{grid-column:1/-1}
.modeButtons{display:grid;grid-template-columns:1fr 1fr;gap:9px}.modeBtn,.mapBtn{background:rgba(255,255,255,.07);color:#dff4ff;border:1px solid var(--line);box-shadow:none;text-align:center}.modeBtn.active,.mapBtn.active{background:linear-gradient(135deg,rgba(34,211,238,.95),rgba(37,99,235,.96) 58%,rgba(124,58,237,.96));border-color:rgba(125,211,252,.65);color:#fff;box-shadow:0 16px 34px rgba(37,99,235,.24)}
.mapButtons{display:grid;grid-template-columns:1fr 1fr;gap:10px}.mapBtn{padding:14px 12px}.mapBtn small{display:block;margin-top:4px;font-weight:700;color:#aecaec}.mapBtn.active small{color:#eaf8ff}.helpBox{background:rgba(255,255,255,.06);border:1px solid var(--line);border-radius:18px;padding:13px;color:#deefff}.teamLegend{display:flex;flex-wrap:wrap;gap:8px;margin-top:9px}.teamChip{padding:6px 10px;border-radius:999px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12);font-size:12px;font-weight:900}.createActions{display:flex;gap:10px;margin-top:18px}.createActions button{flex:1;height:48px}.teacherCredit{margin-top:16px;text-align:right;color:#35516f;font-size:13px;font-weight:900;line-height:1.55;letter-spacing:-.2px}.teacherCredit div{white-space:nowrap}.teacherCredit b{color:#0f2544}
#operateScreen{height:calc(100vh - 64px);display:none;grid-template-rows:auto 1fr;gap:12px;padding:12px;overflow:hidden}.opHeader{display:grid;grid-template-columns:520px minmax(0,1fr);gap:12px;align-items:stretch}.codeBox{display:flex;align-items:center;justify-content:space-between;gap:18px;min-height:172px}.codeTextBlock{min-width:0;flex:1}.codeValue{font-size:54px;letter-spacing:7px;font-weight:1000;color:#fff;line-height:1;text-shadow:0 0 22px rgba(56,189,248,.34)}.joinUrl{margin-top:10px;font-size:12px;line-height:1.35;color:#173b7a;font-weight:900;word-break:break-all;background:rgba(226,238,255,.92);border:1px solid rgba(91,141,204,.30);border-radius:12px;padding:8px 10px}.qrPanel{display:flex;flex-direction:column;align-items:center;gap:7px;flex:0 0 auto}.joinQr{display:block;width:132px;height:132px;padding:8px;border-radius:18px;background:#fff;border:1px solid rgba(15,31,58,.16);box-shadow:0 14px 30px rgba(15,31,58,.18);object-fit:contain;cursor:pointer;transition:.16s ease transform,.16s ease box-shadow}.joinQr:hover{transform:translateY(-2px) scale(1.02);box-shadow:0 18px 36px rgba(15,31,58,.24)}.qrCaption{font-size:11px;font-weight:900;color:#173b7a;text-align:center}.qrModal{position:fixed;inset:0;display:none;align-items:center;justify-content:center;padding:24px;background:rgba(3,7,18,.72);backdrop-filter:blur(10px);z-index:1000}.qrModal.show{display:flex}.qrModalCard{width:min(560px,92vw);background:linear-gradient(180deg,#ffffff,#eaf4ff);border:1px solid rgba(148,197,255,.55);border-radius:30px;padding:22px;box-shadow:0 34px 100px rgba(0,0,0,.42);position:relative;text-align:center}.qrModalClose{position:absolute;right:14px;top:14px;width:42px;height:42px;border-radius:999px;border:1px solid rgba(15,31,58,.12);background:#ffffff;color:#10233f;font-size:28px;line-height:1;font-weight:900;box-shadow:0 10px 24px rgba(15,31,58,.12);cursor:pointer}.qrModalClose:hover{background:#eff6ff}.qrModalTitle{font-size:24px;font-weight:1000;color:#10233f;margin:4px 46px 6px}.qrModalCode{font-size:46px;letter-spacing:8px;font-weight:1000;color:#0f2d59;margin-bottom:10px}.qrModalImg{display:block;width:min(390px,72vw);height:min(390px,72vw);object-fit:contain;background:#fff;border-radius:24px;border:1px solid rgba(15,31,58,.12);padding:16px;margin:0 auto;box-shadow:0 18px 46px rgba(15,31,58,.16)}.qrModalUrl{margin:13px auto 0;max-width:460px;font-size:13px;line-height:1.4;color:#173b7a;font-weight:900;word-break:break-all;background:#f3f8ff;border:1px solid rgba(91,141,204,.24);border-radius:14px;padding:10px}.qrModalHint{margin-top:10px;color:#4b6385;font-size:13px;font-weight:800}.codeTag{padding:7px 10px;border-radius:999px;background:rgba(34,211,238,.15);border:1px solid rgba(125,211,252,.28);font-size:12px;color:#bdefff;font-weight:900}.summaryPanel{padding:10px!important;align-self:stretch;display:flex;flex-direction:column;gap:10px;min-height:172px}.headerMeta{display:grid;grid-template-columns:1.05fr .8fr .9fr 1.4fr .85fr;gap:8px;height:auto}.miniCard{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.11);border-radius:14px;padding:8px 10px;min-width:0;min-height:60px;display:flex;flex-direction:column;justify-content:center}.miniCard strong{display:block;font-size:17px;color:#fff;margin-top:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.miniCard.primary strong{font-size:20px}.miniCard.compact strong{font-size:16px}.opButtons{display:grid;grid-template-columns:1fr 1fr 1fr 1.15fr 1fr;gap:8px;margin-top:auto}.opButtons.hasCeremony{grid-template-columns:1fr 1fr 1fr 1fr 1.15fr 1fr}.ceremonyBtn{display:none;background:linear-gradient(135deg,#fbbf24,#f59e0b)!important;color:#172033!important;box-shadow:0 12px 24px rgba(245,158,11,.22)!important}.ceremonyBtn.show{display:block!important}.opButtons button{min-height:44px;padding:8px 10px;font-size:15px;line-height:1.2}.opButtons #endBtn{box-shadow:0 12px 24px rgba(220,38,38,.16)}
.opMain{display:grid;grid-template-columns:310px minmax(0,1fr) 330px;gap:12px;min-height:0}.col{min-height:0;overflow:auto;padding-right:2px}.col .panel{margin-bottom:10px}.centerCol{min-width:0;min-height:0;display:flex}.mapPanel{width:100%;min-width:0;min-height:0;display:flex;flex-direction:column;padding:15px}.mapHeader{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;gap:10px}.mapTitleWrap h3{margin:0;color:#fff}.mapTitleWrap .mini{margin-top:3px}.mapWrap{flex:1;min-height:0;display:flex;justify-content:center;align-items:center;background:linear-gradient(180deg,rgba(10,28,54,.74),rgba(8,20,38,.80));border:1px solid rgba(255,255,255,.09);border-radius:20px;padding:10px;position:relative;overflow:hidden}.mapWrap::before{content:"";position:absolute;inset:0;background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.10),transparent 45%);pointer-events:none}canvas{position:relative;max-width:100%;max-height:100%;border:2px solid rgba(125,211,252,.38);border-radius:20px;background:#d5e4f7;box-shadow:0 26px 60px rgba(0,0,0,.32),0 0 0 8px rgba(255,255,255,.035)}#teacherCanvas{transform:scale(var(--teacher-map-scale,1));transform-origin:center center;transition:transform .12s ease}.teacherMapZoomControls{position:absolute;right:12px;bottom:12px;z-index:18;display:none;align-items:center;gap:6px;padding:6px;border-radius:999px;background:rgba(238,247,255,.92);border:1px solid rgba(125,211,252,.45);box-shadow:0 10px 24px rgba(15,31,58,.20);backdrop-filter:blur(10px)}.teacherMapZoomControls button{width:34px;height:30px;min-height:30px;padding:0;border-radius:999px;background:linear-gradient(135deg,#eef6ff,#dbeafe);color:#12315a;font-weight:1000;box-shadow:none}.teacherMapZoomControls span{min-width:44px;text-align:center;color:#12315a;font-weight:1000;font-size:12px}
.pillState{display:inline-flex;align-items:center;padding:5px 9px;border-radius:999px;background:rgba(34,197,94,.14);border:1px solid rgba(34,197,94,.28);color:#bbf7d0;font-size:12px;font-weight:900}.submitDone{color:#86efac!important}.submitWait{color:#fca5a5!important}
@media(max-width:1180px){html,body{overflow:auto}.opHeader,.opMain{grid-template-columns:1fr}#operateScreen{overflow:auto}.centerCol{min-height:560px}.summaryPanel{min-height:auto}.headerMeta{grid-template-columns:repeat(2,1fr)}.createGrid{grid-template-columns:1fr}.row{grid-template-columns:1fr}.opButtons{grid-template-columns:1fr 1fr}.createActions{flex-direction:column}.codeBox{align-items:center}.qrPanel{align-self:center}.joinQr{width:138px;height:138px}}@media(max-width:640px){.codeBox{flex-direction:column;text-align:center}.codeValue{font-size:48px}.headerMeta{grid-template-columns:1fr 1fr}.miniCard{min-height:54px}.opButtons{grid-template-columns:1fr}}


/* ===== ReMap contrast patch: bright content cards on dark gradient background ===== */
.card,
.createCard{
  background:linear-gradient(180deg,rgba(255,255,255,.97),rgba(235,244,255,.95))!important;
  border:1px solid rgba(148,197,255,.48)!important;
  color:#10233f!important;
  box-shadow:0 26px 70px rgba(0,0,0,.26), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
.card::before,
.createCard::before{
  background:radial-gradient(circle at 10% 0,rgba(56,189,248,.16),transparent 30%),radial-gradient(circle at 90% 12%,rgba(139,92,246,.12),transparent 30%)!important;
}
.card h1,
.createCard h1{
  color:#0f2a4d!important;
}
.card .mini,
.createCard .mini,
.panel .mini,
.miniCard .mini,
.helpBox .mini{
  color:#5b718f!important;
}
.panel,
.infoCard,
.miniCard,
.helpBox,
#questionsList .qitem{
  background:linear-gradient(180deg,rgba(255,255,255,.88),rgba(225,237,252,.82))!important;
  border:1px solid rgba(91,141,204,.28)!important;
  color:#10233f!important;
  box-shadow:0 14px 34px rgba(15,31,58,.10)!important;
}
.panel h3,
.sectionTitle,
.mapTitleWrap h3{
  color:#12315a!important;
}
.label,
.field label,
.item span:first-child,
.rankItem span:first-child,
.battleItem span:first-child,
.logItem span:first-child{
  color:#526a8a!important;
}
.value,
.item strong,
.rankItem strong,
.battleItem strong,
.statHero strong,
.miniCard strong,
.codeValue{
  color:#0f2544!important;
  text-shadow:none!important;
}
.item,
.rankItem,
.battleItem,
.logItem{
  color:#173455!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
}
input,
textarea,
select{
  background:rgba(255,255,255,.82)!important;
  border:1px solid rgba(91,141,204,.35)!important;
  color:#10233f!important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.92)!important;
}
input::placeholder,
textarea::placeholder{
  color:rgba(82,106,138,.68)!important;
}
input:focus,
textarea:focus,
select:focus{
  border-color:rgba(37,99,235,.62)!important;
  box-shadow:0 0 0 3px rgba(37,99,235,.14), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
button.ghost,
button.soft,
.modeBtn,
.mapBtn,
.teamBtn{
  background:linear-gradient(180deg,rgba(255,255,255,.92),rgba(226,238,255,.86))!important;
  color:#12315a!important;
  border:1px solid rgba(91,141,204,.34)!important;
}
button.ghost:hover,
button.soft:hover,
.modeBtn:hover,
.mapBtn:hover,
.teamBtn:hover{
  filter:brightness(1.02)!important;
}
.modeBtn.active,
.mapBtn.active,
.teamBtn.active{
  background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed)!important;
  color:#fff!important;
  border-color:rgba(37,99,235,.50)!important;
}
.badge,
.codeTag,
.teamChip{
  background:rgba(14,116,144,.10)!important;
  border:1px solid rgba(14,116,144,.22)!important;
  color:#0f5f78!important;
}
.statHero{
  background:linear-gradient(135deg,rgba(224,242,254,.92),rgba(226,232,255,.90))!important;
  border:1px solid rgba(91,141,204,.25)!important;
}
#mapWrap,
.mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.88),rgba(213,228,247,.82))!important;
  border:1px solid rgba(91,141,204,.25)!important;
}
#mapWrap::before,
.mapWrap::before{
  background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.12),transparent 45%)!important;
}
.pillState,
.teamPill{
  background:rgba(34,197,94,.12)!important;
  border:1px solid rgba(34,197,94,.26)!important;
  color:#15803d!important;
}
.submitDone{color:#15803d!important}.submitWait{color:#b91c1c!important}
.colorBtn{border-color:rgba(15,31,58,.18)!important}


/* ===== ReMap hierarchy contrast refinement =====
   Dark background -> bright main card -> darker section panels -> bright inner cells.
   This keeps the big container easy to see while restoring clear boundaries inside it. */
.card,
.createCard{
  background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(231,241,255,.96))!important;
  border:1px solid rgba(191,219,254,.72)!important;
  color:#0f2544!important;
}
.panel,
.helpBox,
.miniCard,
.studentMapPanel,
.mapPanel{
  background:linear-gradient(145deg,rgba(34,58,92,.94),rgba(19,42,72,.92))!important;
  border:1px solid rgba(96,165,250,.38)!important;
  color:#eef7ff!important;
  box-shadow:0 18px 44px rgba(15,31,58,.20)!important;
}
.panel h3,
.sectionTitle,
.mapTitleWrap h3{
  color:#f8fbff!important;
}
.panel .mini,
.helpBox .mini,
.miniCard .mini,
.mapTitleWrap .mini{
  color:#c7d8ee!important;
}
.infoCard,
#questionsList .qitem,
.item,
.rankItem,
.battleItem,
.logItem,
.statHero{
  background:linear-gradient(180deg,rgba(255,255,255,.95),rgba(230,241,255,.90))!important;
  border:1px solid rgba(147,197,253,.46)!important;
  color:#10233f!important;
  box-shadow:0 10px 26px rgba(15,31,58,.10)!important;
}
.infoGrid .infoCard{
  background:linear-gradient(180deg,rgba(255,255,255,.97),rgba(221,236,255,.92))!important;
}
.label,
.field label,
.item span:first-child,
.rankItem span:first-child,
.battleItem span:first-child,
.logItem span:first-child{
  color:#4d6688!important;
}
.value,
.item strong,
.rankItem strong,
.battleItem strong,
.statHero strong,
.miniCard strong,
.codeValue{
  color:#0f2544!important;
}
input,
textarea,
select{
  background:rgba(255,255,255,.96)!important;
  border:1px solid rgba(91,141,204,.42)!important;
  color:#10233f!important;
}
#questionsList .qitem textarea,
#questionsList .qitem input{
  background:rgba(255,255,255,.98)!important;
}
.teamBtn{
  color:#fff!important;
  border:1px solid rgba(255,255,255,.45)!important;
  text-shadow:0 1px 2px rgba(0,0,0,.28)!important;
  box-shadow:0 8px 18px rgba(15,31,58,.14)!important;
}
.teamBtn.active{
  color:#fff!important;
  border-color:rgba(255,255,255,.88)!important;
  box-shadow:0 14px 28px rgba(15,31,58,.22),0 0 0 3px rgba(255,255,255,.78)!important;
}
.team-A{background:#3b82f6!important}.team-B{background:#ef4444!important}.team-C{background:#facc15!important;color:#fff!important;text-shadow:0 1px 2px rgba(0,0,0,.38)!important}.team-D{background:#22c55e!important}.team-E{background:#64748b!important}
.modeBtn,
.mapBtn,
button.ghost,
button.soft{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(226,238,255,.88))!important;
  color:#12315a!important;
}
.modeBtn.active,
.mapBtn.active{
  background:linear-gradient(135deg,#38bdf8,#2563eb 58%,#7c3aed)!important;
  color:#fff!important;
}
.badge,
.codeTag,
.teamChip{
  background:rgba(219,246,255,.92)!important;
  border:1px solid rgba(14,116,144,.22)!important;
  color:#0f5f78!important;
}



/* ===== ReMap contrast hierarchy fix v2 =====
   1) Keep operation dashboard clean/light after room creation.
   2) Fixed team color chips use real team colors.
   3) Student team selection keeps its original team color; selection uses border/glow only. */
#operateScreen .panel,
#operateScreen .miniCard,
#operateScreen .mapPanel{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(222,235,252,.90))!important;
  border:1px solid rgba(147,197,253,.42)!important;
  color:#10233f!important;
  box-shadow:0 16px 38px rgba(15,31,58,.16)!important;
}
#operateScreen .sectionTitle,
#operateScreen .mapTitleWrap h3{
  color:#12315a!important;
}
#operateScreen .mini,
#operateScreen .miniCard .mini,
#operateScreen .mapTitleWrap .mini{
  color:#607692!important;
}
#operateScreen .miniCard strong,
#operateScreen .item strong,
#operateScreen .codeValue{
  color:#0f2544!important;
  text-shadow:none!important;
}
#operateScreen .item{
  background:transparent!important;
  color:#173455!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
  box-shadow:none!important;
}
#operateScreen .item span:first-child{
  color:#526a8a!important;
}
#operateScreen .mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.96),rgba(213,228,247,.90))!important;
  border:1px solid rgba(91,141,204,.28)!important;
}
.teamFixedBox{
  background:linear-gradient(145deg,rgba(34,58,92,.96),rgba(19,42,72,.94))!important;
  color:#eef7ff!important;
}
.teamFixedBox b{color:#fff!important;}
.teamLegend .teamChip,
.teamWrap .teamBtn{
  color:#fff!important;
  border:1px solid rgba(255,255,255,.54)!important;
  text-shadow:0 1px 2px rgba(0,0,0,.34)!important;
  box-shadow:0 8px 18px rgba(15,31,58,.14)!important;
}
.teamLegend .teamChip.team-A,
.teamWrap .teamBtn.team-A{background:#3b82f6!important;}
.teamLegend .teamChip.team-B,
.teamWrap .teamBtn.team-B{background:#ef4444!important;}
.teamLegend .teamChip.team-C,
.teamWrap .teamBtn.team-C{background:#facc15!important;color:#fff!important;text-shadow:0 1px 2px rgba(0,0,0,.45)!important;}
.teamLegend .teamChip.team-D,
.teamWrap .teamBtn.team-D{background:#22c55e!important;}
.teamLegend .teamChip.team-E,
.teamWrap .teamBtn.team-E{background:#64748b!important;}
.teamWrap .teamBtn.active{
  color:#fff!important;
  border-color:rgba(255,255,255,.96)!important;
  outline:3px solid rgba(255,255,255,.82)!important;
  outline-offset:2px!important;
  box-shadow:0 12px 24px rgba(15,31,58,.22)!important;
  transform:none!important;
}
.teamWrap .teamBtn:hover{
  filter:brightness(1.05)!important;
}


/* ===== ReMap student clean light dashboard patch =====
   Match the student running screen to the teacher operation screen:
   dark navy background + clean bright cards + centered bright map. */
#gameScreen .panel,
#gameScreen .studentMapPanel{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(222,235,252,.90))!important;
  border:1px solid rgba(147,197,253,.42)!important;
  color:#10233f!important;
  box-shadow:0 16px 38px rgba(15,31,58,.16)!important;
  backdrop-filter:none!important;
}
#gameScreen .panel h3,
#gameScreen .mapTitleWrap h3{
  color:#12315a!important;
}
#gameScreen .mini,
#gameScreen .mapTitleWrap .mini{
  color:#607692!important;
}
#gameScreen #mapWrap{
  background:linear-gradient(180deg,rgba(235,245,255,.96),rgba(213,228,247,.90))!important;
  border:1px solid rgba(91,141,204,.28)!important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.80)!important;
}
#gameScreen #mapWrap::before{
  background:radial-gradient(circle at 50% 50%,rgba(56,189,248,.10),transparent 45%)!important;
}
#gameScreen canvas{
  border:2px solid rgba(125,211,252,.48)!important;
  background:#d5e4f7!important;
  box-shadow:0 26px 60px rgba(15,31,58,.20),0 0 0 8px rgba(255,255,255,.30)!important;
}
#gameScreen .rankItem,
#gameScreen .battleItem,
#gameScreen .logItem,
#gameScreen .statHero{
  background:transparent!important;
  border-bottom:1px solid rgba(91,141,204,.18)!important;
  color:#173455!important;
  box-shadow:none!important;
}
#gameScreen .rankItem span:first-child,
#gameScreen .battleItem span:first-child,
#gameScreen .logItem span:first-child,
#gameScreen .label{
  color:#526a8a!important;
}
#gameScreen .rankItem strong,
#gameScreen .battleItem strong,
#gameScreen .statHero strong,
#gameScreen .value{
  color:#0f2544!important;
  text-shadow:none!important;
}
#gameScreen .teamPill{
  background:rgba(34,197,94,.12)!important;
  border:1px solid rgba(34,197,94,.26)!important;
  color:#15803d!important;
}
#gameScreen button.ghost{
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(226,238,255,.88))!important;
  color:#12315a!important;
  border:1px solid rgba(91,141,204,.34)!important;
  box-shadow:none!important;
}


/* ===== v3.23 teacher mobile fit-to-screen patch =====
   - No horizontal scrolling on teacher smartphone screens.
   - Header uses real columns instead of an absolute centered title, so the
     status badge never overlaps REMAP.
   - Operation buttons wrap inside the available width instead of becoming a
     horizontally scrollable strip.
*/
@media (max-width: 820px){
  html,body{overflow-x:hidden!important;max-width:100vw!important;}
  .topbar{
    height:58px!important;
    display:grid!important;
    grid-template-columns:auto minmax(0,1fr) auto!important;
    align-items:center!important;
    gap:8px!important;
    padding:0 10px!important;
    overflow:hidden!important;
  }
  .topbar .brand{
    position:relative!important;
    left:auto!important;
    flex:none!important;
    min-width:0!important;
    font-size:clamp(16px,4.2vw,20px)!important;
    white-space:nowrap!important;
    z-index:2!important;
  }
  .topbar .roomTitle{
    position:static!important;
    left:auto!important;
    top:auto!important;
    transform:none!important;
    width:auto!important;
    max-width:none!important;
    min-width:0!important;
    text-align:center!important;
    font-size:clamp(25px,7.2vw,34px)!important;
    letter-spacing:.2px!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
    white-space:nowrap!important;
    z-index:1!important;
  }
  .topbar .statusBadge{
    position:relative!important;
    flex:none!important;
    min-width:0!important;
    max-width:clamp(124px,37vw,178px)!important;
    justify-content:center!important;
    white-space:nowrap!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
    font-size:clamp(10px,2.65vw,12px)!important;
    padding:7px 9px!important;
    gap:4px!important;
    z-index:3!important;
  }
  .topbar .statusBadge::before{flex:0 0 7px!important;width:7px!important;height:7px!important;}

  #operateScreen{overflow-x:hidden!important;max-width:100vw!important;}
  #operateScreen .panel,
  #operateScreen .codeBox,
  #operateScreen .summaryPanel,
  #operateScreen .centerCol,
  #operateScreen .sideCol{max-width:100%!important;min-width:0!important;}
  .opHeader,.opMain{max-width:100%!important;min-width:0!important;}
  .summaryPanel .opButtons,
  .summaryPanel .opButtons.hasCeremony{
    display:grid!important;
    overflow:visible!important;
    gap:8px!important;
    padding:0!important;
    margin-top:10px!important;
    grid-template-columns:repeat(3,minmax(0,1fr))!important;
  }
  .summaryPanel .opButtons button{
    width:100%!important;
    min-width:0!important;
    max-width:100%!important;
    min-height:58px!important;
    padding:8px 6px!important;
    font-size:clamp(12px,3.15vw,15px)!important;
    line-height:1.16!important;
    white-space:normal!important;
    word-break:keep-all!important;
    overflow-wrap:anywhere!important;
  }
  .summaryPanel .opButtons #ceremonyBtn{order:3!important;}
  .summaryPanel .opButtons #resetBtn{order:4!important;}
  .summaryPanel .opButtons #exportQuestionsBtn{order:5!important;}
  .summaryPanel .opButtons #aiReviewBtn{order:6!important;}
  .summaryPanel .opButtons #newRoomBtn{order:7!important;}
}
@media (max-width: 380px){
  .topbar{grid-template-columns:auto minmax(58px,1fr) auto!important;gap:6px!important;padding:0 8px!important;}
  .topbar .brand{font-size:15px!important;}
  .topbar .roomTitle{font-size:24px!important;}
  .topbar .statusBadge{max-width:128px!important;font-size:9.5px!important;padding:6px 7px!important;}
  .summaryPanel .opButtons button{font-size:11.5px!important;min-height:56px!important;padding:7px 4px!important;}
}
@media (max-width: 820px) and (orientation: landscape){
  .topbar{height:54px!important;grid-template-columns:auto minmax(0,1fr) auto!important;}
  .topbar .brand{font-size:17px!important;}
  .topbar .roomTitle{font-size:clamp(27px,5.2vw,38px)!important;}
  .topbar .statusBadge{max-width:220px!important;font-size:12px!important;padding:7px 11px!important;}
  .summaryPanel .opButtons,
  .summaryPanel .opButtons.hasCeremony{grid-template-columns:repeat(6,minmax(0,1fr))!important;}
  .summaryPanel .opButtons button{min-height:54px!important;font-size:clamp(12px,1.85vw,15px)!important;padding:8px 5px!important;}
}
@media (min-width:821px) and (max-height:720px) and (orientation:landscape){
  .summaryPanel .opButtons,
  .summaryPanel .opButtons.hasCeremony{display:grid!important;grid-template-columns:repeat(6,minmax(0,1fr))!important;overflow:visible!important;gap:8px!important;}
  .summaryPanel .opButtons button{min-width:0!important;width:100%!important;white-space:normal!important;word-break:keep-all!important;overflow-wrap:break-word!important;}
}



/* ===== v3.24 teacher mobile usability patch =====
   1) Teacher mobile landscape map has +/- controls and starts slightly smaller.
   2) Teacher mobile portrait room setup keeps compact settings in 2 columns.
*/
@media (max-width: 820px) and (orientation: landscape) and (pointer: coarse){
  .teacherMapZoomControls{display:inline-flex!important;}
  #teacherCanvas{max-width:100%!important;max-height:100%!important;}
}
@media (max-width: 820px) and (orientation: portrait){
  #createScreen .createGrid > .row.full{
    grid-template-columns:repeat(2,minmax(0,1fr))!important;
    gap:10px!important;
  }
  #createScreen .createGrid > .row.full .field label{
    font-size:11px!important;
    line-height:1.25!important;
    min-height:28px!important;
    display:flex!important;
    align-items:flex-end!important;
  }
  #createScreen .createGrid > .row.full .field input{
    min-height:42px!important;
    padding:9px 10px!important;
    font-size:14px!important;
  }
}


/* ===== v3.25 teacher mobile landscape QR + realtime map fit patch =====
   - QR enlargement popup no longer fills the entire landscape phone screen;
     the close button stays visible.
   - Teacher realtime map on landscape phones is sized by the visible vertical
     viewport (svh), so Chrome/Samsung/iPhone browser address bars are considered.
   - Default teacher mobile landscape map scale starts at 75%, with +/- controls
     still available for adjustment.
*/
@media (max-width: 820px) and (orientation: landscape) and (pointer: coarse){
  .qrModal{
    align-items:center!important;
    justify-content:center!important;
    padding:8px!important;
    overflow:auto!important;
  }
  .qrModalCard{
    width:min(440px,86vw)!important;
    max-width:86vw!important;
    max-height:calc(100svh - 16px)!important;
    padding:10px 12px 12px!important;
    border-radius:18px!important;
    overflow:auto!important;
  }
  .qrModalClose{
    right:8px!important;
    top:8px!important;
    width:34px!important;
    height:34px!important;
    font-size:22px!important;
    z-index:3!important;
  }
  .qrModalTitle{
    font-size:15px!important;
    margin:2px 38px 2px!important;
    line-height:1.2!important;
  }
  .qrModalCode{
    font-size:clamp(24px,5.2vw,34px)!important;
    letter-spacing:5px!important;
    margin-bottom:4px!important;
    line-height:1!important;
  }
  .qrModalImg{
    width:min(52vw, calc(100svh - 138px), 300px)!important;
    height:min(52vw, calc(100svh - 138px), 300px)!important;
    min-width:132px!important;
    min-height:132px!important;
    padding:8px!important;
    border-radius:16px!important;
  }
  .qrModalUrl{
    max-width:100%!important;
    margin-top:6px!important;
    font-size:10.5px!important;
    line-height:1.25!important;
    padding:6px 8px!important;
    border-radius:10px!important;
  }
  .qrModalHint{
    margin-top:5px!important;
    font-size:10.5px!important;
    line-height:1.2!important;
  }

  #operateScreen .centerCol{
    min-height:0!important;
    height:auto!important;
  }
  #operateScreen .mapPanel{
    min-height:0!important;
    height:auto!important;
    padding:8px!important;
  }
  #operateScreen .mapHeader{
    margin-bottom:6px!important;
    gap:6px!important;
  }
  #operateScreen .mapTitleWrap h3{
    font-size:16px!important;
    line-height:1.1!important;
  }
  #operateScreen .mapTitleWrap .mini{
    font-size:10.5px!important;
    line-height:1.2!important;
  }
  #operateScreen .mapHeader .ghost{
    min-height:34px!important;
    padding:6px 10px!important;
    font-size:12px!important;
    border-radius:12px!important;
  }
  #operateScreen .mapWrap{
    height:calc(100svh - 118px)!important;
    max-height:calc(100svh - 118px)!important;
    min-height:176px!important;
    padding:6px!important;
    border-radius:18px!important;
  }
  #teacherCanvas{
    max-width:100%!important;
    max-height:calc(100svh - 138px)!important;
  }
  .teacherMapZoomControls{
    right:8px!important;
    bottom:8px!important;
    padding:5px!important;
    gap:5px!important;
  }
  .teacherMapZoomControls button{
    width:32px!important;
    height:28px!important;
    min-height:28px!important;
  }
  .teacherMapZoomControls span{
    min-width:40px!important;
    font-size:11px!important;
  }
}
@media (max-width: 700px) and (orientation: landscape) and (pointer: coarse){
  #operateScreen .mapWrap{
    height:calc(100svh - 106px)!important;
    max-height:calc(100svh - 106px)!important;
    min-height:160px!important;
  }
  #teacherCanvas{max-height:calc(100svh - 126px)!important;}
  .qrModalImg{
    width:min(50vw, calc(100svh - 128px), 260px)!important;
    height:min(50vw, calc(100svh - 128px), 260px)!important;
  }
}

/* ===== v3.26 teacher page credit responsive patch ===== */
@media(max-width:1180px){
  #createScreen .teacherCredit{
    text-align:center!important;
    margin-top:12px!important;
    padding:10px 12px!important;
    border-radius:14px!important;
    background:rgba(255,255,255,.50)!important;
    border:1px solid rgba(91,141,204,.18)!important;
    color:#35516f!important;
    font-size:12px!important;
  }
  #createScreen .teacherCredit div{white-space:normal!important;}
}
@media(min-width:1181px){
  #createScreen .teacherCredit{
    max-width:520px;
    margin-left:auto;
    padding-right:4px;
  }
}

.aiReviewPanel{border-color:rgba(34,211,238,.30)!important;background:linear-gradient(180deg,rgba(14,165,233,.11),rgba(255,255,255,.052))!important}
.aiReviewActions{display:flex;gap:8px;align-items:center;margin:8px 0 10px}.aiReviewActions button{width:100%;min-height:40px;padding:8px 10px;font-size:14px}.aiReviewBox{display:grid;gap:8px;max-height:220px;overflow:auto;padding-right:3px}.aiReviewSummary{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:4px}.aiPill{display:inline-flex;align-items:center;gap:4px;border-radius:999px;padding:5px 8px;font-size:11px;font-weight:1000;border:1px solid rgba(255,255,255,.14);background:rgba(255,255,255,.08)}.aiPill.ok{color:#dcfce7;background:rgba(22,163,74,.20)}.aiPill.warn{color:#fef3c7;background:rgba(245,158,11,.22)}.aiPill.bad{color:#fee2e2;background:rgba(239,68,68,.20)}.aiReviewItem{border:1px solid rgba(255,255,255,.10);background:rgba(255,255,255,.065);border-radius:13px;padding:8px}.aiReviewItem b{display:block;color:#0f2544;font-size:13px;margin-bottom:4px}.aiReviewItem .mini{font-size:11px}.aiStatus{display:inline-block;border-radius:999px;padding:3px 7px;margin-right:5px;font-size:11px;font-weight:1000}.aiStatus.ok{background:rgba(34,197,94,.20);color:#dcfce7}.aiStatus.warn{background:rgba(245,158,11,.20);color:#fef3c7}.aiStatus.bad{background:rgba(239,68,68,.22);color:#fee2e2}.aiStatus.muted{background:rgba(148,163,184,.18);color:#e2e8f0}.aiReviewWarning{border:1px solid rgba(245,158,11,.28);border-radius:12px;padding:7px 8px;background:rgba(245,158,11,.12);color:#78350f;font-weight:900}
#aiReviewBtn{background:linear-gradient(135deg,#22c55e,#0891b2 60%,#2563eb)!important;color:#fff!important;box-shadow:0 12px 24px rgba(14,165,233,.20)!important}
@media(max-width:1180px){.aiReviewBox{max-height:180px}.aiReviewActions{margin-top:6px}}


.winnerCrown{position:absolute;left:50%;top:-20px;transform:translateX(-50%) rotate(-7deg);z-index:6;font-size:34px;filter:drop-shadow(0 7px 9px rgba(0,0,0,.28)) drop-shadow(0 0 10px rgba(251,191,36,.52));animation:crownFloat 1.8s ease-in-out infinite alternate;pointer-events:none}
@keyframes crownFloat{from{transform:translateX(-50%) translateY(0) rotate(-7deg)}to{transform:translateX(-50%) translateY(-3px) rotate(6deg)}}
@media(max-width:760px){.winnerCrown{font-size:26px;top:-15px}}


/* v3.32 teacher mobile AI review visibility */
@media(max-width:1180px){
  #operateScreen .rightCol{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;overflow:visible!important;padding-right:0!important;}
  #operateScreen .rightCol .aiReviewPanel{order:99!important;width:100%!important;display:block!important;margin-bottom:10px!important;}
  #operateScreen .aiReviewBox{max-height:none!important;overflow:visible!important;}
}
@media(max-width:820px) and (orientation:landscape){
  #operateScreen .rightCol{grid-template-columns:1fr 1fr!important;align-items:start!important;}
  #operateScreen .rightCol .aiReviewPanel{grid-column:1 / -1!important;}
  #operateScreen .rightCol .aiReviewBox{max-height:160px!important;overflow:auto!important;}
}


/* v3.33 teacher-only BGM controls */
.teacherMusicPanel{display:flex;align-items:center;justify-content:flex-end;gap:10px;flex-wrap:wrap;margin-top:10px;padding:10px 12px;border-radius:16px;background:linear-gradient(135deg,rgba(15,23,42,.07),rgba(14,165,233,.09));border:1px solid rgba(125,211,252,.24);color:#eaf6ff}
#createScreen .teacherMusicPanel{justify-content:flex-end;color:#dbeafe;background:rgba(255,255,255,.055)}
.musicToggleBtn{width:auto!important;min-height:34px!important;padding:8px 12px!important;border-radius:999px!important;background:linear-gradient(135deg,#0ea5e9,#2563eb)!important;color:#fff!important;font-size:13px!important;font-weight:1000!important;box-shadow:0 12px 24px rgba(37,99,235,.18)!important}
.musicToggleBtn.off{background:linear-gradient(135deg,#64748b,#334155)!important;box-shadow:none!important;color:#e2e8f0!important}
.musicVolumeLabel{display:flex;align-items:center;gap:7px;font-size:12px;font-weight:900;color:#d9efff;white-space:nowrap}.musicVolumeLabel input{width:96px;accent-color:#38bdf8}.musicHint{font-size:11px;color:#b9d3ee;font-weight:800;min-width:72px;text-align:right}
@media(max-width:760px){.teacherMusicPanel{justify-content:center!important;gap:8px;margin-top:8px;padding:8px}.musicToggleBtn{font-size:12px!important;padding:7px 10px!important}.musicVolumeLabel input{width:82px}.musicHint{width:100%;text-align:center;font-size:10px}}
@media(max-width:820px) and (orientation:landscape){#operateScreen .teacherMusicPanel{margin-top:7px;padding:7px 8px;justify-content:flex-start!important}.musicHint{display:none}.musicVolumeLabel input{width:74px}}


/* ===== v3.34 student info compact grid + teacher setup 3-row patch ===== */
/* Student prep: title stays wide, the six small room/nickname fields form 3 columns x 2 rows on phones. */
#prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:1 / -1;}
#prepScreen .studentRoomInfoGrid .compactInfoCard .value{word-break:keep-all;}
@media (orientation:portrait) and (pointer:coarse), (max-width:760px){
  #prepScreen .studentRoomInfoGrid{
    grid-template-columns:repeat(3,minmax(0,1fr))!important;
    gap:8px!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard{
    grid-column:1 / -1!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard{
    min-height:76px!important;
    padding:10px 8px!important;
    display:flex!important;
    flex-direction:column!important;
    justify-content:center!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .label{
    font-size:clamp(10px,2.9vw,12px)!important;
    line-height:1.15!important;
    margin-bottom:6px!important;
    white-space:normal!important;
    letter-spacing:-.45px!important;
  }
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{
    font-size:clamp(18px,5vw,28px)!important;
    line-height:1.05!important;
    letter-spacing:-1.2px!important;
    white-space:nowrap!important;
    overflow:hidden!important;
    text-overflow:ellipsis!important;
  }
  #prepScreen .studentRoomInfoGrid .infoTitleCard .value{
    font-size:clamp(21px,6.2vw,34px)!important;
  }
}
@media (orientation:portrait) and (pointer:coarse) and (max-width:380px){
  #prepScreen .studentRoomInfoGrid{gap:6px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard{padding:8px 6px!important;min-height:68px!important;border-radius:14px!important;}
  #prepScreen .studentRoomInfoGrid .compactInfoCard .value{font-size:clamp(16px,4.7vw,22px)!important;}
}
/* Student landscape keeps the previous compact side panel to avoid squeezing the question editor. */
@media (orientation:landscape) and (pointer:coarse){
  #prepScreen .studentRoomInfoGrid{grid-template-columns:1fr!important;}
  #prepScreen .studentRoomInfoGrid .infoTitleCard{grid-column:auto!important;}
}
/* Teacher PC setup: compact numeric settings become 4 columns, so 10 fields render as 3 rows instead of 5. */
@media (min-width:821px){
  #createScreen .teacherSettingsGrid{
    grid-template-columns:repeat(4,minmax(0,1fr))!important;
    gap:10px!important;
  }
  #createScreen .teacherSettingsGrid .field input{
    min-height:40px!important;
    padding:8px 10px!important;
  }
}

</style>
</head>
<body>
<div class='topbar'>
  <div class='brand'>ReMap</div>
  <div id='roomTitleBar' class='roomTitle'>REMAP</div>
  <div id='statusBar' class='statusBadge'>[현재상황: 준비 중]</div>
</div>
<section id='createScreen'>
  <div class='createCard'>
    <h1>방 생성 / 설정</h1>
    <div class='mini'>처음 사용해도 바로 이해되도록 설정 화면만 먼저 보여줍니다. 방을 만들면 맵 중심 운영 화면으로 전환됩니다.</div>
    <div class='createGrid'>
      <div class='row full teacherRoomMetaRow'><div class='field'><label>방 제목</label><input id='room_title' value='REMAP 방제목'></div><div class='field'><label>방장</label><input id='teacher_owner' value='' placeholder='예: 교사1'></div></div>
      <div class='field'><label>게임 모드</label><input id='game_mode' type='hidden' value='solo'><div class='modeButtons'><button type='button' class='modeBtn active' data-mode='solo'>개인전</button><button type='button' class='modeBtn' data-mode='team'>팀전</button></div></div>
      <div id='teamCountWrap' class='field'><label>팀 수</label><input id='team_count' type='number' min='2' max='5' value='4'></div>
      <div class='field full'><label>맵 종류</label><input id='map_type' type='hidden' value='open'><div class='mapButtons'><button type='button' class='mapBtn active' data-map='open'>오픈 스퀘어<small>넓은 자유 이동형</small></button><button type='button' class='mapBtn' data-map='maze'>미로형 맵<small>벽을 피해 만나는 구조</small></button></div></div>
      <div class='helpBox full teamFixedBox'><b>팀전 색상 고정</b><div class='teamLegend'><span class='teamChip team-A'>A 파랑</span><span class='teamChip team-B'>B 빨강</span><span class='teamChip team-C'>C 노랑</span><span class='teamChip team-D'>D 초록</span><span class='teamChip team-E'>E 회색</span></div></div>
      <div class='row full teacherSettingsGrid'>
        <div class='field'><label>배틀 문제 수</label><input id='question_count' type='number' min='1' value='1'></div>
        <div class='field'><label>문제당 제한시간(초)</label><input id='question_time_limit' type='number' value='20'></div>
        <div class='field'><label>총 게임 시간(초)</label><input id='total_game_time' type='number' value='180'></div>
        <div class='field'><label>이동 속도</label><input id='player_speed' type='number' step='0.1' value='7.5'></div>
        <div class='field'><label>맵 너비</label><input id='map_width' type='number' value='1060'></div>
        <div class='field'><label>맵 높이</label><input id='map_height' type='number' value='612'></div>
        <div class='field'><label>승점</label><input id='score_win' type='number' value='3'></div>
        <div class='field'><label>무승부</label><input id='score_draw' type='number' value='1'></div>
        <div class='field'><label>패배</label><input id='score_lose' type='number' value='0'></div>
        <div class='field'><label>배경 이미지</label><input id='bgFile' type='file' accept='image/*'></div>
      </div>
    </div>
    <div class='createActions'><button id='createBtn'>방 생성</button><button id='clearBgPreBtn' type='button' class='ghost'>기본 배경 복원</button></div><div class='teacherMusicPanel' aria-label='교사용 음악 모드'><button type='button' class='musicToggleBtn'>🎵 음악모드 ON</button><label class='musicVolumeLabel'>음량 <input type='range' class='musicVolume' min='0' max='100' value='35'></label><span class='musicHint'>교사 화면에서만 재생</span></div><div class='teacherCredit'><div>만든이: 서울시교육청 교사 김철원</div><div>문의: churwon@sen.go.kr</div></div>
  </div>
</section>
<section id='operateScreen'>
  <div class='opHeader'>
    <div class='panel codeBox'><div class='codeTextBlock'><div class='mini'>학생 입장 코드</div><div id='codeValue' class='codeValue'>----</div><div id='studentJoinUrl' class='joinUrl'>학생 접속 주소 준비 중</div></div><div class='qrPanel'><img id='joinQr' class='joinQr' alt='학생 입장 QR 코드'><div class='qrCaption'>스마트폰 카메라로 스캔</div></div></div>
    <div class='panel summaryPanel'><div class='headerMeta'><div class='miniCard primary'><span class='mini'>상태</span><strong id='stateValue'>준비 중</strong></div><div class='miniCard compact'><span class='mini'>참가자</span><strong id='playerCountValue'>0명</strong></div><div class='miniCard primary'><span class='mini'>남은 시간</span><strong id='remainValue'>3:00</strong></div><div class='miniCard compact'><span class='mini'>팀별 인원</span><strong id='teamCountValue'>-</strong></div><div class='miniCard compact'><span class='mini'>미제출</span><strong id='unsubmittedValue'>0명</strong></div></div><div class='opButtons' id='opButtons'><button id='startBtn'>게임 시작</button><button id='endBtn' class='danger'>게임 종료</button><button id='ceremonyBtn' class='ceremonyBtn' type='button'>시상식 보기</button><button id='resetBtn' class='ghost'>다음 게임 준비</button><button id='exportQuestionsBtn' class='soft'>문제 엑셀 다운로드</button><button id='aiReviewBtn' class='soft' type='button'>AI 문제 검토</button><button id='newRoomBtn' class='soft'>새 방 설정</button></div><div class='teacherMusicPanel' aria-label='교사용 음악 모드'><button type='button' class='musicToggleBtn'>🎵 음악모드 ON</button><label class='musicVolumeLabel'>음량 <input type='range' class='musicVolume' min='0' max='100' value='35'></label><span class='musicHint'>교사 화면에서만 재생</span></div></div>
  </div>
  <div class='opMain'>
    <div class='col leftCol'>
      <div class='panel'><h3 class='sectionTitle'>방 설정 요약</h3><div id='roomBox'></div></div>
      <div class='panel'><h3 class='sectionTitle'>참가 / 제출 현황</h3><div id='participantBox'></div></div>
    </div>
    <div class='centerCol'>
      <div class='panel mapPanel'>
        <div class='mapHeader'><div class='mapTitleWrap'><h3>실시간 맵 보기</h3><div class='mini'>중앙 맵에서 학생 위치와 미로 벽을 바로 확인합니다.</div></div><button id='clearBgLiveBtn' class='ghost' style='width:auto'>기본 배경 복원</button></div>
        <div class='mapWrap'><div id='teacherMapZoomControls' class='teacherMapZoomControls' aria-label='교사용 맵 크기 조절'><button id='teacherMapZoomOutBtn' type='button'>−</button><span id='teacherMapZoomValue'>75%</span><button id='teacherMapZoomInBtn' type='button'>＋</button></div><canvas id='teacherCanvas' width='1060' height='612'></canvas></div>
      </div>
    </div>
    <div class='col rightCol'>
      <div class='panel'><h3 class='sectionTitle'>개인 순위</h3><div id='rankingBox'></div></div>
      <div class='panel'><h3 class='sectionTitle'>팀 순위</h3><div id='teamRankingBox'></div></div>
      <div class='panel'><h3 class='sectionTitle'>배틀 진행률</h3><div id='battleBox'></div></div>
      <div class='panel aiReviewPanel'><h3 class='sectionTitle'>AI 문제 검토</h3><div class='mini'>AI는 최종 판정이 아니라 교사용 참고 의견만 표시합니다.</div><div id='aiReviewBox' class='aiReviewBox'><div class='mini'>[AI 문제 검토] 버튼을 누르면 학생 제출 문제를 검토합니다.</div></div></div><div class='panel'><h3 class='sectionTitle'>로그</h3><div id='logBox'></div></div>
    </div>
  </div>
</section>
<div id='qrModal' class='qrModal' aria-hidden='true'>
  <div class='qrModalCard' role='dialog' aria-modal='true' aria-labelledby='qrModalTitle'>
    <button id='qrModalClose' class='qrModalClose' type='button' aria-label='QR 확대 닫기'>×</button>
    <div id='qrModalTitle' class='qrModalTitle'>학생 입장 QR</div>
    <div id='qrModalCode' class='qrModalCode'>----</div>
    <img id='qrModalImg' class='qrModalImg' alt='확대된 학생 입장 QR 코드'>
    <div id='qrModalUrl' class='qrModalUrl'>학생 접속 주소 준비 중</div>
    <div class='qrModalHint'>학생 스마트폰 카메라로 이 QR을 스캔하세요.</div>
  </div>
</div>
<script>
let bgDataUrl=null,currentState=null,editingRoom=false,lastGameEndPayload=null;
const createScreen=document.getElementById('createScreen'),operateScreen=document.getElementById('operateScreen');
const statusBar=document.getElementById('statusBar'),roomTitleBar=document.getElementById('roomTitleBar'),roomBox=document.getElementById('roomBox'),participantBox=document.getElementById('participantBox'),rankingBox=document.getElementById('rankingBox'),teamRankingBox=document.getElementById('teamRankingBox'),battleBox=document.getElementById('battleBox'),logBox=document.getElementById('logBox'),aiReviewBox=document.getElementById('aiReviewBox');
const teacherMusic={
  enabled: localStorage.getItem('remapTeacherMusicMode')!=='off',
  volume: Math.max(0,Math.min(1,Number(localStorage.getItem('remapTeacherMusicVolume')||'0.35'))),
  current:null,
  lastPhase:null,
  lastFinishedRoom:null
};
const teacherMusicTracks={
  wait:new Audio('/static/rm_wait_sound.mp3'),
  main:new Audio('/static/rm_main_sound.mp3'),
  win:new Audio('/static/rm_win_sound.mp3')
};
teacherMusicTracks.wait.loop=true;teacherMusicTracks.main.loop=true;teacherMusicTracks.win.loop=false;
Object.values(teacherMusicTracks).forEach(a=>{a.preload='auto';a.volume=teacherMusic.volume;});
function updateMusicControls(message){
  document.querySelectorAll('.musicToggleBtn').forEach(btn=>{btn.textContent=teacherMusic.enabled?'🎵 음악모드 ON':'🔇 음악모드 OFF';btn.classList.toggle('off',!teacherMusic.enabled);});
  document.querySelectorAll('.musicVolume').forEach(sl=>{sl.value=String(Math.round(teacherMusic.volume*100));});
  if(message){document.querySelectorAll('.musicHint').forEach(el=>{el.textContent=message;});}
  else{document.querySelectorAll('.musicHint').forEach(el=>{el.textContent='교사 화면에서만 재생';});}
}
function setTeacherMusicVolume(value){teacherMusic.volume=Math.max(0,Math.min(1,Number(value)/100));localStorage.setItem('remapTeacherMusicVolume',String(teacherMusic.volume));Object.values(teacherMusicTracks).forEach(a=>a.volume=teacherMusic.volume);updateMusicControls();}
function stopTeacherMusic(){Object.values(teacherMusicTracks).forEach(a=>{try{a.pause();a.currentTime=0;}catch(e){}});teacherMusic.current=null;}
function playTeacherTrack(name,{loop=true,restart=false}={}){
  if(!teacherMusic.enabled)return;
  const audio=teacherMusicTracks[name];if(!audio)return;
  Object.entries(teacherMusicTracks).forEach(([k,a])=>{if(k!==name){try{a.pause();a.currentTime=0;}catch(e){}}});
  audio.loop=loop;audio.volume=teacherMusic.volume;
  if(restart||teacherMusic.current!==name){try{audio.currentTime=0;}catch(e){}}
  teacherMusic.current=name;
  const promise=audio.play();
  if(promise&&typeof promise.catch==='function'){
    promise.catch(()=>{updateMusicControls('음악 재생 대기: 음악모드 버튼을 눌러주세요');});
  }
}
function syncTeacherMusic(status, force=false){
  if(!teacherMusic.enabled)return;
  if(status==='running'||status==='countdown'){
    teacherMusic.lastPhase='main';
    playTeacherTrack('main',{loop:true,restart:force||teacherMusic.current!=='main'});
  }else if(status==='finished'){
    const roomKey=(currentState&&currentState.room&&currentState.room.code)||teacherRoomCode||'room';
    if(force||teacherMusic.lastPhase!=='finished'||teacherMusic.lastFinishedRoom!==roomKey){
      teacherMusic.lastPhase='finished';teacherMusic.lastFinishedRoom=roomKey;
      playTeacherTrack('win',{loop:false,restart:true});
    }
  }else if(status==='lobby'||status==='idle'||status==='ready'||status==='waiting'||status==='countdown_ready'){
    teacherMusic.lastPhase='wait';
    playTeacherTrack('wait',{loop:true,restart:force||teacherMusic.current!=='wait'});
  }
}
function setTeacherMusicEnabled(enabled){teacherMusic.enabled=!!enabled;localStorage.setItem('remapTeacherMusicMode',teacherMusic.enabled?'on':'off');if(!teacherMusic.enabled){stopTeacherMusic();}else{syncTeacherMusic((currentState&&currentState.game_status)||'lobby',true);}updateMusicControls();}
document.querySelectorAll('.musicToggleBtn').forEach(btn=>btn.addEventListener('click',()=>setTeacherMusicEnabled(!teacherMusic.enabled)));
document.querySelectorAll('.musicVolume').forEach(sl=>sl.addEventListener('input',e=>setTeacherMusicVolume(e.target.value)));
updateMusicControls();

const canvas=document.getElementById('teacherCanvas'),ctx=canvas.getContext('2d');let bgImage=null;
const teacherMapZoomOutBtn=document.getElementById('teacherMapZoomOutBtn'), teacherMapZoomInBtn=document.getElementById('teacherMapZoomInBtn'), teacherMapZoomValue=document.getElementById('teacherMapZoomValue');
let teacherMapScale=(window.matchMedia && window.matchMedia('(max-width: 820px) and (orientation: landscape) and (pointer: coarse)').matches)?0.75:1;
function applyTeacherMapScale(){document.documentElement.style.setProperty('--teacher-map-scale', String(teacherMapScale));if(teacherMapZoomValue)teacherMapZoomValue.textContent=Math.round(teacherMapScale*100)+'%';}
function changeTeacherMapScale(delta){teacherMapScale=Math.max(0.55,Math.min(1.15,Number((teacherMapScale+delta).toFixed(2))));applyTeacherMapScale();}
if(teacherMapZoomOutBtn)teacherMapZoomOutBtn.onclick=(ev)=>{ev.stopPropagation();changeTeacherMapScale(-0.05);};
if(teacherMapZoomInBtn)teacherMapZoomInBtn.onclick=(ev)=>{ev.stopPropagation();changeTeacherMapScale(0.05);};
applyTeacherMapScale();
const ids=['room_title','teacher_owner','game_mode','team_count','map_type','question_count','question_time_limit','total_game_time','map_width','map_height','player_speed','score_win','score_draw','score_lose'];
const gameModeSelect=document.getElementById('game_mode');const mapTypeSelect=document.getElementById('map_type');const teamCountWrap=document.getElementById('teamCountWrap');
const modeButtons=[...document.querySelectorAll('.modeBtn')];const mapButtons=[...document.querySelectorAll('.mapBtn')];
function syncModeUI(){const isTeam=gameModeSelect.value==='team';teamCountWrap.style.display=isTeam?'block':'none';modeButtons.forEach(btn=>btn.classList.toggle('active',btn.dataset.mode===gameModeSelect.value));}
function syncMapUI(){mapButtons.forEach(btn=>btn.classList.toggle('active',btn.dataset.map===mapTypeSelect.value));}
modeButtons.forEach(btn=>btn.onclick=()=>{gameModeSelect.value=btn.dataset.mode;syncModeUI();});mapButtons.forEach(btn=>btn.onclick=()=>{mapTypeSelect.value=btn.dataset.map;syncMapUI();});syncModeUI();syncMapUI();
function statusText(s){return s==='countdown'?'시작 대기':s==='running'?'게임 진행 중':s==='finished'?'게임 종료':s==='lobby'?'대기 중':'준비 중'}
function fmt(sec){sec=Number(sec||0);return `${Math.floor(sec/60)}:${String(sec%60).padStart(2,'0')}`}
function escapeHtml(value){return String(value??'').replace(/[&<>'"]/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[ch]))}
let lastAccessInfoCode='';
const qrModal=document.getElementById('qrModal');
const qrModalImg=document.getElementById('qrModalImg');
const qrModalCode=document.getElementById('qrModalCode');
const qrModalUrl=document.getElementById('qrModalUrl');
const qrModalClose=document.getElementById('qrModalClose');
function openQrModal(){
  const qrEl=document.getElementById('joinQr');
  const code=(document.getElementById('codeValue')?.textContent||'').trim();
  const urlText=(document.getElementById('studentJoinUrl')?.textContent||'').trim();
  if(!qrEl||!qrEl.getAttribute('src')||!qrModal)return;
  qrModalImg.src=qrEl.getAttribute('src');
  qrModalImg.alt=`${code||''} 확대 QR 코드`;
  qrModalCode.textContent=code&&code!=='----'?code:'----';
  qrModalUrl.textContent=urlText||'학생 접속 주소 준비 중';
  qrModal.classList.add('show');
  qrModal.setAttribute('aria-hidden','false');
}
function closeQrModal(){
  if(!qrModal)return;
  qrModal.classList.remove('show');
  qrModal.setAttribute('aria-hidden','true');
}
document.getElementById('joinQr')?.addEventListener('click',openQrModal);
qrModalClose?.addEventListener('click',closeQrModal);
qrModal?.addEventListener('click',e=>{if(e.target===qrModal)closeQrModal();});
window.addEventListener('keydown',e=>{if(e.key==='Escape')closeQrModal();});
async function refreshAccessInfo(code){
  code=(code||'').trim().toUpperCase();
  const urlEl=document.getElementById('studentJoinUrl');
  const qrEl=document.getElementById('joinQr');
  if(!urlEl||!qrEl)return;
  if(!code){
    lastAccessInfoCode='';
    urlEl.textContent='학생 접속 주소 준비 중';
    qrEl.removeAttribute('src');
    qrEl.alt='학생 입장 QR 코드';
    return;
  }
  if(lastAccessInfoCode===code)return;
  lastAccessInfoCode=code;
  qrEl.alt=`${code} 학생 입장 QR 코드`;
  qrEl.src=`/api/room/qr.svg?code=${encodeURIComponent(code)}&v=${Date.now()}`;
  if(qrModal&&qrModal.classList.contains('show')){qrModalImg.src=qrEl.src;qrModalCode.textContent=code;}
  urlEl.textContent='학생 접속 주소 불러오는 중...';
  try{
    const res=await fetch(`/api/access_info?code=${encodeURIComponent(code)}`,{cache:'no-store'});
    const data=await res.json();
    urlEl.textContent=data.student_url||`${location.origin}/?code=${code}`;
    if(qrModal&&qrModal.classList.contains('show'))qrModalUrl.textContent=urlEl.textContent;
  }catch(e){
    urlEl.textContent=`${location.origin}/?code=${code}`;
    if(qrModal&&qrModal.classList.contains('show'))qrModalUrl.textContent=urlEl.textContent;
  }
}
function applyStatusStyle(status){statusBar.classList.remove('status-running','status-finished','status-waiting');if(status==='running'){statusBar.classList.add('status-running');}else if(status==='finished'){statusBar.classList.add('status-finished');}else if(status==='lobby'||status==='countdown'){statusBar.classList.add('status-waiting');}}
function updateCeremonyButton(status){const btn=document.getElementById('ceremonyBtn');const box=document.getElementById('opButtons');const show=status==='finished';if(btn){btn.classList.toggle('show',show);}if(box){box.classList.toggle('hasCeremony',show);}}
function showCreate(){editingRoom=true;createScreen.style.display='flex';operateScreen.style.display='none';roomTitleBar.textContent='REMAP';statusBar.textContent='[현재상황: 설정 중]';applyStatusStyle('idle');updateCeremonyButton('idle')}
function showOperate(){editingRoom=false;createScreen.style.display='none';operateScreen.style.display='grid'}

function aiStatusClass(status){if(status==='통과')return 'ok';if(status==='오류 가능성')return 'bad';if(status==='확인 필요'||status==='표현 모호')return 'warn';return 'muted';}
function renderAiReviews(items,meta){
  if(!aiReviewBox)return;
  items=Array.isArray(items)?items:[];
  if(!items.length){aiReviewBox.innerHTML='<div class="mini">[AI 문제 검토] 버튼을 누르면 학생 제출 문제를 검토합니다.</div>';return;}
  const counts={};items.forEach(it=>{counts[it.status||'검토 불가']=(counts[it.status||'검토 불가']||0)+1;});
  const pillHtml=['통과','확인 필요','오류 가능성','표현 모호','검토 불가'].filter(k=>counts[k]).map(k=>`<span class="aiPill ${aiStatusClass(k)}">${escapeHtml(k)} ${counts[k]}</span>`).join('');
  const reviewedAt=meta&&meta.reviewed_at?`<div class="mini">검토 시각: ${escapeHtml(meta.reviewed_at)} · AI는 참고 의견이며 최종 판단은 교사가 확인합니다.</div>`:'<div class="mini">AI는 참고 의견이며 최종 판단은 교사가 확인합니다.</div>';
  const warningHtml=meta&&meta.warning?`<div class="mini aiReviewWarning">${escapeHtml(meta.warning)}</div>`:'';
  const rows=items.slice().sort((a,b)=>String(a.nickname||'').localeCompare(String(b.nickname||''))||Number(a.question_no||0)-Number(b.question_no||0)).map(it=>`<div class="aiReviewItem"><b>${escapeHtml(it.nickname||'-')} ${Number(it.question_no||0)}번 <span class="aiStatus ${aiStatusClass(it.status)}">${escapeHtml(it.status||'검토 불가')}</span></b><div class="mini">${escapeHtml(it.summary||'')}</div>${it.suggested_answer?`<div class="mini">추천/확인 답: ${escapeHtml(it.suggested_answer)}</div>`:''}</div>`).join('');
  aiReviewBox.innerHTML=`<div class="aiReviewSummary">${pillHtml}</div>${reviewedAt}${warningHtml}${rows}`;
}
function updateScreen(msg){const hasRoom=!!(msg.room&&msg.room.code);if(!hasRoom){createScreen.style.display='flex';operateScreen.style.display='none';roomTitleBar.textContent='REMAP';statusBar.textContent='[현재상황: 설정 전]';applyStatusStyle('idle');updateCeremonyButton('idle');return;}if(editingRoom&&!teacherRoomCode){return;}showOperate();}
document.getElementById('bgFile').onchange=(e)=>{const file=e.target.files[0];if(!file)return;const reader=new FileReader();reader.onload=()=>{bgDataUrl=reader.result;};reader.readAsDataURL(file);};
document.getElementById('clearBgPreBtn').onclick=()=>{bgDataUrl=null;document.getElementById('bgFile').value='';};
let teacherRoomCode=(new URLSearchParams(location.search).get('room')||'').trim().toUpperCase().slice(0,4);
let teacherWs=null;
function teacherApi(path){const u=new URL(path,location.href);if(teacherRoomCode)u.searchParams.set('room',teacherRoomCode);return u.href;}
function setTeacherRoomCode(code){teacherRoomCode=String(code||'').trim().toUpperCase().slice(0,4);const u=new URL(location.href);if(teacherRoomCode){u.searchParams.set('room',teacherRoomCode);}else{u.searchParams.delete('room');}history.replaceState(null,'',u.href);}
function connectTeacherSocket(){
  if(teacherWs){try{teacherWs.close();}catch(e){}}
  const wsUrl=new URL('/ws/teacher', location.href);
  wsUrl.protocol=location.protocol==='https:'?'wss:':'ws:';
  if(teacherRoomCode)wsUrl.searchParams.set('room',teacherRoomCode);
  const ws=new WebSocket(wsUrl.href);
  teacherWs=ws;
  ws.onmessage=(ev)=>{
  const msg=JSON.parse(ev.data);
  if(msg.type==='game_end'){lastGameEndPayload=msg;updateCeremonyButton('finished');return;}
  if(msg.type!=='state')return;
  currentState=msg;
  updateScreen(msg);
  const room=msg.room||{};
  const settings=msg.settings||{};
  const players=msg.players||[];
  const participants=msg.participants||[];
  const rankings=msg.rankings||[];
  const teamRankings=msg.team_rankings||[];
  const battles=msg.battles||[];
  const logs=msg.logs||[];
  const teamMode=room.game_mode==='team';
  statusBar.textContent=`[현재상황: ${statusText(msg.game_status)}]`;applyStatusStyle(msg.game_status);updateCeremonyButton(msg.game_status);syncTeacherMusic(msg.game_status);
  roomTitleBar.textContent='REMAP';
  document.getElementById('codeValue').textContent=room.code||'----';
  refreshAccessInfo(room.code||'');
  document.getElementById('stateValue').textContent=statusText(msg.game_status);
  document.getElementById('playerCountValue').textContent=(players.length||0)+'명';
  document.getElementById('remainValue').textContent=fmt(msg.remaining_time);
  document.getElementById('teamCountValue').textContent=(msg.team_counts&&msg.team_counts.length?msg.team_counts.map(t=>`${t.team}:${Number(t.count||0)}`).join(' / '):'-');
  document.getElementById('unsubmittedValue').textContent=(Number(msg.unsubmitted_count||0))+'명';
  roomBox.innerHTML=`<div class='item'><span>방 제목</span><strong>${escapeHtml(room.title||'-')}</strong></div><div class='item'><span>방장</span><strong>${escapeHtml(room.owner||settings.teacher_owner||'-')}</strong></div><div class='item'><span>방 코드</span><strong>${escapeHtml(room.code||'-')}</strong></div><div class='item'><span>게임 모드</span><strong>${teamMode?'팀전':'개인전'}</strong></div>${teamMode?`<div class='item'><span>팀 수</span><strong>${Number(room.team_count||0)||'-'}</strong></div>`:''}<div class='item'><span>맵 종류</span><strong>${escapeHtml(room.map_label||'-')}</strong></div><div class='item'><span>배틀 문제 수</span><strong>${Number(settings.question_count||0)}</strong></div><div class='item'><span>제한시간</span><strong>${Number(settings.question_time_limit||0)}초</strong></div><div class='item'><span>점수</span><strong>${Number(settings.score_win||0)}/${Number(settings.score_draw||0)}/${Number(settings.score_lose||0)}</strong></div>`;
  participantBox.innerHTML=participants.length?participants.map(p=>`<div class='item'><span>${escapeHtml(p.nickname)}${p.team?` (${escapeHtml(p.team)})`:''}</span><strong class='${p.submitted?'submitDone':'submitWait'}'>${p.submitted?'제출 완료':'미제출'}</strong></div>`).join(''):'<div class="mini">아직 참가자가 없습니다.</div>';
  rankingBox.innerHTML=rankings.length?rankings.map(r=>`<div class='item'><span>${Number(r.rank)||'-'}. ${escapeHtml(r.nickname)}${teamMode&&r.team?` (${escapeHtml(r.team)})`:''}</span><strong>${Number(r.score||0)}</strong></div>`).join(''):'<div class="mini">순위 없음</div>';
  teamRankingBox.innerHTML=teamMode?(teamRankings.length?teamRankings.map(r=>`<div class='item'><span>${Number(r.rank)||'-'}. ${escapeHtml(r.team)}팀</span><strong>${Number(r.score||0)}</strong></div>`).join(''):'<div class="mini">팀 순위 없음</div>'):'<div class="mini">개인전 모드</div>';
  battleBox.innerHTML=battles.length?battles.map(b=>`<div class='item'><span>${(b.players||[]).map(escapeHtml).join(' vs ')}</span><strong>${Number(b.progress||0)}/${Number(b.total||0)}</strong></div>`).join(''):'<div class="mini">진행 중인 배틀 없음</div>';
  logBox.innerHTML=logs.length?logs.slice().reverse().map(l=>`<div class='item'><span>${escapeHtml(l.time)}</span><span>${escapeHtml(l.message)}</span></div>`).join(''):'<div class="mini">로그 없음</div>';
  renderAiReviews(msg.ai_reviews||[], msg.ai_review_meta||{});
  ids.forEach(id=>{const el=document.getElementById(id);if(el&&settings[id]!==undefined&&!editingRoom)el.value=settings[id];});
  syncModeUI();
  syncMapUI();
  resizeCanvas(settings);
  renderMap(msg);
  if(room.code && room.code!==teacherRoomCode){setTeacherRoomCode(room.code);}
};
  ws.onclose=()=>{if(teacherWs===ws){teacherWs=null;}};
}
document.getElementById('createBtn').onclick=async()=>{const btn=document.getElementById('createBtn');btn.disabled=true;btn.textContent='방 생성 중...';const payload={};ids.forEach(id=>{const el=document.getElementById(id);payload[id]=(id==='room_title'||id==='teacher_owner'||id==='game_mode'||id==='map_type')?el.value:Number(el.value);});payload.background_data_url=bgDataUrl||null;editingRoom=false;const res=await fetch('/api/teacher/create_room',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});if(res.ok){const data=await res.json();if(data.room_code){setTeacherRoomCode(data.room_code);connectTeacherSocket();}showOperate();syncTeacherMusic('lobby',true);}else{editingRoom=true;showCreate();}btn.disabled=false;btn.textContent='방 생성';};
document.getElementById('startBtn').onclick=()=>{syncTeacherMusic('running',true);return fetch(teacherApi('/api/teacher/start'),{method:'POST'});};
document.getElementById('endBtn').onclick=()=>{syncTeacherMusic('finished',true);return fetch(teacherApi('/api/teacher/end'),{method:'POST'});};
document.getElementById('resetBtn').onclick=()=>{syncTeacherMusic('lobby',true);return fetch(teacherApi('/api/teacher/reset'),{method:'POST'});};
document.getElementById('exportQuestionsBtn').onclick=()=>{window.location.href=teacherApi('/api/teacher/export_questions.xlsx');};
document.getElementById('aiReviewBtn').onclick=async()=>{const btn=document.getElementById('aiReviewBtn');btn.disabled=true;const old=btn.textContent;btn.textContent='AI 검토 중...';if(aiReviewBox){aiReviewBox.innerHTML='<div class="mini">AI가 학생 문제를 보수적으로 검토하는 중입니다. 무료 API 환경에서는 잠시 걸릴 수 있습니다.</div>';}try{const res=await fetch(teacherApi('/api/teacher/ai_review'),{method:'POST'});const data=await res.json().catch(()=>({ok:false,message:'응답 해석 실패',items:[]}));if(data.ok){renderAiReviews(data.items||[],data.meta||{});}else if(aiReviewBox){aiReviewBox.innerHTML=`<div class="mini">${escapeHtml(data.message||'AI 검토를 실행하지 못했습니다.')}</div>`;}}catch(e){if(aiReviewBox){aiReviewBox.innerHTML=`<div class="mini">AI 검토 연결 오류: ${escapeHtml(e.message||e)}</div>`;}}finally{btn.disabled=false;btn.textContent=old;}};
document.getElementById('newRoomBtn').onclick=async()=>{const btn=document.getElementById('newRoomBtn');btn.disabled=true;try{if(currentState&&['countdown','running'].includes(currentState.game_status)){await fetch(teacherApi('/api/teacher/end'),{method:'POST'});}else{await fetch(teacherApi('/api/teacher/new_room_setup'),{method:'POST'});}}catch(e){console.error(e);}finally{stopTeacherMusic();setTeacherRoomCode('');connectTeacherSocket();btn.disabled=false;showCreate();}};
document.getElementById('clearBgLiveBtn').onclick=()=>fetch(teacherApi('/api/teacher/clear_background'),{method:'POST'});

connectTeacherSocket();
function resizeCanvas(settings){const w=Number(settings.map_width)||1060;const h=Number(settings.map_height)||612;if(canvas.width!==w)canvas.width=w;if(canvas.height!==h)canvas.height=h;if(settings.background_data_url){if(!bgImage||bgImage.src!==settings.background_data_url){const img=new Image();img.onload=()=>currentState&&renderMap(currentState);img.src=settings.background_data_url;bgImage=img;}}else{bgImage=null;}}
function buildTeacherEndPayload(){const msg=currentState||{};const players=(msg.players||[]).slice().sort((a,b)=>(Number(b.score||0)-Number(a.score||0))||(Number(b.correct_count||0)-Number(a.correct_count||0))||String(a.nickname||'').localeCompare(String(b.nickname||'')));const best=players.slice().sort((a,b)=>(Number(b.correct_count||0)-Number(a.correct_count||0))||(Number(a.answer_count||0)-Number(b.answer_count||0))||String(a.nickname||'').localeCompare(String(b.nickname||'')))[0]||null;const most=players.slice().sort((a,b)=>(Number(b.battles_played||0)-Number(a.battles_played||0))||(Number(b.correct_count||0)-Number(a.correct_count||0))||String(a.nickname||'').localeCompare(String(b.nickname||'')))[0]||null;return {rankings:msg.rankings||players.map((p,i)=>({...p,rank:i+1})),team_rankings:msg.team_rankings||[],logs:msg.student_logs||msg.logs||[],player_stats:players.map(p=>({nickname:p.nickname,team:p.team,score:p.score,correct_count:p.correct_count,answer_count:p.answer_count,battles_played:p.battles_played,color:p.color})),best_correct:best?{nickname:best.nickname,correct_count:best.correct_count}:null,most_battles:most?{nickname:most.nickname,battles_played:most.battles_played}:null,winner_team:(msg.team_rankings&&msg.team_rankings[0])||null};}
let teacherCeremonyWindow=null;
function openTeacherCeremony(){
  const url=teacherApi('/teacher/ceremony');
  try{
    if(teacherCeremonyWindow&&!teacherCeremonyWindow.closed){
      teacherCeremonyWindow.location.href=url;
      teacherCeremonyWindow.focus();
      return;
    }
  }catch(e){}
  teacherCeremonyWindow=window.open(url,'remap_teacher_ceremony','width=1280,height=900');
  if(teacherCeremonyWindow){try{teacherCeremonyWindow.focus();}catch(e){}}
}
const ceremonyBtnEl=document.getElementById('ceremonyBtn');
if(ceremonyBtnEl){ceremonyBtnEl.onclick=openTeacherCeremony;}
function drawWalls(walls){(walls||[]).forEach(w=>{const radius=Math.min(14,Math.min(w.w,w.h)*0.22);const g=ctx.createLinearGradient(w.x,w.y,w.x+w.w,w.y+w.h);g.addColorStop(0,'#7fb6ef');g.addColorStop(.55,'#5d8fda');g.addColorStop(1,'#4d78c6');ctx.save();ctx.shadowColor='rgba(59,130,246,0.18)';ctx.shadowBlur=8;ctx.shadowOffsetY=3;ctx.fillStyle=g;roundRect(w.x,w.y,w.w,w.h,radius,true,false);ctx.restore();ctx.strokeStyle='rgba(255,255,255,.42)';ctx.lineWidth=2;roundRect(w.x,w.y,w.w,w.h,radius,false,true);ctx.fillStyle='rgba(255,255,255,.16)';roundRect(w.x+4,w.y+4,Math.max(8,w.w-8),Math.max(6,Math.min(w.h*0.24,16)),Math.max(4,radius*0.5),true,false);});ctx.lineWidth=1;}
function renderMap(msg){ctx.setTransform(1,0,0,1,0,0);ctx.clearRect(0,0,canvas.width,canvas.height);const hasMaze=(msg.map_walls||[]).length>0;if(bgImage&&bgImage.complete){ctx.drawImage(bgImage,0,0,canvas.width,canvas.height)}else{const bg=ctx.createLinearGradient(0,0,canvas.width,canvas.height);bg.addColorStop(0,'#eaf2ff');bg.addColorStop(.55,'#dbeafe');bg.addColorStop(1,'#cfe2ff');ctx.fillStyle=bg;ctx.fillRect(0,0,canvas.width,canvas.height);ctx.fillStyle='rgba(255,255,255,.30)';ctx.fillRect(14,14,canvas.width-28,canvas.height-28)}ctx.strokeStyle=hasMaze?'rgba(37,99,235,0.08)':'rgba(37,99,235,0.09)';for(let x=0;x<canvas.width;x+=50){ctx.beginPath();ctx.moveTo(x,0);ctx.lineTo(x,canvas.height);ctx.stroke()}for(let y=0;y<canvas.height;y+=50){ctx.beginPath();ctx.moveTo(0,y);ctx.lineTo(canvas.width,y);ctx.stroke()}drawWalls(msg.map_walls||[]);(msg.players||[]).forEach(p=>{try{drawPlayer(p)}catch(e){console.warn('teacher drawPlayer failed',e,p);}})}
function hexToRgb(hex){const clean=(hex||'#60a5fa').replace('#','');const normalized=clean.length===3?clean.split('').map(c=>c+c).join(''):clean;const n=parseInt(normalized,16);return {r:(n>>16)&255,g:(n>>8)&255,b:n&255};}function darkenColor(hex,factor=0.22){const {r,g,b}=hexToRgb(hex);return `rgb(${Math.max(0,Math.floor(r*(1-factor)))},${Math.max(0,Math.floor(g*(1-factor)))},${Math.max(0,Math.floor(b*(1-factor)))})`;}function lightenColor(hex,factor=0.18){const {r,g,b}=hexToRgb(hex);return `rgb(${Math.min(255,Math.floor(r+(255-r)*factor))},${Math.min(255,Math.floor(g+(255-g)*factor))},${Math.min(255,Math.floor(b+(255-b)*factor))})`;}function alphaColor(hex,alpha){const {r,g,b}=hexToRgb(hex);return `rgba(${r},${g},${b},${alpha})`;}
function normalizeCharacterColor(color){const raw=String(color||'').trim();const cleaned=raw.replace(/[^0-9a-fA-F]/g,'');if(cleaned.length===3){return cleaned.split('').map(ch=>ch+ch).join('').toLowerCase();}if(cleaned.length>=6){return cleaned.slice(0,6).toLowerCase();}return '60a5fa';}
function buildCharacterSvgUrl(color){return `/character/${normalizeCharacterColor(color)}.svg`; }
const playerMascotCache={};
function getPlayerMascot(color){const key=normalizeCharacterColor(color);if(!playerMascotCache[key]){const img=new Image();img.onload=()=>{if(currentState)renderMap(currentState);};img.onerror=()=>{playerMascotCache[key]=null;};img.src=buildCharacterSvgUrl(key);playerMascotCache[key]=img;}return playerMascotCache[key];}
function drawPlayer(p){const size=30;const bodyColor=p.color||'#60a5fa';const mascot=getPlayerMascot(bodyColor);ctx.save();if(p.state==='battling'){ctx.globalAlpha=0.45}ctx.translate(p.x,p.y);ctx.shadowColor='rgba(15,23,42,0.18)';ctx.shadowBlur=4;ctx.shadowOffsetY=1;if(mascot&&mascot.complete&&mascot.naturalWidth>0){ctx.drawImage(mascot,-size/2,-size/2,size,size);}else{ctx.fillStyle=bodyColor;ctx.beginPath();ctx.arc(0,0,size/2.4,0,Math.PI*2);ctx.fill();}ctx.shadowColor='transparent';if(p.state==='battling'){ctx.beginPath();ctx.arc(0,0,size/2+5.8,0,Math.PI*2);ctx.strokeStyle='rgba(239,68,68,0.85)';ctx.lineWidth=2.2;ctx.stroke();}ctx.restore();ctx.fillStyle='#173b7a';ctx.font='12px Arial';ctx.textAlign='center';const teamLabel=(currentState&&currentState.room&&currentState.room.game_mode==='team'&&p.team)?` [${p.team}]`:'';ctx.fillText(`${p.nickname}${teamLabel}`,p.x,p.y-25)}
function roundRect(x,y,w,h,r,fill,stroke){ctx.beginPath();ctx.moveTo(x+r,y);ctx.arcTo(x+w,y,x+w,y+h,r);ctx.arcTo(x+w,y+h,x,y+h,r);ctx.arcTo(x,y+h,x,y,r);ctx.arcTo(x,y,x+w,y,r);ctx.closePath();if(fill)ctx.fill();if(stroke)ctx.stroke()}
showCreate();
</script>
</body>
</html>
"""
STUDENT_HTML = STUDENT_HTML.replace('__CHAR_COLORS__', json.dumps(CHAR_COLORS))
TEACHER_HTML = TEACHER_HTML
